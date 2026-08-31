import openpyxl, re, json, os

XLSX = os.path.join(os.path.dirname(__file__), 'test_files_by_category_20260723.xlsx')

# column map (1-indexed) on 'Test Files' sheet
C_PATH=3; C_FILE=4; C_PERSON=5; C_PR=6; C_DEVREL=7; C_XPU=8
C_TEAM=12; C_COMMPR=15; C_ASSIGNEE=16; C_STATUS=17

PULL_RE=re.compile(r'pytorch/pytorch/pull/(\d+)')
HASH_RE=re.compile(r'#(\d+)')

def parse_prs(v):
    """Return list of pytorch/pytorch PR numbers found in a cell (string)."""
    if v is None: return []
    s=str(v)
    nums=[]
    for m in PULL_RE.finditer(s):
        nums.append(m.group(1))
    # bare #123 references (only when no full URL captured them)
    if not nums:
        for m in HASH_RE.finditer(s):
            nums.append(m.group(1))
    # dedupe preserving order
    seen=set(); out=[]
    for n in nums:
        if n not in seen: seen.add(n); out.append(n)
    return out

def norm_xpu(v):
    if v is True or v is False: return v
    if v is None: return None
    s=str(v).strip().lower()
    if s in ('true','yes','1'): return True
    if s in ('false','no','0'): return False
    return None

wb=openpyxl.load_workbook(XLSX, data_only=True)
ws=wb['Test Files']

rows=[]; all_prs=set()
for ri in range(2, ws.max_row+1):
    path=ws.cell(ri,C_PATH).value
    team=ws.cell(ri,C_TEAM).value
    if not path or team in (None,''):   # "owned" = team (Owner col L) set
        continue
    prs=parse_prs(ws.cell(ri,C_PR).value)
    comm_prs=parse_prs(ws.cell(ri,C_COMMPR).value)
    all_prs.update(prs); all_prs.update(comm_prs)
    rows.append({
        'path':path,
        'file':ws.cell(ri,C_FILE).value or os.path.basename(str(path)),
        'team':str(team).strip(),
        'person':ws.cell(ri,C_PERSON).value,
        'assignee':ws.cell(ri,C_ASSIGNEE).value,
        'device_relevance':ws.cell(ri,C_DEVREL).value,
        'xpu':norm_xpu(ws.cell(ri,C_XPU).value),
        'status':ws.cell(ri,C_STATUS).value,
        'prs':prs,
        'comm_prs':comm_prs,
        'distributed':'distributed' in str(path),
    })

owned={'rows':rows,'prs':sorted(all_prs,key=int)}
json.dump(owned, open('/tmp/owned.json','w'), indent=1)
print('owned rows:',len(rows),'unique PRs:',len(all_prs))
