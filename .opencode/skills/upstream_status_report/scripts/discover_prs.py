#!/usr/bin/env python3
"""Discover new pytorch/pytorch PRs for files tracked in the xlsx and fill in
their PR links.

Logic:
  1. Collect the set of authors from PRs already recorded in the table
     (read from cached PR JSON in pr_cache/, keyed by /tmp/owned.json['prs']).
  2. Determine a cutoff = the newest createdAt among recorded PRs.
  3. For each author, ask GitHub for their pytorch/pytorch PRs created after the
     cutoff (`gh search prs`).
  4. For each candidate PR, fetch the list of changed files. If any changed file
     matches a file path tracked in the table (and the PR isn't already
     recorded), record the PR URL against that row.
  5. Write the PR URL into the xlsx PR column (append, de-duplicated). A backup
     is made before saving.

Run extract_owned.py first (build_report.sh step 1) so /tmp/owned.json exists.

Usage:
    python3 discover_prs.py                 # dry-run: show what would change
    python3 discover_prs.py --apply         # write links into the xlsx (backs up)
    python3 discover_prs.py --apply --rebuild   # then refresh report.html
    python3 discover_prs.py --since 2026-08-01  # override cutoff date
    python3 discover_prs.py --limit 200     # per-author PR search cap
"""
import json, os, re, sys, shutil, datetime, subprocess

HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, 'test_files_by_category_20260723.xlsx')
CACHE = os.path.join(HERE, 'pr_cache')
OWNED = '/tmp/owned.json'

# column map (1-indexed) on 'Test Files' sheet (must match extract_owned.py)
C_PATH = 3; C_PR = 6; C_TEAM = 12; C_COMMPR = 15

PULL_RE = re.compile(r'pytorch/pytorch/pull/(\d+)')
HASH_RE = re.compile(r'#(\d+)')


def parse_prs(v):
    if v is None:
        return []
    s = str(v)
    nums = [m.group(1) for m in PULL_RE.finditer(s)]
    if not nums:
        nums = [m.group(1) for m in HASH_RE.finditer(s)]
    seen = set(); out = []
    for x in nums:
        if x not in seen:
            seen.add(x); out.append(x)
    return out


def gh_json(args):
    try:
        out = subprocess.run(args, capture_output=True, text=True, check=True).stdout
        return json.loads(out) if out.strip() else None
    except subprocess.CalledProcessError as e:
        tail = (e.stderr or '').strip().splitlines()
        print(f'  ! gh failed ({tail[-1] if tail else "?"}): {" ".join(args[:6])}')
        return None


def arg(name, default=None):
    if name in sys.argv:
        i = sys.argv.index(name)
        if i + 1 < len(sys.argv):
            return sys.argv[i + 1]
    return default


def main():
    apply = '--apply' in sys.argv
    rebuild = '--rebuild' in sys.argv
    limit = int(arg('--limit', '100'))

    if not os.path.exists(OWNED):
        sys.exit('missing /tmp/owned.json — run extract_owned.py first')
    owned = json.load(open(OWNED))
    recorded = set(owned['prs'])

    # ---- 1. authors + 2. cutoff from recorded (cached) PRs -----------------
    authors = set(); cutoff = None
    for n in owned['prs']:
        f = os.path.join(CACHE, f'{n}.json')
        if not os.path.exists(f):
            continue
        d = json.load(open(f))
        a = (d.get('author') or {}).get('login')
        if a:
            authors.add(a)
        c = d.get('createdAt')
        if c and (cutoff is None or c > cutoff):
            cutoff = c
    since = arg('--since') or (cutoff or '')[:10]
    if not since:
        sys.exit('could not determine cutoff date; pass --since YYYY-MM-DD')
    print(f'{len(authors)} authors, cutoff (PRs created after) = {since}')

    # ---- owned file paths -> set (for matching) ----------------------------
    owned_paths = {r['path'] for r in owned['rows'] if r.get('path')}
    print(f'{len(owned_paths)} tracked file paths')

    # ---- 3. per-author candidate PRs ---------------------------------------
    candidates = {}   # pr number(str) -> {url,title,author,createdAt}
    for i, a in enumerate(sorted(authors), 1):
        res = gh_json(['gh', 'search', 'prs', '--repo', 'pytorch/pytorch',
                       '--author', a, '--created', f'>{since}',
                       '--limit', str(limit),
                       '--json', 'number,title,url,createdAt,author'])
        found = 0
        for pr in (res or []):
            n = str(pr['number'])
            if n in recorded:
                continue
            candidates[n] = {
                'url': pr.get('url') or f'https://github.com/pytorch/pytorch/pull/{n}',
                'title': pr.get('title', ''),
                'author': (pr.get('author') or {}).get('login', a),
                'createdAt': pr.get('createdAt', ''),
            }
            found += 1
        print(f'[{i}/{len(authors)}] {a}: {found} new candidate PR(s)')
    print(f'total unique new candidate PRs: {len(candidates)}')

    # ---- 4. match candidates to tracked files by changed files -------------
    # matches: path -> list of (pr_number, url)
    matches = {}
    for j, (n, info) in enumerate(sorted(candidates.items(), key=lambda x: int(x[0])), 1):
        d = gh_json(['gh', 'pr', 'view', n, '--repo', 'pytorch/pytorch',
                     '--json', 'files'])
        files = [f['path'] for f in (d or {}).get('files', [])] if d else []
        hit = owned_paths.intersection(files)
        if hit:
            for p in hit:
                matches.setdefault(p, []).append((n, info['url']))
            print(f'[{j}/{len(candidates)}] PR {n} ({info["author"]}) -> {sorted(hit)}')
    print(f'\nPRs matching tracked files touch {len(matches)} file(s)')

    if not matches:
        print('nothing to fill.')
        return

    # ---- 5. write links into xlsx ------------------------------------------
    import openpyxl
    wb = openpyxl.load_workbook(XLSX)
    wbv = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb['Test Files']; wsv = wbv['Test Files']

    # map tracked path -> excel row index
    path_row = {}
    for ri in range(2, wsv.max_row + 1):
        path = wsv.cell(ri, C_PATH).value
        team = wsv.cell(ri, C_TEAM).value
        if path and team not in (None, '') and path in matches:
            path_row.setdefault(path, ri)

    changes = []
    for path, prs in matches.items():
        ri = path_row.get(path)
        if not ri:
            print(f'  ! no owned row for {path}; skipped')
            continue
        cur = ws.cell(ri, C_PR).value
        existing = set(parse_prs(cur))
        add = [(n, url) for (n, url) in prs if n not in existing]
        if not add:
            continue
        new_links = '\n'.join(url for _, url in add)
        ws.cell(ri, C_PR).value = (str(cur).rstrip() + '\n' + new_links) if cur else new_links
        changes.append((ri, path, [n for n, _ in add]))

    print(f'\nrows to update: {len(changes)}')
    for ri, path, ns in changes:
        print(f'  row {ri}: {path}  += PR {ns}')

    if not changes:
        print('all matched PRs already recorded; nothing to write.')
        return

    if not apply:
        print('\n(dry-run) re-run with --apply to write these into the xlsx.')
        return

    bak = XLSX.replace('.xlsx', f'.predisc_{datetime.datetime.now():%Y%m%d_%H%M%S}.bak.xlsx')
    shutil.copy(XLSX, bak)
    wb.save(XLSX)
    print(f'\nbackup: {os.path.basename(bak)}')
    print(f'saved:  {os.path.basename(XLSX)}')

    if rebuild:
        print('\n== rebuilding report ==')
        subprocess.run(['bash', os.path.join(HERE, 'build_report.sh'), '--refresh'], check=False)


if __name__ == '__main__':
    main()
