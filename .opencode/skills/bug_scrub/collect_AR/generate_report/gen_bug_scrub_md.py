"""Generate bug_scrub.md from Issues sheet + AR classification.

Phase 5 grouping is driven entirely by the `AR` column in the Issues sheet,
which holds one or more of the 7 canonical action-required buckets joined
with '; ':

    1. Close/Skip
    2. Need Owner
    3. Land PR
    4. Wait for PR
    5. Need Response
    6. Need check case existence
    7. Verify

An issue appears in EVERY AR bucket it lists (multi-membership). Rows with
an empty AR cell are collected into UNCLASSIFIED.

The legacy `action_Type` column is NO LONGER read by this script.

Usage:
    # default full report
    python3 gen_bug_scrub_md.py

    # filtered report (e.g. UT-scoped)
    python3 gen_bug_scrub_md.py \\
        --issues-file ut_issues.txt \\
        --out bug_scrub_ut.md \\
        --title-suffix " — UT scope"
"""
from __future__ import annotations

import argparse
import re
import sys
import textwrap
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

import openpyxl

_COMMON_DIR = Path(__file__).resolve().parents[2] / "_common"
if str(_COMMON_DIR) not in sys.path:
    sys.path.insert(0, str(_COMMON_DIR))
from header_utils import row_dict  # type: ignore[reportMissingImports] # noqa: E402
from paths import RESULT_DIR, AGENT_SPACE  # type: ignore[reportMissingImports] # noqa: E402

_AGENT_SPACE = AGENT_SPACE
if str(_AGENT_SPACE) not in sys.path:
    sys.path.insert(0, str(_AGENT_SPACE))
from linkify_action_tbd import linkify_md  # type: ignore[reportMissingImports] # noqa: E402

_p = argparse.ArgumentParser(description=__doc__.splitlines()[0])
_p.add_argument("--issues-file", type=Path, default=None,
                help="Optional path to a file with whitespace/newline-"
                     "separated issue IDs. When given, only those issues "
                     "are included in the report.")
_p.add_argument("--out", type=Path, default=None,
                help="Output .md path (relative paths resolve against "
                     "the current working directory). Defaults to "
                     "$BUG_SCRUB_RESULT_DIR/bug_scrub.md.")
_p.add_argument("--title-suffix", default="",
                help="String appended to the report's top-level heading.")
_args = _p.parse_args()

EXCEL = RESULT_DIR / "torch_xpu_ops_issues.xlsx"
if _args.out is None:
    OUT = RESULT_DIR / "bug_scrub.md"
elif _args.out.is_absolute():
    OUT = _args.out
else:
    OUT = Path.cwd() / _args.out
REPO  = "intel/torch-xpu-ops"
TODAY = datetime.now(timezone.utc)
RECENT_CUTOFF = TODAY - timedelta(days=7)

AR_SECTIONS = [
    "Close/Skip",
    "Need Owner",
    "Land PR",
    "Wait for PR",
    "Need Response",
    "Need check case existence",
    "Verify",
]

AR_TITLES = {
    "Close/Skip":                 "Close/Skip — terminal QA action (close fixed, verify-and-close merged fix, skip not-target / wontfix)",
    "Need Owner":                 "Need Owner — awaiting triage-lead to assign an owner",
    "Land PR":                    "Land PR — numbered PR in action_TBD is the next concrete action",
    "Wait for PR":                "Wait for PR — fix path is known but no PR is filed yet; awaiting PR submission (or external non-PR tracker)",
    "Need Response":              "Need Response — owner / reporter must answer an open question (or no response yet on a new issue)",
    "Need check case existence":  "Need check case existence — XPU test case missing in repo; QA must verify case existence before action",
    "Verify":                     "Verify — referenced PR in action_TBD has merged AND owner_transferred=Reporter; reporter must verify the fix and confirm closure",
    "UNCLASSIFIED":               "UNCLASSIFIED — Phase 4d produced no AR verdict; should be empty after AR backfill",
}

PRIO_RANK = {"P0": 0, "P1": 1, "P2": 2, "P3": 3, "": 9, None: 9}


# -------- load -------------------------------------------------------------
wb  = openpyxl.load_workbook(EXCEL)
ws  = wb["Issues"]
C = {k: k for k in [
    "Issue ID","Title","Status","Assignee","Reporter","Labels","Created Time",
    "Category","Priority","Root Cause","Fix Approach","action_TBD","action_reason",
    "owner_transferred","duplicated_issue","Dependency","AR",
]}

def split_ar(v) -> list[str]:
    """Split the AR cell on '; ' into the canonical bucket labels.

    Skips empty / 'none' values and trims whitespace. Returns a deduped
    list preserving canonical AR_SECTIONS order when multiple buckets
    are present.
    """
    s = clean(v) if v is not None else ""
    if not s:
        return []
    raw = [tok.strip() for tok in s.split(";") if tok.strip()]
    seen, out = set(), []
    for b in raw:
        if b not in seen:
            seen.add(b)
            out.append(b)
    return out

# Test case index: {issue_id: [ {source, name, file, error, traceback, ...status cols}, ... ]}
# Includes ALL test case rows (not just those with traceback) so detail files
# can render a complete results table AND inline tracebacks.
tc_by_issue: dict[int, list[dict]] = defaultdict(list)
for sheet_name, src_label, id_col in [
    ("Test Cases", "UT", "Issue ID"),
    ("E2E Test Cases", "E2E", "Issue ID"),
    ("Others", "Others", "ID"),
]:
    ws_tb = wb[sheet_name]
    headers = set(row_dict(ws_tb, 1))
    def _safe_value(data, name):
        return data.get(name) if name in headers else None
    if id_col not in headers:
        continue
    for row_idx in range(2, ws_tb.max_row + 1):
        r_tb = row_dict(ws_tb, row_idx)
        iid_v = r_tb.get(id_col)
        if iid_v is None:
            continue
        try:
            iid_int = int(iid_v)
        except (TypeError, ValueError):
            continue
        name = ""
        if _safe_value(r_tb, "Test Case"):
            name = str(r_tb["Test Case"])
        elif _safe_value(r_tb, "Model"):
            name = f"{r_tb.get('Benchmark') or ''}/{r_tb['Model']}".strip("/")
        elif _safe_value(r_tb, "Title"):
            name = str(r_tb["Title"])
        tb_v = _safe_value(r_tb, "Traceback")
        xpu_exist = _safe_value(r_tb, "XPU Case Exist") or _safe_value(r_tb, "xpu_case_existence")
        entry = {
            "source": src_label,
            "name": name,
            "file": str(_safe_value(r_tb, "Test File")) if _safe_value(r_tb, "Test File") else "",
            "test_class": str(_safe_value(r_tb, "Test Class")) if _safe_value(r_tb, "Test Class") else "",
            "error": str(_safe_value(r_tb, "Error Message")) if _safe_value(r_tb, "Error Message") else "",
            "traceback": str(tb_v) if tb_v else "",
            "xpu_status": str(_safe_value(r_tb, "XPU Status")) if _safe_value(r_tb, "XPU Status") else "",
            "stock_status": str(_safe_value(r_tb, "Stock Status")) if _safe_value(r_tb, "Stock Status") else "",
            "xpu_accuracy": str(_safe_value(r_tb, "XPU Accuracy Status")) if _safe_value(r_tb, "XPU Accuracy Status") else "",
            "xpu_case_exist": str(xpu_exist) if xpu_exist else "",
            "cuda_case_exist": str(_safe_value(r_tb, "CUDA Case Exist")) if _safe_value(r_tb, "CUDA Case Exist") else "",
            "dup": str(_safe_value(r_tb, "duplicated_issue")) if _safe_value(r_tb, "duplicated_issue") else "",
            "phase": str(_safe_value(r_tb, "Phase")) if _safe_value(r_tb, "Phase") else "",
            "dtype": str(_safe_value(r_tb, "Dtype")) if _safe_value(r_tb, "Dtype") else "",
            "local_status": str(_safe_value(r_tb, "Local Status")) if _safe_value(r_tb, "Local Status") else "",
        }
        tc_by_issue[iid_int].append(entry)

# Legacy alias used by write_detail — rows with non-empty traceback only
tb_by_issue: dict[int, list[dict]] = defaultdict(list)
for iid, entries in tc_by_issue.items():
    for e in entries:
        if e["traceback"]:
            tb_by_issue[iid].append(e)

rows = [row_dict(ws, row_idx) for row_idx in range(2, ws.max_row + 1)]
print(f"loaded {len(rows)} rows")

def _clean_pre(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "none" else s

# Open status is derived ONLY from the workbook Status column (the
# authoritative source refreshed by Phase 1). Never read a cached open-id
# list: a stale cache silently drops newly-opened issues from the report.
open_ids = {int(r[C["Issue ID"]]) for r in rows
            if _clean_pre(r[C["Status"]]).lower() == "open"}
rows = [r for r in rows if r[C["Issue ID"]] is not None
        and int(r[C["Issue ID"]]) in open_ids]
print(f"filtered to {len(rows)} open issues")

if _args.issues_file is not None:
    ifile = _args.issues_file
    if not ifile.is_absolute():
        # resolve relative to the script's directory first, then CWD
        cand = Path(__file__).resolve().parent / ifile
        ifile = cand if cand.exists() else ifile
    wanted = {int(tok.lstrip("#")) for tok in ifile.read_text().split()
              if tok.strip().lstrip("#").isdigit()}
    _iid = C["Issue ID"]
    rows = [r for r in rows if r[_iid] is not None and int(r[_iid]) in wanted]
    found_ids = {int(r[_iid]) for r in rows}
    missing = sorted(wanted - found_ids)
    print(f"filtered to {len(rows)} rows matching {len(wanted)} requested IDs"
          f"{' (missing: ' + ', '.join(str(m) for m in missing) + ')' if missing else ''}")


# -------- helpers ----------------------------------------------------------
def clean(v) -> str:
    if v is None: return ""
    s = str(v).strip()
    if s.lower() == "none": return ""
    return s

def owner(r) -> str:
    a = clean(r[C["Assignee"]])
    if a: return a
    o = clean(r[C["owner_transferred"]])
    return o

def assignee_only(r) -> str:
    """Raw GitHub Assignee value, with the literal string 'None' (which is
    how the workbook represents an unassigned issue) treated as empty.
    Used in report tables alongside the separate `Owner Transferred`
    column so the two cells are actually different on rows where
    triage transferred ownership."""
    a = clean(r[C["Assignee"]])
    return "" if a.lower() == "none" else a

def owner_transferred_cell(r) -> str:
    return clean(r[C["owner_transferred"]])

def esc(s: str, max_len: int = 0) -> str:
    s = s.replace("\r", " ").replace("\n", " ").replace("|", "\\|")
    s = re.sub(r"\s+", " ", s)
    if max_len and len(s) > max_len:
        s = s[: max_len - 1] + "…"
    return s

def fmt_list(v) -> str:
    """action_TBD / action_reason cells are JSON arrays; join with '; '."""
    s = clean(v)
    if not s:
        return ""
    if s.startswith("["):
        try:
            import json as _j
            items = _j.loads(s)
            if isinstance(items, list):
                return "; ".join(str(x) for x in items)
        except Exception:
            pass
    return s


def split_action_tbd(v) -> list[str]:
    """Split an action_TBD cell into a list of items.

    The cell may be:
      - JSON array (legacy)
      - Phase 4b style: "tok_a, tok_b | sentence c | sentence d"
        where ' | ' separates Phase 4b additions and ', ' separates 4a/4c tokens.
    Returns a flat list of trimmed non-empty items.
    """
    s = clean(v)
    if not s:
        return []
    # JSON array
    if s.startswith("["):
        try:
            import json as _j
            parsed = _j.loads(s)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            pass
    # Pipe-separated; first segment may itself be comma-separated 4a/4c tokens
    items: list[str] = []
    for part in s.split("|"):
        part = part.strip()
        if not part:
            continue
        # Only split commas in the FIRST segment if it looks like short tokens
        # (4a/4c tokens never contain spaces in their canonical form). To be
        # safe, split only when no embedded sentence punctuation is present.
        if "," in part and len(part) < 80 and ":" not in part:
            for sub in part.split(","):
                sub = sub.strip()
                if sub:
                    items.append(sub)
        else:
            items.append(part)
    return items


def fmt_action_tbd_cell(v, max_item_len: int = 0) -> str:
    """Render action_TBD as a multi-line bulleted cell.

    Each pipe-separated segment becomes its own bullet line. We emit inline
    HTML (`<ul><li>...</li></ul>`) rather than markdown bullets because
    markdown table cells don't honor list syntax, and `* item *` would
    collide with the inline italic regex in the HTML renderer. Pipes inside
    the cell are escaped as `\\|`. Long items can be optionally truncated.
    """
    items = split_action_tbd(v)
    if not items:
        return ""
    out = []
    for it in items:
        it = it.replace("\r", " ").replace("\n", " ").replace("|", "\\|")
        it = re.sub(r"\s+", " ", it).strip()
        if max_item_len and len(it) > max_item_len:
            it = it[: max_item_len - 1] + "…"
        out.append(f"<li>{linkify_md(it)}</li>")
    return "<ul>" + "".join(out) + "</ul>"


def wrap_cell(s, width: int = 80) -> str:
    """Soft-wrap a cell to `width` chars per visual line using <br>.
    Escapes pipes, collapses whitespace, word-wraps at word boundaries."""
    s = clean(s).replace("\r", " ").replace("\n", " ").replace("|", "\\|")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""
    return "<br>".join(
        textwrap.wrap(s, width=width, break_long_words=True,
                      break_on_hyphens=False)
    )


# Patterns for Fix Approach beautification.
# Match file paths with common source/config extensions.
_PATH_RE = re.compile(
    r"(?<![`\w/.])"                               # not preceded by backtick or path char
    r"([\w./-]+\.(?:py|cpp|cmake|h|hpp|cu|cuh|xml|md|rst|yaml|yml|json))"
    r"(?![\w/.])"                                 # not followed by path char
)
# Sentence boundary: period+space before an uppercase ASCII letter, OR "; ".
# Avoids splitting on "e.g.", "vs.", "i.e." because they end with a period+space+lowercase.
_SPLIT_RE = re.compile(r"(?:\.\s+(?=[A-Z])|;\s+)")


def format_fix_approach(s, width: int = 80) -> str:
    """Bulletize Fix Approach text and wrap paths / quoted identifiers in
    backticks for readability.

    Pipeline:
      1. Clean & normalise whitespace.
      2. Wrap `'…'` single-quoted tokens and file paths in backticks
         (skipping content already inside backticks).
      3. Split on sentence boundaries ('. ' before uppercase, '; ').
      4. Soft-wrap each bullet to `width`; join with '<br>• '.
    """
    s = clean(s).replace("\r", " ").replace("\n", " ").replace("|", "\\|")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""

    # Pass 1: protect existing backtick spans so we don't double-wrap.
    spans: list[str] = []
    def _save(m):
        spans.append(m.group(0))
        return f"\x00{len(spans)-1}\x00"
    s = re.sub(r"`[^`]+`", _save, s)

    # Pass 2: wrap single-quoted identifiers → backticks.
    s = re.sub(r"'([^'\s][^']{0,120}?)'", r"`\1`", s)

    # Pass 3: wrap file paths in backticks.
    s = _PATH_RE.sub(r"`\1`", s)

    # Restore protected spans.
    s = re.sub(r"\x00(\d+)\x00", lambda m: spans[int(m.group(1))], s)

    # Split into bullets on sentence boundaries. Preserve trailing '.' on
    # each bullet except those produced by '; ' splits.
    bullets: list[str] = []
    buf = s
    while True:
        m = _SPLIT_RE.search(buf)
        if not m:
            if buf.strip():
                bullets.append(buf.strip())
            break
        head = buf[:m.start()].strip()
        sep = m.group(0)
        if head:
            # If the split was on '. ', restore the period.
            if sep.startswith("."):
                head = head + "."
            bullets.append(head)
        buf = buf[m.end():]

    if not bullets:
        return ""

    # Wrap each bullet; prefix '• '; join with '<br>'.
    lines: list[str] = []
    for b in bullets:
        wrapped = textwrap.wrap(b, width=width - 2,   # -2 for '• ' prefix
                                break_long_words=False,
                                break_on_hyphens=False)
        if not wrapped:
            continue
        lines.append("• " + wrapped[0])
        # Continuation lines of the same bullet: indent with 2 spaces so
        # they visually align under the bullet glyph.
        for cont in wrapped[1:]:
            lines.append("&nbsp;&nbsp;" + cont)
    return "<br>".join(lines)


DUP_TOKEN = re.compile(r"#?(\d+)")


# ---- per-issue detail files ----------------------------------------------
DETAILS_DIR = OUT.parent / "details"
DETAILS_REL = "details"

def _wrap_para(s: str, width: int = 100) -> str:
    """Wrap free-form prose for detail-file body (plain markdown, no <br>)."""
    s = clean(s).replace("\r", " ")
    s = re.sub(r"[ \t]+", " ", s)
    paras = [p.strip() for p in re.split(r"\n\s*\n", s) if p.strip()]
    out = []
    for p in paras:
        p_flat = re.sub(r"\s+", " ", p)
        out.append("\n".join(textwrap.wrap(p_flat, width=width,
                                           break_long_words=False,
                                           break_on_hyphens=False)))
    return "\n\n".join(out)

def _bullets(raw, linkify: bool = False) -> str:
    """Render action_TBD / action_reason JSON list cells as markdown bullets."""
    s = clean(raw)
    if not s:
        return "_(none)_"
    items: list[str] = []
    if s.startswith("["):
        try:
            import json as _j
            parsed = _j.loads(s)
            if isinstance(parsed, list):
                items = [str(x) for x in parsed]
        except Exception:
            items = [s]
    if not items:
        # Phase 4b pipe-separated list (and possibly comma-separated 4a/4c tokens)
        items = split_action_tbd(s)
    if not items:
        items = [s]
    if linkify:
        items = [linkify_md(it) for it in items]
    return "\n".join(f"* {it}" for it in items)

def _fix_approach_md(raw) -> str:
    """Render Fix Approach as bulleted markdown for detail files."""
    s = clean(raw)
    if not s:
        return "_(none)_"
    s = re.sub(r"\s+", " ", s).strip()
    parts = [p.strip() for p in _SPLIT_RE.split(s) if p.strip()]
    if not parts:
        return s
    out = []
    for i, p in enumerate(parts):
        if i < len(parts) - 1 and not p.endswith(".") and not p.endswith(":"):
            p = p + "."
        out.append(f"- {p}")
    return "\n".join(out)

def _preview(raw, max_chars: int = 100) -> str:
    """Short preview for Fix Approach table cell: first sentence or max_chars."""
    s = clean(raw).replace("|", "\\|")
    s = re.sub(r"\s+", " ", s).strip()
    if not s:
        return ""
    m = _SPLIT_RE.search(s)
    if m and m.start() <= max_chars:
        head = s[:m.start()].rstrip()
        if m.group(0).startswith("."):
            head += "."
        return head
    if len(s) <= max_chars:
        return s
    return s[:max_chars - 1].rstrip() + "…"

def fix_approach_cell(r) -> str:
    """Truncated Fix Approach + link to per-issue detail file."""
    iid = r[C["Issue ID"]]
    preview = _preview(r[C["Fix Approach"]])
    link = f"[→ details]({DETAILS_REL}/{iid}.md)"
    return f"{preview}<br>{link}" if preview else link

def write_detail(r) -> None:
    iid = r[C["Issue ID"]]
    if iid is None:
        return
    title = clean(r[C["Title"]])
    buf: list[str] = []
    a = buf.append
    a(f"# Issue #{iid}: {title}")
    a("")
    a(f"- **GitHub**: https://github.com/{REPO}/issues/{iid}")
    for key, label in [
        ("Category", "Category"),
        ("Priority", "Priority"),
        ("Status", "Status"),
        ("Assignee", "Assignee"),
        ("owner_transferred", "owner_transferred"),
        ("Reporter", "Reporter"),
        ("Labels", "Labels"),
        ("Dependency", "Dependency"),
        ("AR", "AR"),
    ]:
        v = clean(r[C[key]])
        a(f"- **{label}**: {v if v else '_(blank)_'}")
    a("")
    a("## action_TBD")
    a("")
    a(_bullets(r[C["action_TBD"]], linkify=True))
    a("")
    a("## action_reason")
    a("")
    a(_bullets(r[C["action_reason"]]))
    a("")
    a("## Root Cause")
    a("")
    rc = clean(r[C["Root Cause"]])
    a(_wrap_para(rc) if rc else "_(none)_")
    a("")
    a("## Fix Approach")
    a("")
    a(_fix_approach_md(r[C["Fix Approach"]]))
    a("")
    try:
        iid_int = int(iid)
    except (TypeError, ValueError):
        iid_int = None

    all_cases = tc_by_issue.get(iid_int, []) if iid_int is not None else []
    if all_cases:
        ut_cases = [e for e in all_cases if e["source"] == "UT"]
        e2e_cases = [e for e in all_cases if e["source"] == "E2E"]
        other_cases = [e for e in all_cases if e["source"] == "Others"]

        def _short(s, n=120):
            if not s:
                return ""
            return (s[:n] + ("…" if len(s) > n else "")).replace("|", "\\|").replace("\n", " ")

        if ut_cases:
            a(f"## UT Test Case Results ({len(ut_cases)})")
            a("")
            a("| # | Test Case | Test File | XPU Status | Stock Status | Local Status | XPU Case Exist | Error Message |")
            a("|---|---|---|---|---|---|---|---|")
            for i, e in enumerate(ut_cases, start=1):
                file_cell = f"`{e['file']}`" if e["file"] else ""
                a(f"| {i} | {e['name']} | {file_cell}"
                  f" | {e['xpu_status']} | {e['stock_status']} | {e['local_status']}"
                  f" | {e['xpu_case_exist']} | {_short(e['error'])} |")
            a("")

        if e2e_cases:
            a(f"## E2E Test Case Results ({len(e2e_cases)})")
            a("")
            a("| # | Model | Phase | Dtype | XPU Accuracy Status | Local Status | Error Message |")
            a("|---|---|---|---|---|---|---|")
            for i, e in enumerate(e2e_cases, start=1):
                a(f"| {i} | {e['name']} | {e['phase']} | {e['dtype']}"
                  f" | {e['xpu_accuracy']} | {e['local_status']} | {_short(e['error'])} |")
            a("")

        if other_cases:
            a(f"## Others Test Case Results ({len(other_cases)})")
            a("")
            a("| # | Title | Local Status | Error Message |")
            a("|---|---|---|---|")
            for i, e in enumerate(other_cases, start=1):
                title_cell = _short(e['name'], 80)
                a(f"| {i} | {title_cell} | {e['local_status']} | {_short(e['error'])} |")
            a("")

    tbs = tb_by_issue.get(iid_int, []) if iid_int is not None else []
    if tbs:
        a(f"## Test Cases & Traceback ({len(tbs)})")
        a("")
        for i, e in enumerate(tbs, start=1):
            hdr_bits = [f"{e['source']}"]
            if e["name"]:
                hdr_bits.append(e["name"])
            a(f"### {i}. {' · '.join(hdr_bits)}")
            a("")
            if e["file"]:
                a(f"- **Test File**: `{e['file']}`")
            if e["error"]:
                err = e["error"].replace("\r", " ").strip()
                err = re.sub(r"\s+", " ", err)
                a(f"- **Error**: {err[:300] + ('…' if len(err) > 300 else '')}")
            a("")
            a("```")
            a(e["traceback"].rstrip())
            a("```")
            a("")
    (DETAILS_DIR / f"{iid}.md").write_text("\n".join(buf))





def parse_dup_ids(dup_cell, action_tbd) -> list[int]:
    """Extract duplicate issue IDs from duplicated_issue cell, with
    action_TBD 'duplicate of …' fallback. Forward refs only."""
    ids: set[int] = set()
    for tok in re.split(r"[,;\s]+", clean(dup_cell)):
        m = DUP_TOKEN.fullmatch(tok)
        if m:
            ids.add(int(m.group(1)))
    if not ids:
        s = clean(action_tbd).lower()
        idx = s.find("duplicate of")
        if idx >= 0:
            for m in DUP_TOKEN.finditer(s[idx:idx+200]):
                ids.add(int(m.group(1)))
    return sorted(ids)


BACK = '_[↑ Back to Index](#sec-2)_'

def issue_link(iid) -> str:
    return f"[#{iid}](https://github.com/{REPO}/issues/{iid})"

def prio_key(r):
    p = clean(r[C["Priority"]]) or None
    return PRIO_RANK.get(p, 9)

def parse_dt(s) -> datetime | None:
    s = clean(s)
    if not s: return None
    try:
        return datetime.fromisoformat(s.replace("Z","+00:00"))
    except Exception:
        return None


# -------- bucket rows into AR sections ------------------------------------
by_section: dict[str, list] = defaultdict(list)
per_ar = Counter()
per_prio = Counter()
per_status = Counter()
per_category_col = Counter()
empty_ar = 0
check_cases_ids: list = []
unclassified_rows: list = []

for r in rows:
    buckets = split_ar(r[C["AR"]])
    if buckets:
        for b in buckets:
            per_ar[b] += 1
            if b in AR_SECTIONS:
                by_section[b].append(r)
        if "Need check case existence" in buckets:
            check_cases_ids.append(r[C["Issue ID"]])
    else:
        empty_ar += 1
        unclassified_rows.append(r)
    per_prio[clean(r[C["Priority"]]) or "(blank)"] += 1
    per_status[clean(r[C["Status"]]) or "(blank)"] += 1
    per_category_col[clean(r[C["Category"]]) or "(blank)"] += 1

def has_ar(r, bucket: str) -> bool:
    return bucket in split_ar(r[C["AR"]])

def is_terminal(r) -> bool:
    return has_ar(r, "Close/Skip")

# Duplicated: duplicated_issue non-empty OR action_TBD mentions "duplicate of"
dup_rows = [r for r in rows if clean(r[C["duplicated_issue"]]) or
            "duplicate of" in clean(r[C["action_TBD"]]).lower()]

def stale_items(r) -> list[str]:
    return [it for it in split_action_tbd(r[C["action_TBD"]]) if "(>1 week)" in it]
stale_rows = [r for r in rows if stale_items(r)]

def dep_ok(r) -> bool:
    d = clean(r[C["Dependency"]]).lower()
    if not d: return False
    if d == "upstream-pytorch": return False
    if d.startswith("sycl kernel"): return False
    if d == "cpu fallback": return False
    return True
dep_rows = [r for r in rows if dep_ok(r)]

# Third-party blockers: rows whose Dependency points at code Intel does not
# own (oneDNN/oneMKL/oneAPI/triton/driver/xccl). They are surfaced in §6 only;
# §3 hides them unless AR includes a live next action (Land PR or Wait for PR).
THIRD_PARTY_DEPS = {"onednn", "onemkl", "oneapi", "triton", "driver", "xccl"}
def is_third_party_blocked_for_sec3(r) -> bool:
    if has_ar(r, "Land PR") or has_ar(r, "Wait for PR"):
        return False
    return clean(r[C["Dependency"]]).lower() in THIRD_PARTY_DEPS

dep_rows = [r for r in dep_rows if not is_terminal(r)]

upstream_rows = [r for r in rows
                 if clean(r[C["Dependency"]]).lower() == "upstream-pytorch"
                 and not is_terminal(r)]
cpu_fb_rows   = [r for r in rows
                 if clean(r[C["Dependency"]]).lower() == "cpu fallback"
                 and not is_terminal(r)]

recent_rows = [r for r in rows
               if (dt := parse_dt(r[C["Created Time"]])) and dt >= RECENT_CUTOFF
               and not is_terminal(r)]


# -------- render -----------------------------------------------------------
def render_table(row_list) -> str:
    """Standard table: Issue | Title | Owner | Owner Transferred | action_TBD | Fix Approach | Priority | action_reason | Reporter | Labels"""
    if not row_list:
        return "_No issues._\n"
    head = "| Issue | Title | Owner | Owner Transferred | action_TBD | Fix Approach | Priority | action_reason | Reporter | Labels |"
    sep  = "|---|---|---|---|---|---|---|---|---|---|"
    out = [head, sep]
    sorted_rows = sorted(row_list, key=lambda r: (
        prio_key(r), str(r[C["Issue ID"]])
    ))
    for r in sorted_rows:
        out.append("| " + " | ".join([
            issue_link(r[C["Issue ID"]]),
            wrap_cell(r[C["Title"]], 50),
            esc(assignee_only(r), 25),
            esc(owner_transferred_cell(r), 25),
            fmt_action_tbd_cell(r[C["action_TBD"]]),
            fix_approach_cell(r),
            esc(clean(r[C["Priority"]]), 6),
            esc(fmt_list(r[C["action_reason"]]), 140),
            esc(clean(r[C["Reporter"]]), 20),
            esc(clean(r[C["Labels"]]), 40),
        ]) + " |")
    return "\n".join(out) + "\n"


def slug(s: str) -> str:
    """GitHub-style anchor slug."""
    s = (s or "").lower().strip()
    s = s.replace("_", "-")
    s = re.sub(r"[^a-z0-9\s-]", "", s)
    s = re.sub(r"\s+", "-", s)
    return s


def render_section_by_category(row_list, section_num: str, cat_prefix: str) -> list[str]:
    """Split rows by their `Category` column; emit a `#### cat_prefix.N <Cat>` sub-heading with a table per group.
    Returns list of markdown lines and list of (anchor, label) for TOC."""
    buckets: dict[str, list] = defaultdict(list)
    for r in row_list:
        buckets[clean(r[C["Category"]]) or "(blank)"].append(r)
    out: list[str] = []
    toc: list[tuple[str, str]] = []
    for idx, cat in enumerate(sorted(buckets), start=1):
        rows_c = buckets[cat]
        anchor = f"sec-{section_num.replace('.','-')}-{idx}-{slug(cat)}"
        out.append(f'<a id="{anchor}"></a>')
        out.append(f"#### {section_num}.{idx} {cat}  ·  {len(rows_c)} issues")
        out.append("")
        out.append(BACK)
        out.append("")
        out.append(render_table(rows_c))
        out.append("")
        toc.append((anchor, f"{section_num}.{idx} {cat} ({len(rows_c)})"))
    return out, toc


def render_dep_table(row_list) -> str:
    head = "| Issue | Dependency | Title | Owner | Owner Transferred | action_TBD | Fix Approach | Priority | action_reason | Reporter | Labels |"
    sep  = "|---|---|---|---|---|---|---|---|---|---|---|"
    out = [head, sep]
    sorted_rows = sorted(row_list, key=lambda r: (
        prio_key(r), clean(r[C["Dependency"]]), str(r[C["Issue ID"]])
    ))
    for r in sorted_rows:
        out.append("| " + " | ".join([
            issue_link(r[C["Issue ID"]]),
            esc(clean(r[C["Dependency"]]), 30),
            wrap_cell(r[C["Title"]], 50),
            esc(assignee_only(r), 25),
            esc(owner_transferred_cell(r), 25),
            fmt_action_tbd_cell(r[C["action_TBD"]]),
            fix_approach_cell(r),
            esc(clean(r[C["Priority"]]), 6),
            esc(fmt_list(r[C["action_reason"]]), 140),
            esc(clean(r[C["Reporter"]]), 20),
            esc(clean(r[C["Labels"]]), 40),
        ]) + " |")
    return "\n".join(out) + "\n"


def render_stale_table(row_list) -> str:
    """Table of issues with one or more action_TBD items containing '(>1 week)'."""
    if not row_list:
        return "_No issues._\n"
    head = "| Issue | Title | Owner | Stale Requests | Priority | Reporter | Labels |"
    sep  = "|---|---|---|---|---|---|---|"
    out = [head, sep]
    sorted_rows = sorted(row_list, key=lambda r: (
        prio_key(r), str(r[C["Issue ID"]])
    ))
    for r in sorted_rows:
        items = stale_items(r)
        bullets = "<ul>" + "".join(f"<li>{esc(it, 0)}</li>" for it in items) + "</ul>"
        out.append("| " + " | ".join([
            issue_link(r[C["Issue ID"]]),
            wrap_cell(r[C["Title"]], 50),
            esc(assignee_only(r), 25),
            bullets,
            esc(clean(r[C["Priority"]]), 6),
            esc(clean(r[C["Reporter"]]), 20),
            esc(clean(r[C["Labels"]]), 40),
        ]) + " |")
    return "\n".join(out) + "\n"


def render_recent(row_list) -> str:
    head = "| Issue | Created | Title | Owner | Owner Transferred | action_TBD | Fix Approach | Priority | action_reason | Reporter | Labels |"
    sep  = "|---|---|---|---|---|---|---|---|---|---|---|"
    out = [head, sep]
    sorted_rows = sorted(
        row_list,
        key=lambda r: (prio_key(r), -(parse_dt(r[C["Created Time"]]) or TODAY).timestamp(), str(r[C["Issue ID"]])),
    )
    for r in sorted_rows:
        dt = parse_dt(r[C["Created Time"]])
        created = dt.strftime("%Y-%m-%d") if dt else ""
        out.append("| " + " | ".join([
            issue_link(r[C["Issue ID"]]),
            created,
            wrap_cell(r[C["Title"]], 50),
            esc(assignee_only(r), 25),
            esc(owner_transferred_cell(r), 25),
            fmt_action_tbd_cell(r[C["action_TBD"]]),
            fix_approach_cell(r),
            esc(clean(r[C["Priority"]]), 6),
            esc(fmt_list(r[C["action_reason"]]), 140),
            esc(clean(r[C["Reporter"]]), 20),
            esc(clean(r[C["Labels"]]), 40),
        ]) + " |")
    return "\n".join(out) + "\n"


def render_dup_table(row_list) -> str:
    """§5 table: adds a `Duplicates` column after `Issue` with clickable issue links
    parsed from `duplicated_issue` (fallback: 'duplicate of …' clause in action_TBD)."""
    if not row_list:
        return "_No issues._\n"
    head = "| Issue | Duplicates | Title | Owner | Owner Transferred | action_TBD | Fix Approach | Priority | action_reason | Reporter | Labels |"
    sep  = "|---|---|---|---|---|---|---|---|---|---|---|"
    out = [head, sep]
    sorted_rows = sorted(row_list, key=lambda r: (
        prio_key(r), str(r[C["Issue ID"]])
    ))
    for r in sorted_rows:
        ids = parse_dup_ids(r[C["duplicated_issue"]], r[C["action_TBD"]])
        if ids:
            dup_cell = ", ".join(issue_link(i) for i in ids)
        else:
            dup_cell = esc(clean(r[C["duplicated_issue"]]), 30)
        out.append("| " + " | ".join([
            issue_link(r[C["Issue ID"]]),
            dup_cell,
            wrap_cell(r[C["Title"]], 50),
            esc(assignee_only(r), 25),
            esc(owner_transferred_cell(r), 25),
            fmt_action_tbd_cell(r[C["action_TBD"]]),
            fix_approach_cell(r),
            esc(clean(r[C["Priority"]]), 6),
            esc(fmt_list(r[C["action_reason"]]), 140),
            esc(clean(r[C["Reporter"]]), 20),
            esc(clean(r[C["Labels"]]), 40),
        ]) + " |")
    return "\n".join(out) + "\n"


# ---- write per-issue detail files ----------------------------------------
DETAILS_DIR.mkdir(parents=True, exist_ok=True)
for _old_detail in DETAILS_DIR.glob("*.md"):
    _old_detail.unlink()
for _r in rows:
    write_detail(_r)
print(f"wrote {len(rows)} detail files to {DETAILS_DIR}")


# ---- assemble report ------------------------------------------------------
lines: list[str] = []
def w(s=""): lines.append(s)

w(f"# XPU Ops Bug Scrub Report{_args.title_suffix}")
w()
w(f"- **Repository**: `{REPO}`")
w(f"- **Generated**: {TODAY.strftime('%Y-%m-%d')} (cutoff for Section 6: {RECENT_CUTOFF.strftime('%Y-%m-%d')})")
w(f"- **Total issues in workbook**: {len(rows)}")
w(f"- **Classified (non-empty `AR`)**: {len(rows) - empty_ar}")
w(f"- **Empty `AR` (no verdict)**: {empty_ar}")
w()

# -- Section 1: Summary ----------------------------------------------------
w("## 1. Summary")
w()
w(f"This report groups the {len(rows)} tracked torch-xpu-ops issues by the "
  f"`AR` (Action Required) column in the workbook. An issue may appear in "
  f"multiple AR buckets if its `AR` cell contains more than one value "
  f"(joined with `; `). Cross-cutting slices (duplicated issues, external "
  f"dependency blockers, newly filed issues, stale requests) are listed "
  f"separately for visibility.")
w()
w("**Headline counts (multi-membership — an issue with N AR values is counted N times):**")
w()
w("| AR Bucket | Issues |")
w("|---|---:|")
for c in AR_SECTIONS:
    w(f"| {c} | {per_ar[c]} |")
w(f"| UNCLASSIFIED | {empty_ar} |")
w(f"| Duplicated | {len(dup_rows)} |")
w(f"| External dependency (non-upstream-pytorch, non-SYCL-kernel) | {len(dep_rows)} |")
w(f"| Upstream-pytorch | {len(upstream_rows)} |")
w(f"| CPU fallback | {len(cpu_fb_rows)} |")
w(f"| Filed within last 7 days | {len(recent_rows)} |")
w(f"| Requests pending > 1 week | {len(stale_rows)} |")
w()

# -- Section 2: Index ------------------------------------------------------
w('<a id="sec-2"></a>')
w("## 2. Index")
w()
w('- [3. Action Required (by AR bucket)](#sec-3)')
w('  - [UNCLASSIFIED](#sec-3-0-unclassified)')
for i, c in enumerate(AR_SECTIONS, start=1):
    w(f'  - [{c}](#sec-3-{i}-{slug(c)})')
w('- [4. Duplicated issues](#sec-4)')
w('- [5. Dependency (external blockers)](#sec-5)')
w('  - [Third Parties](#sec-5-1-third-parties)')
w('  - [upstream-pytorch](#sec-5-2-upstream-pytorch)')
w('  - [CPU fallback](#sec-5-3-cpu-fallback)')
w('- [6. New submitted issues (<7 days)](#sec-6)')
w('- [7. Requests pending > 1 week](#sec-7)')
w('- [8. Statistics](#sec-8)')
w()

# -- Section 3: Action Required (by AR bucket) -----------------------------
w('<a id="sec-3"></a>')
w("## 3. Action Required (by AR bucket)")
w()
w(BACK)
w()
w("Issues are grouped by the `AR` column from the Issues sheet. Each issue "
  "appears in every AR bucket it lists. Rows inside each bucket are split by "
  "`Category` (existing taxonomy column); rows within a category table are "
  "sorted by `Priority` (P0 → P3).")
w()
w("Issues whose `Dependency` is a third-party blocker "
  "(`oneDNN` / `oneMKL` / `oneAPI` / `triton` / `driver` / `xccl`) are "
  "hidden here and listed only under §5 Dependency, except when their "
  "AR includes `Land PR` or `Wait for PR` (a live next action makes the row actionable).")
w()

sec3_unclassified = [r for r in unclassified_rows
                     if not is_third_party_blocked_for_sec3(r)]
w('<a id="sec-3-0-unclassified"></a>')
w(f"- **UNCLASSIFIED**  ·  {len(sec3_unclassified)} issues")
w()
w(BACK)
w()
w(f"**{AR_TITLES['UNCLASSIFIED']}**")
w()
w(render_table(sec3_unclassified))
w()

for i, cat in enumerate(AR_SECTIONS, start=1):
    section_num = f"3.{i}"
    anchor = f"sec-3-{i}-{slug(cat)}"
    bucket = [r for r in by_section[cat]
              if not is_third_party_blocked_for_sec3(r)]
    w(f'<a id="{anchor}"></a>')
    w(f"- **{cat}**  ·  {len(bucket)} issues")
    w()
    w(f"**{AR_TITLES[cat]}**")
    w()
    sub_lines, _sub_toc = render_section_by_category(bucket, section_num, cat)
    for line in sub_lines:
        w(line)
w()

# -- Section 4: Duplicated -------------------------------------------------
w('<a id="sec-4"></a>')
w("## 4. Duplicated issues")
w()
w(BACK)
w()
w(f"Rows where `duplicated_issue` is set or `action_TBD` contains "
  f"\"duplicate of\".  —  {len(dup_rows)} issues.")
w()
w(render_dup_table(dup_rows))
w()

# -- Section 5: Dependency -------------------------------------------------
w('<a id="sec-5"></a>')
w("## 5. Dependency (external blockers)")
w()
w(BACK)
w()
w("Issues with a non-blank `Dependency` value, excluding `upstream-pytorch`, "
  "`CPU fallback`, and `SYCL kernel:*` (in-repo kernel pointers). "
  "Rows whose AR is `Close/Skip` are also excluded.  —  "
  f"{len(dep_rows)} issues.")
w()
w('<a id="sec-5-1-third-parties"></a>')
w("- **Third Parties**")
w()
w(BACK)
w()
w(render_dep_table(dep_rows))
w()

w('<a id="sec-5-2-upstream-pytorch"></a>')
w("- **upstream-pytorch**")
w()
w(BACK)
w()
w("Issues whose fix lives in `pytorch/pytorch` (Dynamo/Inductor, AOTAutograd, "
  "`_prims_common`, benchmark harness, test-list sync, etc.). Close/Skip rows "
  f"excluded.  —  {len(upstream_rows)} issues.")
w()
w(render_dep_table(upstream_rows))
w()

w('<a id="sec-5-3-cpu-fallback"></a>')
w("- **CPU fallback**")
w()
w(BACK)
w()
w("Issues where the XPU operator is missing and a CPU fallback is registered "
  "in torch-xpu-ops. Close/Skip rows excluded.  —  "
  f"{len(cpu_fb_rows)} issues.")
w()
w(render_dep_table(cpu_fb_rows))
w()

# -- Section 6: New <=7 days -----------------------------------------------
w('<a id="sec-6"></a>')
w("## 6. New submitted issues (<7 days)")
w()
w(BACK)
w()
w(f"Issues created on or after {RECENT_CUTOFF.strftime('%Y-%m-%d')}, "
  "excluding Close/Skip rows.  —  "
  f"{len(recent_rows)} issues.")
w()
w(render_recent(recent_rows))
w()

# -- Section 7: Requests pending > 1 week ----------------------------------
w('<a id="sec-7"></a>')
w("## 7. Requests pending > 1 week")
w()
w(BACK)
w()
w("Issues whose `action_TBD` contains one or more verbs flagged `(>1 week)` — "
  "an unresolved comment AR, unresolved PR review comments, or unaddressed CI "
  "failures that have been sitting more than 7 days. These are the highest-"
  "priority candidates for owner follow-up.")
w()
w(render_stale_table(stale_rows))
w()

# -- Section 8: Statistics -------------------------------------------------
w('<a id="sec-8"></a>')
w("## 8. Statistics")
w()
w(BACK)
w()
w(f"- Total rows: **{len(rows)}**")
w(f"- Classified (non-empty `AR`): **{len(rows) - empty_ar}**")
w(f"- Empty `AR` (no verdict yet): **{empty_ar}**")
w(f"- Issues flagged for test-case existence check (`Need check case existence`): **{len(check_cases_ids)}**")
w()

w("- **AR bucket distribution (multi-membership — an issue with N AR values is counted N times)**")
w()
w(BACK)
w()
w("| AR Bucket | Issues |")
w("|---|---:|")
for c in AR_SECTIONS:
    if per_ar[c]:
        w(f"| {c} | {per_ar[c]} |")
if empty_ar:
    w(f"| UNCLASSIFIED | {empty_ar} |")
w()

w("- **Priority distribution**")
w()
w(BACK)
w()
w("| Priority | Issues |")
w("|---|---:|")
for p in ["P0","P1","P2","P3","(blank)"]:
    if per_prio.get(p):
        w(f"| {p} | {per_prio[p]} |")
w()

w("- **Status distribution**")
w()
w(BACK)
w()
w("| Status | Issues |")
w("|---|---:|")
for s, n in per_status.most_common():
    w(f"| {s} | {n} |")
w()

w("- **Category column distribution (top 20)**")
w()
w(BACK)
w()
w("| Category | Issues |")
w("|---|---:|")
for c, n in per_category_col.most_common(20):
    w(f"| {c} | {n} |")
w()

w("- **`Need check case existence` issue IDs**")
w()
w(BACK)
w()
w(f"{len(check_cases_ids)} issues flagged for XPU test-case existence check:")
w()
w("> " + ", ".join(f"#{i}" for i in sorted(check_cases_ids)))
w()

# ---- write ----------------------------------------------------------------
OUT.write_text("\n".join(lines))
print(f"wrote {OUT} ({OUT.stat().st_size} bytes, {len(lines)} lines)")
