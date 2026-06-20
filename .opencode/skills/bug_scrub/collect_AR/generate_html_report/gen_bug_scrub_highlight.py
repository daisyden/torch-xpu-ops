"""Generate bug_scrub_highlight.html — priority-organized highlight report.

Reads the Excel workbook directly and produces an interactive single-file
HTML report organized by Priority (P0/P1/P2/P3), with:
  - AR filter (Close/Skip, Need Owner, Land PR, Wait for PR, Need Response, Need check case existence, Verify)
  - Duplicate issues filter button
  - Tables sorted by Category within each priority section
  - Duplicate and Dependency columns
  - Open-request >1 week rows shown by default; others behind a "more" link
  - Priority/Category tooltips showing reasons on hover
  - Done checkbox with localStorage persistence
  - Date filter to show only issues after a selected date
  - All filters from bug_scrub.html preserved (Assignee, Category, Dependency, Milestone, search)

Usage:
    python3 gen_bug_scrub_highlight.py
"""

from __future__ import annotations

import html
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta
from pathlib import Path

THIS = Path(__file__).resolve()
_COMMON_DIR = THIS.parents[2] / "_common"
if str(_COMMON_DIR) not in sys.path:
    sys.path.insert(0, str(_COMMON_DIR))
from paths import RESULT_DIR, AGENT_SPACE  # type: ignore[reportMissingImports] # noqa: E402

_MARKED_LIB_SRC = THIS.parent / "marked.min.js"
_MARKED_LIB: str = ""
if _MARKED_LIB_SRC.exists():
    _MARKED_LIB = _MARKED_LIB_SRC.read_text(encoding="utf-8")

XLSX_PATH = RESULT_DIR / "torch_xpu_ops_issues.xlsx"
HTML_PATH = RESULT_DIR / "bug_scrub_highlight.html"
OPEN_IDS_CACHE = RESULT_DIR / "open_issue_ids.json"
# Maps issue id -> close timestamp for issues dropped from the visible report.
# These feed the open-over-time trend (so closures show up) without appearing
# in the table/filters.
CLOSED_DATES_CACHE = RESULT_DIR / "closed_issue_dates.json"
REPO = "intel/torch-xpu-ops"
TODAY = datetime.now(timezone.utc)

_AGENT_SPACE = AGENT_SPACE
if str(_AGENT_SPACE) not in sys.path:
    sys.path.insert(0, str(_AGENT_SPACE))
from linkify_action_tbd import linkify_html  # noqa: E402

AR_LABELS = [
    "Close/Skip",
    "Need Owner",
    "Land PR",
    "Wait for PR",
    "Need Response",
    "Need check case existence",
    "Verify",
]

PRIO_ORDER = ["P0", "P1", "P2", "P3"]


def _blank_sycl_dep(dep: str) -> str:
    # Per-file "SYCL kernel:" entries are internal kernel work, not an external
    # dependency; blank them so they neither render nor appear in the filter.
    return "" if dep.strip().lower().startswith("sycl kernel:") else dep


def clean_ar(raw: str) -> str:
    """Normalise the AR cell to the canonical '; '-joined form.

    Strips empty / 'none' tokens and dedupes while preserving order.
    Returns '' if no recognisable AR value is present.
    """
    if not raw:
        return ""
    s = str(raw).strip()
    if not s or s.lower() == "none":
        return ""
    seen, out = set(), []
    for tok in s.split(";"):
        t = tok.strip()
        if not t or t.lower() == "none":
            continue
        if t not in seen:
            seen.add(t)
            out.append(t)
    return "; ".join(out)


def clean(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() == "none" else s


def esc(s: str) -> str:
    return html.escape(s, quote=True)


DUP_TOKEN = re.compile(r"#?(\d+)")


def parse_dup_ids(dup_cell, action_tbd) -> list[int]:
    ids: set[int] = set()
    for tok in re.split(r"[,;\s]+", clean(dup_cell)):
        m = DUP_TOKEN.fullmatch(tok)
        if m:
            ids.add(int(m.group(1)))
    if not ids:
        s = clean(action_tbd).lower()
        idx = s.find("duplicate of")
        if idx >= 0:
            for m in DUP_TOKEN.finditer(s[idx : idx + 200]):
                ids.add(int(m.group(1)))
    return sorted(ids)


def stale_items(action_tbd: str) -> list[str]:
    items = []
    for part in clean(action_tbd).replace("[", "").replace("]", "").replace('"', "").split("|"):
        part = part.strip()
        if "(>1 week)" in part:
            items.append(part)
    return items


def parse_dt(s) -> datetime | None:
    s = clean(s)
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


def load_issues(xlsx_path: Path) -> list[dict]:
    from openpyxl import load_workbook

    # read_only=False is required so that cell.comment is accessible.
    wb = load_workbook(xlsx_path, read_only=False, data_only=True)
    ws = wb["Issues"]
    header_row = [c.value for c in ws[1]]
    headers = [str(c) if c is not None else "" for c in header_row]
    col = {h: i for i, h in enumerate(headers)}
    at_col_idx = col.get("action_TBD")

    def g(row_vals, key):
        i = col.get(key)
        if i is None or i >= len(row_vals):
            return ""
        v = row_vals[i]
        if v is None:
            return ""
        s = str(v).strip()
        return "" if s.lower() == "none" else s

    open_ids: set[int] | None = None
    if OPEN_IDS_CACHE.exists():
        open_ids = {int(x) for x in json.loads(OPEN_IDS_CACHE.read_text())}

    issues = []
    for r in range(2, ws.max_row + 1):
        row_cells = ws[r]
        row_vals = [c.value for c in row_cells]
        iid = g(row_vals, "Issue ID")
        if not iid:
            continue
        iid_int = int(iid)
        if open_ids is not None and iid_int not in open_ids:
            continue
        if open_ids is None and g(row_vals, "Status").lower() != "open":
            continue
        action_tbd = g(row_vals, "action_TBD")
        # Full untruncated text lives in the openpyxl Comment on the
        # action_TBD cell (written by apply_need_response_cache.py).
        action_tbd_full = action_tbd
        if at_col_idx is not None and at_col_idx < len(row_cells):
            cmt = row_cells[at_col_idx].comment
            if cmt and cmt.text:
                action_tbd_full = cmt.text.strip()
        ar_value = clean_ar(g(row_vals, "AR"))
        dup_cell = g(row_vals, "duplicated_issue")
        dup_ids = parse_dup_ids(dup_cell, action_tbd)
        is_dup = bool(dup_ids) or "duplicate of" in action_tbd.lower()
        created = g(row_vals, "Created Time")
        stale = stale_items(action_tbd)

        issues.append({
            "id": iid_int,
            "title": g(row_vals, "Title"),
            "status": g(row_vals, "Status"),
            "assignee": g(row_vals, "Assignee"),
            "reporter": g(row_vals, "Reporter"),
            "labels": g(row_vals, "Labels"),
            "created": created,
            "created_dt": parse_dt(created),
            "closed_dt": None,
            "milestone": g(row_vals, "Milestone"),
            "test_module": g(row_vals, "Test Module"),
            "dependency": _blank_sycl_dep(g(row_vals, "Dependency")),
            "dependency_reason": g(row_vals, "dependency_reason"),
            "category": g(row_vals, "Category"),
            "priority": g(row_vals, "Priority"),
            "priority_reason": g(row_vals, "Priority Reason"),
            "category_reason": g(row_vals, "Category Reason"),
            "root_cause": g(row_vals, "Root Cause"),
            "fix_approach": g(row_vals, "Fix Approach"),
            "action_tbd": action_tbd,
            "action_tbd_full": action_tbd_full,
            "action_reason": g(row_vals, "action_reason"),
            "owner_transferred": g(row_vals, "owner_transferred"),
            "dup_cell": dup_cell,
            "dup_ids": dup_ids,
            "is_dup": is_dup,
            "ar": ar_value,
            "stale_items": stale,
            "is_stale": bool(stale),
        })
    wb.close()
    return issues


def load_closed_trend_issues(xlsx_path: Path) -> list[dict]:
    """Load issues that were dropped from the visible report because they are
    now closed, enriched with a close timestamp from CLOSED_DATES_CACHE. These
    are used only to make the open-over-time trend reflect closures; they are
    never rendered as table rows."""
    if not CLOSED_DATES_CACHE.exists():
        return []
    closed_at = {int(k): v for k, v in json.loads(CLOSED_DATES_CACHE.read_text()).items()}
    if not closed_at:
        return []
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(c.value) if c.value is not None else "" for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    def g(row_vals, key):
        i = col.get(key)
        if i is None or i >= len(row_vals):
            return ""
        v = row_vals[i]
        if v is None:
            return ""
        s = str(v).strip()
        return "" if s.lower() == "none" else s

    closed_issues = []
    for r in range(2, ws.max_row + 1):
        row_vals = [c.value for c in ws[r]]
        iid = g(row_vals, "Issue ID")
        if not iid:
            continue
        iid_int = int(iid)
        if iid_int not in closed_at:
            continue
        created = g(row_vals, "Created Time")
        action_tbd = g(row_vals, "action_TBD")
        closed_issues.append({
            "id": iid_int,
            "title": g(row_vals, "Title"),
            "assignee": g(row_vals, "Assignee"),
            "owner_transferred": g(row_vals, "owner_transferred"),
            "milestone": g(row_vals, "Milestone"),
            "test_module": g(row_vals, "Test Module"),
            "dependency": _blank_sycl_dep(g(row_vals, "Dependency")),
            "category": g(row_vals, "Category"),
            "priority": g(row_vals, "Priority"),
            "action_tbd": action_tbd,
            "created": created,
            "created_dt": parse_dt(created),
            "closed_dt": parse_dt(closed_at[iid_int]),
            "ar": clean_ar(g(row_vals, "AR")),
            "is_dup": bool(parse_dup_ids(g(row_vals, "duplicated_issue"), action_tbd))
            or "duplicate of" in action_tbd.lower(),
        })
    wb.close()
    return closed_issues


def truncate(s: str, n: int = 80) -> str:
    if len(s) <= n:
        return s
    return s[: n - 1] + "\u2026"


_PLEASE_REPLY_RE = re.compile(r"@([\w-]+):\s*please reply", re.IGNORECASE)


def summarize_action_tbd(text: str, limit: int = 80) -> str:
    if not text:
        return ""
    users = _PLEASE_REPLY_RE.findall(text)
    if users:
        seen = []
        for u in users:
            if u not in seen:
                seen.append(u)
        mentions = ", ".join("@" + u for u in seen)
        n = len(seen)
        summary = f"{mentions}: please reply ({n} pending)" if n > 1 else f"{mentions}: please reply"
        return truncate(summary, limit)
    return truncate(text, limit)


def split_owner_tokens(*values: str) -> list[str]:
    tokens: set[str] = set()
    for value in values:
        for tok in re.split(r"[,;|]", clean(value)):
            tok = tok.strip()
            if tok:
                tokens.add(tok)
    return sorted(tokens)


def render_issue_row(issue: dict, show_done: bool = True) -> str:
    """Render a single <tr> for an issue."""
    i = issue
    iid = i["id"]
    pr = esc(i["priority_reason"])
    cr = esc(i["category_reason"])

    attrs = [
        f'data-issue="{iid}"',
        f'data-priority="{esc(i["priority"])}"',
        f'data-category="{esc(i["category"])}"',
        f'data-assignee="{esc(i["assignee"])}"',
        f'data-owner_transferred="{esc(i["owner_transferred"])}"',
        f'data-dependency="{esc(i["dependency"])}"',
        f'data-milestone="{esc(i["milestone"])}"',
        f'data-test_module="{esc(i["test_module"])}"',
        f'data-ar="{esc(i["ar"])}"',
        f'data-is-dup="{1 if i["is_dup"] else 0}"',
        f'data-is-stale="{1 if i["is_stale"] else 0}"',
        f'data-created="{esc(i["created"][:10] if i["created"] else "")}"',
        f'data-search="{esc((str(iid) + " " + i["title"] + " " + (i.get("action_tbd_full") or i["action_tbd"])).lower())}"',
    ]

    cells = []
    if show_done:
        cells.append(
            f'<td class="done-col"><input type="checkbox" class="ar-done" '
            f'data-issue="{iid}" id="done-{iid}"></td>'
        )
    cells.append(
        f'<td><a href="https://github.com/{REPO}/issues/{iid}" '
        f'target="_blank" rel="noopener noreferrer">#{iid}</a></td>'
    )
    cells.append(f"<td>{esc(truncate(i['title'], 80))}</td>")
    cells.append(
        f'<td class="tip-cell" title="{pr}">{esc(i["priority"])}</td>'
    )
    cells.append(
        f'<td class="tip-cell" title="{cr}">{esc(i["category"])}</td>'
    )
    cells.append(f"<td>{esc(i['ar'])}</td>")
    owner = i["assignee"] or i["owner_transferred"]
    cells.append(f"<td>{esc(truncate(owner, 25))}</td>")
    cells.append(f"<td>{esc(truncate(i['owner_transferred'], 25))}</td>")
    cells.append(f"<td>{esc(truncate(i['reporter'], 25))}</td>")
    _dep_reason = i.get('dependency_reason') or ''
    if _dep_reason:
        cells.append(f'<td class="tip-cell" title="{esc(_dep_reason)}">{esc(truncate(i["dependency"], 30))}</td>')
    else:
        cells.append(f"<td>{esc(truncate(i['dependency'], 30))}</td>")
    _at_raw = i['action_tbd']
    _at_full = i.get('action_tbd_full') or _at_raw
    _at_disp = summarize_action_tbd(_at_raw, 80)
    cells.append(
        f'<td class="tip-cell action-tbd-cell" title="{esc(_at_full)}">{linkify_html(_at_disp)}</td>'
    )
    preview = esc(truncate(i["fix_approach"], 80))
    details_path = RESULT_DIR / "details" / f"{iid}.md"
    if details_path.exists():
        detail_link = (
            f'<a href="#" class="fix-modal-link" data-iid="{iid}" '
            f'data-title="Issue #{iid} \u2014 Details" '
            f'data-github="https://github.com/{REPO}/issues/{iid}">\u2192 details</a>'
        )
        cells.append(f"<td>{preview}<br>{detail_link}</td>")
    else:
        cells.append(f"<td>{preview}</td>")
    if i["dup_ids"]:
        dup_links = ", ".join(
            f'<a href="https://github.com/{REPO}/issues/{d}">#{d}</a>'
            for d in i["dup_ids"]
        )
        cells.append(f"<td>{dup_links}</td>")
    else:
        cells.append("<td></td>")
    cells.append(f"<td>{esc(truncate(i['labels'], 40))}</td>")
    cells.append(f"<td>{esc(i['created'][:10] if i['created'] else '')}</td>")

    return f'<tr {" ".join(attrs)}>{"".join(cells)}</tr>'


HEADERS = [
    "Done", "Issue", "Title", "Priority", "Category", "AR",
    "Owner", "Owner Transferred", "Reporter", "Dependency", "action_TBD",
    "Fix Approach", "Duplicates", "Labels", "Created",
]


def render_table_header() -> str:
    ths = []
    for h in HEADERS:
        cls = ' class="done-col"' if h == "Done" else ""
        ths.append(f"<th{cls}>{h}</th>")
    return "<thead><tr>" + "".join(ths) + "</tr></thead>"


def _biweekly_boundaries() -> list[tuple[str, datetime]]:
    """Return the last 12 bi-weekly boundaries (end-of-period dates) up to now."""
    from datetime import timedelta
    boundaries = []
    current = TODAY
    year, week, _ = current.isocalendar()
    biweek = (week - 1) // 2
    for i in range(12):
        bw = biweek - i
        yr = year
        while bw < 0:
            yr -= 1
            bw += 26
        ww_start = bw * 2 + 1
        label = f"{yr}-WW{ww_start:02d}"
        from datetime import date
        jan1 = date(yr, 1, 1)
        day_offset = (ww_start - 1) * 7 - jan1.weekday() + 13
        boundary_date = datetime(yr, 1, 1, tzinfo=timezone.utc) + timedelta(days=day_offset)
        boundaries.append((label, boundary_date))
    return list(reversed(boundaries))


def _open_cases_at_boundaries(issues: list[dict], boundaries: list[tuple[str, datetime]],
                              key: str, top_n: int = 5,
                              exclude_vals: set[str] | None = None,
                              exclude_prefixes: tuple[str, ...] = ()) -> tuple[list[str], dict[str, list[int]], list[str]]:
    """For each boundary, count issues open at that time (created <= boundary
    and not yet closed) grouped by key.
    Returns (labels, {value: [count_per_boundary]}, top_keys)."""
    def _excluded(v: str) -> bool:
        if exclude_vals and v in exclude_vals:
            return True
        if exclude_prefixes:
            vl = v.lower()
            if any(vl.startswith(p.lower()) for p in exclude_prefixes):
                return True
        return False

    all_vals: Counter = Counter()
    for i in issues:
        v = i.get(key, "") or "(none)"
        if _excluded(v):
            continue
        all_vals[v] += 1
    top_keys = [k for k, _ in all_vals.most_common(top_n)]

    labels = [b[0] for b in boundaries]
    series: dict[str, list[int]] = {k: [] for k in top_keys}
    for _, boundary_dt in boundaries:
        counts: Counter = Counter()
        for i in issues:
            dt = i["created_dt"]
            cdt = i.get("closed_dt")
            v = i.get(key, "") or "(none)"
            if _excluded(v):
                continue
            if dt and dt <= boundary_dt and (cdt is None or cdt > boundary_dt) and v in top_keys:
                counts[v] += 1
        for k in top_keys:
            series[k].append(counts.get(k, 0))
    return labels, series, top_keys


def _bar_chart_svg(title: str, data: list[tuple[str, int]], width: int = 500, bar_h: int = 18,
                   colors: dict[str, str] | None = None,
                   filter_dim: str | None = None,
                   show_hidden_on_filter: bool = True,
                   stale_only: bool = False) -> str:
    """Render a horizontal bar chart as inline SVG."""
    if not data:
        return ""
    max_val = max(v for _, v in data) or 1
    left_margin = 140
    chart_w = width - left_margin - 40
    height = len(data) * (bar_h + 6) + 40
    lines = [f'<svg class="bar-chart" data-filter-dim="{esc(filter_dim or "")}" '
             f'data-chart-width="{chart_w}" data-left-margin="{left_margin}" '
             f'data-stale-only="{1 if stale_only else 0}" '
             f'width="{width}" height="{height}" xmlns="http://www.w3.org/2000/svg" '
             f'style="font-family:sans-serif;font-size:11px;">']
    lines.append(f'<text x="{width // 2}" y="14" text-anchor="middle" '
                 f'font-weight="bold" font-size="12">{esc(title)}</text>')
    y = 28
    palette = ["#4e79a7", "#f28e2b", "#e15759", "#76b7b2", "#59a14f",
               "#edc948", "#b07aa1", "#ff9da7", "#9c755f", "#bab0ac",
               "#1f77b4", "#2ca02c", "#d62728", "#9467bd", "#8c564b"]
    for idx, (label, val) in enumerate(data):
        bar_w = int(val / max_val * chart_w) if max_val else 0
        color = (colors or {}).get(label, palette[idx % len(palette)])
        if filter_dim:
            lines.append(f'<g class="chart-filter-target chart-bar" role="button" tabindex="0" '
                         f'data-filter-dim="{esc(filter_dim)}" data-filter-value="{esc(label)}" '
                         f'data-show-hidden="{1 if show_hidden_on_filter else 0}">')
            lines.append(f'<title>Filter table: {esc(label)}</title>')
        lines.append(f'<text x="{left_margin - 4}" y="{y + bar_h // 2 + 4}" '
                     f'text-anchor="end">{esc(label[:20])}</text>')
        lines.append(f'<rect class="chart-bar-rect" x="{left_margin}" y="{y}" '
                     f'width="{bar_w}" height="{bar_h}" fill="{color}" rx="2"/>')
        lines.append(f'<text class="chart-bar-value" x="{left_margin + bar_w + 4}" '
                     f'y="{y + bar_h // 2 + 4}">{val}</text>')
        if filter_dim:
            lines.append('</g>')
        y += bar_h + 6
    lines.append("</svg>")
    return "\n".join(lines)


def _line_chart_svg(title: str, labels: list[str], series: dict[str, list[int]],
                    top_keys: list[str], width: int = 700, height: int = 260,
                    fixed_colors: dict[str, str] | None = None,
                    filter_dim: str | None = None,
                    legend_position: str = "right",
                    show_hidden_on_filter: bool = True,
                    boundary_dates: list[str] | None = None) -> str:
    """Render a multi-line trend chart as inline SVG with data point labels."""
    if not labels or not series:
        return ""
    palette = ["#4e79a7", "#f28e2b", "#e15759", "#76b7b2", "#59a14f",
               "#edc948", "#b07aa1", "#ff9da7", "#9c755f", "#bab0ac",
               "#1f77b4", "#2ca02c", "#d62728", "#9467bd", "#8c564b"]
    colors = dict(fixed_colors) if fixed_colors else {}
    for i, k in enumerate(top_keys):
        if k not in colors:
            colors[k] = palette[i % len(palette)]

    legend_w = max(len(k[:20]) for k in top_keys) * 7 + 30 if top_keys and legend_position == "right" else 0
    total_w = width + legend_w
    margin_l, margin_r, margin_t, margin_b = 50, 30, 30, 90
    plot_w = width - margin_l - margin_r
    plot_h = height - margin_t - margin_b
    n = len(labels)

    max_val = max((max(vals) for vals in series.values() if vals), default=1) or 1

    boundary_data = esc(json.dumps(boundary_dates or []))
    lines = [f'<svg class="line-chart" data-filter-dim="{esc(filter_dim or "")}" '
             f'data-max-val="{max_val}" data-boundaries="{boundary_data}" '
             f'data-margin-l="{margin_l}" data-margin-t="{margin_t}" '
             f'data-plot-w="{plot_w}" data-plot-h="{plot_h}" '
             f'width="{total_w}" height="{height}" '
             f'xmlns="http://www.w3.org/2000/svg" '
             f'style="font-family:sans-serif;font-size:10px;">']
    lines.append(f'<text x="{total_w // 2}" y="16" text-anchor="middle" '
                 f'font-weight="bold" font-size="12">{esc(title)}</text>')

    for gi in range(5):
        gy = margin_t + plot_h - int(gi / 4 * plot_h)
        gv = int(gi / 4 * max_val)
        lines.append(f'<line x1="{margin_l}" y1="{gy}" x2="{margin_l + plot_w}" y2="{gy}" '
                     f'stroke="#eee" stroke-width="1"/>')
        lines.append(f'<text class="chart-y-label" data-grid-index="{gi}" x="{margin_l - 6}" y="{gy + 4}" text-anchor="end" '
                     f'fill="#999">{gv}</text>')

    for bi, label in enumerate(labels):
        x = margin_l + int(bi / max(n - 1, 1) * plot_w) if n > 1 else margin_l + plot_w // 2
        y = margin_t + plot_h + 14
        display_label = label.split("-")[1] if "-" in label else label
        lines.append(f'<text x="{x}" y="{y}" text-anchor="end" fill="#666" '
                     f'transform="rotate(-45 {x} {y})">{esc(display_label)}</text>')

    for k in top_keys:
        vals = series[k]
        points = []
        for bi, v in enumerate(vals):
            x = margin_l + int(bi / max(n - 1, 1) * plot_w) if n > 1 else margin_l + plot_w // 2
            y = margin_t + plot_h - int(v / max_val * plot_h)
            points.append(f"{x},{y}")
        polyline = " ".join(points)
        if filter_dim:
            lines.append(f'<g class="chart-filter-target chart-series" role="button" tabindex="0" '
                         f'data-filter-dim="{esc(filter_dim)}" data-filter-value="{esc(k)}" '
                         f'data-show-hidden="{1 if show_hidden_on_filter else 0}">')
            lines.append(f'<title>Filter table: {esc(k)}</title>')
            lines.append(f'<polyline class="chart-filter-hit" points="{polyline}" fill="none" '
                         f'stroke="#000" stroke-width="12" stroke-linejoin="round"/>')
        lines.append(f'<polyline class="chart-line-main" points="{polyline}" fill="none" stroke="{colors[k]}" '
                     f'stroke-width="2" stroke-linejoin="round"/>')
        for bi, v in enumerate(vals):
            x = margin_l + int(bi / max(n - 1, 1) * plot_w) if n > 1 else margin_l + plot_w // 2
            y = margin_t + plot_h - int(v / max_val * plot_h)
            lines.append(f'<circle class="chart-point" cx="{x}" cy="{y}" r="3" fill="{colors[k]}"/>')
            if bi == n - 1:
                lines.append(f'<text class="chart-line-value" x="{x + 6}" y="{y + 4}" fill="{colors[k]}" '
                             f'font-size="9" font-weight="bold">{v}</text>')
        if filter_dim:
            lines.append('</g>')

    if legend_position == "bottom":
        lx = margin_l
        ly = height - 22
        for k in top_keys:
            if filter_dim:
                lines.append(f'<g class="chart-filter-target" role="button" tabindex="0" '
                             f'data-filter-dim="{esc(filter_dim)}" data-filter-value="{esc(k)}" '
                             f'data-show-hidden="{1 if show_hidden_on_filter else 0}">')
                lines.append(f'<title>Filter table: {esc(k)}</title>')
            lines.append(f'<rect x="{lx}" y="{ly}" width="12" height="4" fill="{colors[k]}" rx="1"/>')
            lines.append(f'<text x="{lx + 16}" y="{ly + 5}" fill="#333">{esc(k[:20])}</text>')
            if filter_dim:
                lines.append('</g>')
            lx += len(k[:20]) * 7 + 34
    else:
        lx = width + 5
        ly = margin_t + 10
        for k in top_keys:
            if filter_dim:
                lines.append(f'<g class="chart-filter-target" role="button" tabindex="0" '
                             f'data-filter-dim="{esc(filter_dim)}" data-filter-value="{esc(k)}" '
                             f'data-show-hidden="{1 if show_hidden_on_filter else 0}">')
                lines.append(f'<title>Filter table: {esc(k)}</title>')
            lines.append(f'<rect x="{lx}" y="{ly}" width="12" height="4" fill="{colors[k]}" rx="1"/>')
            lines.append(f'<text x="{lx + 16}" y="{ly + 5}" fill="#333">{esc(k[:20])}</text>')
            if filter_dim:
                lines.append('</g>')
            ly += 16

    lines.append("</svg>")
    return "\n".join(lines)


def _build_charts(issues: list[dict], prio_counts, cat_counts, ar_counts,
                  closed_trend: list[dict] | None = None) -> str:
    parts = []
    parts.append('<div class="charts-section">')

    boundaries = _biweekly_boundaries()
    boundary_dates = [dt.strftime("%Y-%m-%d") for _, dt in boundaries]

    trend_issues = issues + (closed_trend or [])

    prio_colors = {"P0": "#dc3545", "P1": "#fd7e14", "P2": "#ffc107", "P3": "#28a745"}
    labels, series, keys = _open_cases_at_boundaries(trend_issues, boundaries, "priority", top_n=4)
    parts.append(_line_chart_svg("Open Cases Trend: Priority", labels, series, keys,
                                 width=800, height=320, fixed_colors=prio_colors,
                                 filter_dim="priority", legend_position="bottom",
                                 boundary_dates=boundary_dates))

    labels, series, keys = _open_cases_at_boundaries(trend_issues, boundaries, "category",
                                                     top_n=len(cat_counts))
    parts.append(_line_chart_svg("Open Cases Trend: Category", labels, series, keys,
                                 width=900, height=400, filter_dim="category",
                                 boundary_dates=boundary_dates))

    labels, series, keys = _open_cases_at_boundaries(trend_issues, boundaries, "dependency", top_n=5,
                                                     exclude_vals={"(none)"},
                                                     exclude_prefixes=("SYCL kernel:",))
    parts.append(_line_chart_svg("Open Cases Trend: Dependency", labels, series, keys,
                                 width=700, height=320, filter_dim="dependency",
                                 boundary_dates=boundary_dates))

    labels, series, keys = _open_cases_at_boundaries(trend_issues, boundaries, "test_module",
                                                     top_n=5, exclude_vals={"(none)"})
    parts.append(_line_chart_svg("Open Cases Trend: Test Module", labels, series, keys,
                                 width=700, height=320, filter_dim="test_module",
                                 boundary_dates=boundary_dates))

    assignee_counts = Counter()
    for i in issues:
        for owner in split_owner_tokens(i["assignee"], i["owner_transferred"]):
            assignee_counts[owner] += 1
    assignee_data = assignee_counts.most_common(15)
    parts.append(_bar_chart_svg("Distribution: Assignee (top 15)", assignee_data,
                                filter_dim="assignee"))

    stale_assignee_counts = Counter()
    for i in issues:
        if i["is_stale"]:
            for owner in split_owner_tokens(i["assignee"], i["owner_transferred"]):
                stale_assignee_counts[owner] += 1
    stale_assignee_data = stale_assignee_counts.most_common(15)
    parts.append(_bar_chart_svg("Distribution: Assignee — Open Response >1 Week (top 15)",
                                stale_assignee_data, width=600, filter_dim="assignee",
                                show_hidden_on_filter=False, stale_only=True))

    ar_data = sorted(ar_counts.items(), key=lambda x: -x[1])
    parts.append(_bar_chart_svg("Distribution: AR", ar_data, filter_dim="ar"))

    parts.append("</div>")
    return "\n".join(parts)


def build_html(issues: list[dict], closed_trend: list[dict] | None = None) -> str:
    by_prio: dict[str, list[dict]] = defaultdict(list)
    for i in issues:
        p = i["priority"] if i["priority"] in PRIO_ORDER else "Other"
        by_prio[p].append(i)

    for p in by_prio:
        by_prio[p].sort(key=lambda x: (x["category"], x["id"]))

    total = len(issues)
    ar_counts: Counter = Counter()
    for i in issues:
        if not i["ar"]:
            continue
        for tok in (t.strip() for t in i["ar"].split(";")):
            if tok:
                ar_counts[tok] += 1
    prio_counts = Counter(i["priority"] for i in issues)
    cat_counts = Counter(i["category"] for i in issues)
    dup_count = sum(1 for i in issues if i["is_dup"])

    body_parts = []

    body_parts.append(f"<h1>XPU Ops Bug Scrub — Highlight Report</h1>")
    body_parts.append(f'<p class="meta">Repository: <code>{REPO}</code> &middot; '
                      f'Generated: {TODAY.strftime("%Y-%m-%d")} &middot; '
                      f'Total issues: {total}</p>')

    body_parts.append("<h2>Summary</h2>")
    body_parts.append('<div class="summary-grid">')
    for p in PRIO_ORDER:
        count = prio_counts.get(p, 0)
        body_parts.append(
            f'<div class="summary-card priority-{p.lower()}" data-summary-priority="{esc(p)}">'
            f'<div class="sc-num" data-total="{count}">{count}/{count}</div>'
            f'<div class="sc-label">{p}</div></div>'
        )
    body_parts.append("</div>")

    body_parts.append(_build_charts(issues, prio_counts, cat_counts, ar_counts, closed_trend))

    body_parts.append('<h2 id="sec-index">Index</h2>')
    body_parts.append("<ul>")
    for p in PRIO_ORDER:
        cnt = len(by_prio.get(p, []))
        body_parts.append(f'<li><a data-index-priority="{esc(p)}" href="#sec-{p.lower()}">{p} ({cnt} issues)</a></li>')
    if by_prio.get("Other"):
        body_parts.append(f'<li><a href="#sec-other">Other ({len(by_prio["Other"])} issues)</a></li>')
    body_parts.append("</ul>")

    all_prios = [p for p in PRIO_ORDER if p in by_prio]
    if "Other" in by_prio:
        all_prios.append("Other")

    for p in all_prios:
        group = by_prio[p]
        anchor = p.lower()
        body_parts.append(f'<h2 id="sec-{anchor}" data-priority-title="{esc(p)}" '
                          f'data-title-prefix="{esc(p)}">{p} &mdash; {len(group)} issues</h2>')

        # Sub-group by category
        by_cat: dict[str, list[dict]] = defaultdict(list)
        for i in group:
            by_cat[i["category"] or "(blank)"].append(i)

        for cat in sorted(by_cat.keys()):
            cat_issues = by_cat[cat]
            cat_anchor = f"sec-{anchor}-{cat.lower().replace(' ', '-').replace('/', '-')}"

            # Split into stale (>1 week open request) and non-stale
            stale = [i for i in cat_issues if i["is_stale"]]
            non_stale = [i for i in cat_issues if not i["is_stale"]]

            body_parts.append(
                f'<h3 id="{esc(cat_anchor)}" data-category-title="{esc(cat)}" '
                f'data-title-prefix="{esc(cat)}">{esc(cat)} &middot; {len(cat_issues)} issues</h3>'
            )

            table_id = f"tbl-{anchor}-{cat.lower().replace(' ', '-').replace('/', '-')}"
            body_parts.append(f'<table class="ar-table" id="{esc(table_id)}">')
            body_parts.append(render_table_header())
            body_parts.append("<tbody>")

            # Stale rows shown by default
            for i in stale:
                body_parts.append(render_issue_row(i))

            # Non-stale rows hidden by default
            for i in non_stale:
                row_html = render_issue_row(i)
                # Add a class to hide by default
                row_html = row_html.replace("<tr ", '<tr class="more-hidden" ', 1)
                body_parts.append(row_html)

            body_parts.append("</tbody></table>")

            # "more" link if there are non-stale rows
            if non_stale:
                body_parts.append(
                    f'<a href="#" class="more-link" data-table="{esc(table_id)}">'
                    f"Show {len(non_stale)} more (no open request &gt;1 week)</a>"
                )

    body_parts.append("""
<h2 id="sec-tips">Tips: Classification Rules</h2>
<div class="tips-section">
<h3>Priority</h3>
<ul>
<li><b>P0</b> — Crash / segfault / build failure / &gt;5% perf regression / custom model blocker</li>
<li><b>P1</b> — UT &gt;6 failures or regression / accuracy regression / hang</li>
<li><b>P2</b> — Benchmark issue / 1-6 UT failures / feature gap / correctness issue</li>
<li><b>P3</b> — Enhancement / minor issue / docs / nice-to-have</li>
</ul>
<h3>Category (first-match priority)</h3>
<ol>
<li>Distributed</li>
<li>Flash Attention</li>
<li>Inductor</li>
<li>TorchAO</li>
<li>Sparse</li>
<li>Torch Ops — gemm / eltwise / reduction / others</li>
<li>Torch Runtime</li>
<li>Others</li>
</ol>
<h3>Open Request (&gt;1 week)</h3>
<p>The <code>action_TBD</code> field is generated by analyzing related issue and PR
comments, status, and activity. When a pending action item (e.g., waiting for
reply, waiting for PR review, waiting for upstream fix) has been open for more
than one week, it is tagged with <code>(&gt;1 week)</code>. Issues with at least one
such tag are shown by default as they need attention; other issues are hidden
behind the "Show hidden issues" checkbox.</p>
</div>
""")

    return "\n".join(body_parts)


CSS = """
:root {
  --bg: #f8f9fa; --fg: #212529; --muted: #6c757d;
  --border: #dee2e6; --accent: #0066cc; --done-bg: #e9ecef; --done-fg: #adb5bd;
  --filter-bg: #ffffff; --hl: #fff3cd;
}
* { box-sizing: border-box; }
body { font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
  font-size: 13px; line-height: 1.5; color: var(--fg); background: var(--bg);
  margin: 0; padding: 0; }
h1, h2, h3, h4 { line-height: 1.2; margin-top: 1.5em; }
h1 { font-size: 1.8em; } h2 { font-size: 1.4em; border-bottom: 2px solid var(--border); padding-bottom: .3em; }
h3 { font-size: 1.15em; } h4 { font-size: 1em; color: var(--muted); }
a { color: var(--accent); text-decoration: none; }
a:hover { text-decoration: underline; }
code { background: #f1f3f5; padding: 1px 4px; border-radius: 3px; font-size: .9em; }
.meta { color: var(--muted); font-size: .9em; margin: .3em 0; }
.content { max-width: 100%; padding: 1em 1.5em 4em; }

/* Summary cards */
.summary-grid { display: flex; gap: 1em; margin: .5em 0 1em; flex-wrap: wrap; }
.summary-card { border: 1px solid var(--border); border-radius: 6px; padding: .6em 1.2em;
  text-align: center; min-width: 90px; background: white; }
.summary-card .sc-num { font-size: 1.8em; font-weight: 700; }
.summary-card .sc-label { font-size: .9em; color: var(--muted); }
.priority-p0 { border-left: 4px solid #dc3545; }
.priority-p1 { border-left: 4px solid #fd7e14; }
.priority-p2 { border-left: 4px solid #ffc107; }
.priority-p3 { border-left: 4px solid #28a745; }

/* Sticky filter bar */
.filter-bar {
  position: sticky; top: 0; z-index: 100;
  background: var(--filter-bg); border-bottom: 1px solid var(--border);
  padding: .6em 1em; display: flex; flex-wrap: wrap; gap: .8em; align-items: center;
  box-shadow: 0 1px 3px rgba(0,0,0,.06);
}
.filter-bar > label { font-size: 12px; color: var(--muted); display: flex; flex-direction: column; gap: 2px; }
.filter-bar select, .filter-bar input[type=text], .filter-bar input[type=date] {
  font-size: 13px; padding: 3px 6px; border: 1px solid var(--border); border-radius: 3px;
  background: white; min-width: 120px;
}
.filter-bar input[type=text] { min-width: 200px; }
.filter-bar button {
  font-size: 12px; padding: 4px 10px; border: 1px solid var(--border); border-radius: 3px;
  background: white; cursor: pointer;
}
.filter-bar button:hover { background: #f1f3f5; }
.filter-bar button.active { background: var(--accent); color: white; border-color: var(--accent); }
.filter-bar .stats { margin-left: auto; color: var(--muted); font-size: 12px; }

/* Tables */
table.ar-table { border-collapse: collapse; width: 100%; margin: .5em 0 .3em; font-size: 12px; }
table.ar-table th, table.ar-table td { border: 1px solid var(--border); padding: 4px 6px;
  vertical-align: top; text-align: left; }
table.ar-table thead th { background: #e7eaf0; position: sticky; top: 56px; z-index: 50; }
table.ar-table td.done-col, table.ar-table th.done-col { width: 40px; text-align: center; }
table.ar-table tr:nth-child(even) td { background: #fafbfc; }
table.ar-table tr.done td { background: var(--done-bg) !important; color: var(--done-fg); }
table.ar-table tr.done td a { color: var(--done-fg); }
table.ar-table tr.done td:not(.done-col) { text-decoration: line-through; }
table.ar-table tr.hidden { display: none; }
table.ar-table tr.more-hidden { display: none; }
table.ar-table tr.more-shown { display: table-row; }
table.ar-table tr.more-shown.hidden { display: none; }

/* Tooltip for priority/category cells */
.tip-cell { cursor: help; }
.tip-cell[title]:hover { background: var(--hl) !important; }

/* More link */
.more-link { display: inline-block; margin: 0 0 1em 0.5em; font-size: 12px; color: var(--accent); cursor: pointer; }
.more-link.expanded { color: var(--muted); }

/* Section hidden by filter */
.section-hidden { display: none !important; }

/* Multi-select dropdown */
.ms-dd { position: relative; display: inline-block; }
.ms-dd-btn {
  font-size: 13px; padding: 3px 24px 3px 6px; border: 1px solid var(--border);
  border-radius: 3px; background: white; cursor: pointer; min-width: 140px;
  text-align: left; position: relative; white-space: nowrap; overflow: hidden;
  text-overflow: ellipsis; max-width: 220px;
}
.ms-dd-btn::after {
  content: "\\25BE"; position: absolute; right: 6px; top: 50%;
  transform: translateY(-50%); color: var(--muted); font-size: 10px;
}
.ms-dd-panel {
  display: none; position: absolute; top: calc(100% + 2px); left: 0;
  background: white; border: 1px solid var(--border); border-radius: 3px;
  box-shadow: 0 2px 8px rgba(0,0,0,.12); z-index: 200;
  max-height: 320px; overflow-y: auto; min-width: 200px; padding: 4px 0;
}
.ms-dd.open .ms-dd-panel { display: block; }
.ms-dd-search {
  width: calc(100% - 12px); margin: 4px 6px; padding: 3px 5px;
  border: 1px solid var(--border); border-radius: 3px; font-size: 12px;
}
.ms-dd-item {
  display: flex; flex-direction: row; align-items: center; gap: 4px;
  padding: 3px 10px; cursor: pointer; font-size: 12px; user-select: none;
}
.ms-dd-item:hover { background: #f1f3f5; }
.ms-dd-item input[type=checkbox] { margin: 0; flex: 0 0 auto; width: 14px; height: 14px; accent-color: var(--accent); }
.ms-dd-item .label { flex: 0 1 auto; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; max-width: 200px; }
.ms-dd-item .count { flex: 0 0 auto; color: var(--muted); font-size: 11px; background: #f1f3f5;
  padding: 0 6px; border-radius: 8px; min-width: 22px; text-align: center; }
.ms-dd-item.none-opt .label { font-style: italic; color: var(--muted); }
.ms-dd-item.zero-count { opacity: .45; }
.ms-dd-actions {
  display: flex; gap: 4px; padding: 4px 6px; border-top: 1px solid var(--border);
  margin-top: 2px; position: sticky; bottom: 0; background: white;
}
.ms-dd-actions button { flex: 1; font-size: 11px; padding: 2px 4px; border: 1px solid var(--border);
  border-radius: 3px; background: white; cursor: pointer; }
.ms-dd-actions button:hover { background: #f1f3f5; }

/* Charts */
.charts-section { margin: 1em 0 2em; display: flex; flex-wrap: wrap; gap: 1.5em; }
.charts-section svg { background: white; border: 1px solid var(--border); border-radius: 4px; padding: 8px; }
.chart-filter-target { cursor: pointer; }
.chart-filter-target:hover text { font-weight: 700; text-decoration: underline; }
.chart-filter-target:hover polyline:not(.chart-filter-hit), .chart-filter-target:hover rect { filter: brightness(.9); }
.chart-filter-hit { opacity: 0; }
.chart-filter-target.filtered-out { opacity: .18; }
.chart-filter-target.filtered-zero { opacity: .28; }

/* Custom tooltip popup (replaces native title) - plain white */
.tip-popup {
  display: none; position: fixed; z-index: 9999; max-width: 560px;
  background: #ffffff;
  color: #1a1a1a;
  border: 1px solid #ccc;
  border-radius: 4px;
  padding: 8px 12px;
  font-size: 12.5px;
  line-height: 1.5;
  box-shadow: 0 2px 8px rgba(0,0,0,.15);
  white-space: pre-wrap;
  word-wrap: break-word;
  pointer-events: none;
}
.tip-popup.visible { display: block; }

/* In-page modal (for issue iframe + fix-approach markdown) */
#modal-overlay {
  display: none; position: fixed; top: 0; left: 0; right: 0; bottom: 0;
  background: rgba(0,0,0,.5); z-index: 10000;
  align-items: center; justify-content: center;
}
#modal-overlay.visible { display: flex; }
#modal-frame {
  width: 85vw; height: 85vh; max-width: 1200px;
  background: white; border-radius: 6px;
  box-shadow: 0 10px 30px rgba(0,0,0,.3);
  display: flex; flex-direction: column;
  overflow: hidden;
}
#modal-header {
  display: flex; align-items: center;
  padding: 8px 14px; border-bottom: 1px solid #eee;
  font-weight: 600; font-size: 14px;
  gap: 12px;
}
#modal-title { flex: 1 1 auto; min-width: 0; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
#modal-close {
  cursor: pointer; background: transparent; border: 0;
  font-size: 22px; line-height: 1; padding: 0 4px; color: #555;
}
#modal-close:hover { color: #000; }
#modal-body { flex: 1 1 auto; overflow: auto; padding: 0; }
#modal-body.markdown { padding: 14px 20px; }
#modal-body iframe { width: 100%; height: 100%; border: 0; display: block; }
#modal-body.markdown pre { background: #f6f8fa; padding: 10px; border-radius: 4px; overflow-x: auto; }
#modal-body.markdown code { background: #f1f3f5; padding: 1px 4px; border-radius: 3px; font-family: ui-monospace, SFMono-Regular, Menlo, monospace; font-size: 12.5px; }
#modal-body.markdown pre code { background: transparent; padding: 0; }
#modal-body.markdown table { border-collapse: collapse; margin: .6em 0; }
#modal-body.markdown th, #modal-body.markdown td { border: 1px solid #ddd; padding: 4px 8px; }
.tips-section { background: #fff; border: 1px solid var(--border); border-radius: 6px;
  padding: 1em 1.5em; margin: 1em 0; }
.tips-section h3 { margin-top: .8em; }
.tips-section ul, .tips-section ol { margin: .3em 0; padding-left: 1.5em; }
.tips-section li { margin: .2em 0; }
"""

JS = """
// ---- Done checkbox persistence ----
const STORAGE_KEY_PREFIX = 'bug_scrub_done_';
function loadDoneState() {
  document.querySelectorAll('input.ar-done').forEach(cb => {
    const id = cb.dataset.issue;
    if (!id) return;
    if (localStorage.getItem(STORAGE_KEY_PREFIX + id) === '1') {
      cb.checked = true;
      cb.closest('tr').classList.add('done');
    }
  });
}
function onDoneToggle(e) {
  const cb = e.target;
  const id = cb.dataset.issue;
  if (!id) return;
  const tr = cb.closest('tr');
  if (cb.checked) {
    localStorage.setItem(STORAGE_KEY_PREFIX + id, '1');
    tr.classList.add('done');
  } else {
    localStorage.removeItem(STORAGE_KEY_PREFIX + id);
    tr.classList.remove('done');
  }
}

// ---- "More" toggle ----
document.addEventListener('click', e => {
  const link = e.target.closest('.more-link');
  if (!link) return;
  e.preventDefault();
  const tableId = link.dataset.table;
  const table = document.getElementById(tableId);
  if (!table) return;
  const expanded = link.classList.toggle('expanded');
  table.querySelectorAll('tr.more-hidden, tr.more-shown').forEach(tr => {
    if (expanded) {
      tr.classList.remove('more-hidden');
      tr.classList.add('more-shown');
    } else {
      tr.classList.remove('more-shown');
      tr.classList.add('more-hidden');
    }
  });
  link.textContent = expanded
    ? link.textContent.replace('Show', 'Hide')
    : link.textContent.replace('Hide', 'Show');
});

// ---- Filter infrastructure ----
const NONE_TOKEN = '(none)';
const MULTI_VALUE_DIMS = new Set(['ar', 'assignee', 'owner_transferred']);
const MULTI_SPLIT_RE = /[,;|]/;

const FILTER_DIMS = ['ar', 'assignee', 'priority', 'category', 'milestone', 'dependency', 'test_module'];
const FILTER_LABELS = {
  ar: 'AR', assignee: 'Assignee', priority: 'Priority', category: 'Category',
  milestone: 'Milestone', dependency: 'Dependency', test_module: 'Test Module'
};
const SELECTED = Object.fromEntries(FILTER_DIMS.map(d => [d, new Set()]));

// Closed issues kept out of the table but folded into the open-over-time trend.
// Each carries a row-like `dataset` so the same filter predicates apply.
const TREND_CLOSED = (window.TREND_CLOSED_DATA || []);

function tokensFor(dim, raw) {
  const r = (raw || '').trim();
  if (!r) return [NONE_TOKEN];
  if (MULTI_VALUE_DIMS.has(dim)) {
    const out = r.split(MULTI_SPLIT_RE).map(s => s.trim()).filter(Boolean);
    return out.length ? Array.from(new Set(out)) : [NONE_TOKEN];
  }
  return [r];
}

function collectValues(dim) {
  const counts = new Map();
  const sourceDims = (dim === 'assignee') ? ['assignee', 'owner_transferred'] : [dim];
  document.querySelectorAll('table.ar-table tbody tr[data-issue]').forEach(tr => {
    const tokenSet = new Set();
    for (const sd of sourceDims) {
      for (const tok of tokensFor(sd, tr.dataset[sd] || tr.dataset[sd.replace('_', '-')])) tokenSet.add(tok);
    }
    if (sourceDims.length > 1 && tokenSet.size > 1) tokenSet.delete(NONE_TOKEN);
    for (const tok of tokenSet) {
      // Hide per-file SYCL kernel entries from the Dependency filter dropdown -
      // they are noise (hundreds of distinct values). Rows are still filterable
      // via the search box if needed.
      if (dim === 'dependency' && /^sycl kernel:/i.test(tok)) continue;
      counts.set(tok, (counts.get(tok) || 0) + 1);
    }
  });
  return Array.from(counts.entries()).sort((a, b) => {
    if (a[0] === NONE_TOKEN) return 1;
    if (b[0] === NONE_TOKEN) return -1;
    return a[0].localeCompare(b[0], undefined, {numeric: true});
  });
}

function buildMultiSelect(dim, items) {
  const dd = document.createElement('div');
  dd.className = 'ms-dd'; dd.dataset.dim = dim;
  dd.dataset.totalOptions = String(items.length);

  const btn = document.createElement('button');
  btn.type = 'button'; btn.className = 'ms-dd-btn'; btn.textContent = '(all)';
  dd.appendChild(btn);

  const panel = document.createElement('div');
  panel.className = 'ms-dd-panel';
  const search = document.createElement('input');
  search.type = 'text'; search.className = 'ms-dd-search'; search.placeholder = 'filter...';
  panel.appendChild(search);

  const list = document.createElement('div');
  list.className = 'ms-dd-list';
  panel.appendChild(list);

  for (const [tok, count] of items) {
    const label = document.createElement('label');
    label.className = 'ms-dd-item' + (tok === NONE_TOKEN ? ' none-opt' : '');
    const cb = document.createElement('input');
    cb.type = 'checkbox'; cb.value = tok; cb.checked = true;
    SELECTED[dim].add(tok);
    cb.addEventListener('change', () => {
      if (cb.checked) SELECTED[dim].add(tok); else SELECTED[dim].delete(tok);
      updateButtonLabel(dd, dim); applyFilters();
    });
    const txt = document.createElement('span');
    txt.className = 'label'; txt.textContent = tok; txt.title = tok;
    const cnt = document.createElement('span');
    cnt.className = 'count'; cnt.textContent = count;
    label.appendChild(cb); label.appendChild(txt); label.appendChild(cnt);
    list.appendChild(label);
  }

  search.addEventListener('input', () => {
    const q = search.value.trim().toLowerCase();
    list.querySelectorAll('.ms-dd-item').forEach(it => {
      it.style.display = (!q || it.querySelector('.label').textContent.toLowerCase().includes(q)) ? '' : 'none';
    });
  });

  const actions = document.createElement('div');
  actions.className = 'ms-dd-actions';
  const allBtn = document.createElement('button'); allBtn.type = 'button'; allBtn.textContent = 'All';
  allBtn.addEventListener('click', () => {
    for (const [tok] of items) SELECTED[dim].add(tok);
    list.querySelectorAll('input[type=checkbox]').forEach(c => c.checked = true);
    updateButtonLabel(dd, dim); applyFilters();
  });
  const noneBtn = document.createElement('button'); noneBtn.type = 'button'; noneBtn.textContent = 'None';
  noneBtn.addEventListener('click', () => {
    SELECTED[dim].clear();
    list.querySelectorAll('input[type=checkbox]').forEach(c => c.checked = false);
    updateButtonLabel(dd, dim); applyFilters();
  });
  const closeBtn = document.createElement('button'); closeBtn.type = 'button'; closeBtn.textContent = 'Close';
  closeBtn.addEventListener('click', () => dd.classList.remove('open'));
  actions.appendChild(allBtn); actions.appendChild(noneBtn); actions.appendChild(closeBtn);
  panel.appendChild(actions);
  dd.appendChild(panel);

  btn.addEventListener('click', e => {
    e.stopPropagation();
    const wasOpen = dd.classList.contains('open');
    document.querySelectorAll('.ms-dd.open').forEach(d => d.classList.remove('open'));
    if (!wasOpen) dd.classList.add('open');
  });
  panel.addEventListener('click', e => e.stopPropagation());
  updateButtonLabel(dd, dim);
  return dd;
}

function updateButtonLabel(dd, dim) {
  const btn = dd.querySelector('.ms-dd-btn');
  const sel = SELECTED[dim];
  const total = parseInt(dd.dataset.totalOptions || '0', 10);
  if (sel.size === 0) { btn.textContent = '(none)'; btn.title = ''; }
  else if (sel.size === total) { btn.textContent = '(all)'; btn.title = ''; }
  else if (sel.size === 1) { const v = Array.from(sel)[0]; btn.textContent = v; btn.title = v; }
  else { btn.textContent = sel.size + ' of ' + total; btn.title = Array.from(sel).join(', '); }
}

function rowMatchesAllFiltersExcept(tr, skipDim) {
  const showHidden = document.getElementById('filter-show-hidden').checked;
  const isMoreRow = tr.classList.contains('more-hidden') || tr.classList.contains('more-shown');
  if (isMoreRow && !showHidden) return false;
  for (const dim of FILTER_DIMS) {
    if (dim === skipDim) continue;
    if (!rowMatchesDim(tr, dim, SELECTED[dim])) return false;
  }
  const search = (document.getElementById('filter-search').value || '').trim().toLowerCase();
  if (search && !(tr.dataset.search || '').includes(search)) return false;
  if (document.getElementById('filter-hide-done').checked && tr.classList.contains('done')) return false;
  if (showDupsOnly && tr.dataset.isDup !== '1') return false;
  const dateVal = document.getElementById('filter-date').value;
  if (dateVal) {
    const created = tr.dataset.created || '';
    if (created && created < dateVal) return false;
  }
  return true;
}

function updateFacetCounts() {
  const allRows = Array.from(document.querySelectorAll('table.ar-table tbody tr[data-issue]'));
  for (const dim of FILTER_DIMS) {
    const dd = document.getElementById('filter-' + dim);
    if (!dd) continue;
    const counts = new Map();
    for (const tr of allRows) {
      if (!rowMatchesAllFiltersExcept(tr, dim)) continue;
      for (const v of rowValuesForDim(tr, dim)) counts.set(v, (counts.get(v) || 0) + 1);
    }
    dd.querySelectorAll('.ms-dd-item').forEach(item => {
      const cb = item.querySelector('input[type=checkbox]');
      const count = counts.get(cb.value) || 0;
      item.querySelector('.count').textContent = String(count);
      item.classList.toggle('zero-count', count === 0);
    });
  }
}

function setSingleFilter(dim, value, showHiddenRows = true) {
  const dd = document.getElementById('filter-' + dim);
  if (!dd) return;
  SELECTED[dim].clear();
  SELECTED[dim].add(value);
  dd.querySelectorAll('input[type=checkbox]').forEach(cb => { cb.checked = (cb.value === value); });
  updateButtonLabel(dd, dim);
  const showHidden = document.getElementById('filter-show-hidden');
  if (showHidden) showHidden.checked = showHiddenRows;
  applyFilters();
  const firstVisible = document.querySelector('table.ar-table tbody tr[data-issue]:not(.hidden)');
  if (firstVisible) firstVisible.scrollIntoView({behavior: 'smooth', block: 'center'});
}

function initChartFilters() {
  document.querySelectorAll('.chart-filter-target').forEach(target => {
    const applyChartFilter = () => setSingleFilter(
      target.dataset.filterDim,
      target.dataset.filterValue,
      target.dataset.showHidden !== '0'
    );
    target.addEventListener('click', applyChartFilter);
    target.addEventListener('keydown', e => {
      if (e.key === 'Enter' || e.key === ' ') {
        e.preventDefault();
        applyChartFilter();
      }
    });
  });
}

function visibleIssueRows() {
  return Array.from(document.querySelectorAll('table.ar-table tbody tr[data-issue]:not(.hidden):not(.more-hidden)'));
}

function rowValuesForDim(tr, dim) {
  const sourceDims = (dim === 'assignee') ? ['assignee', 'owner_transferred'] : [dim];
  const vals = new Set();
  for (const sd of sourceDims) {
    const key = sd.replace('_', '-');
    const raw = tr.dataset[sd] || tr.dataset[key] || '';
    for (const t of tokensFor(sd, raw)) vals.add(t);
  }
  if (sourceDims.length > 1 && vals.size > 1) vals.delete(NONE_TOKEN);
  return vals;
}

function countVisibleRowsByDim(rows, dim, extraFilter) {
  const counts = new Map();
  for (const tr of rows) {
    if (extraFilter && !extraFilter(tr)) continue;
    for (const v of rowValuesForDim(tr, dim)) counts.set(v, (counts.get(v) || 0) + 1);
  }
  return counts;
}

function updateVisibleCounts(rows) {
  const priorityCounts = countVisibleRowsByDim(rows, 'priority');
  document.querySelectorAll('[data-summary-priority]').forEach(card => {
    const priority = card.dataset.summaryPriority;
    const num = card.querySelector('.sc-num');
    if (num) {
      const total = num.dataset.total || '0';
      num.textContent = String(priorityCounts.get(priority) || 0) + '/' + total;
    }
  });
  document.querySelectorAll('[data-index-priority]').forEach(link => {
    const priority = link.dataset.indexPriority;
    link.textContent = priority + ' (' + (priorityCounts.get(priority) || 0) + ' issues)';
  });
  document.querySelectorAll('table.ar-table').forEach(table => {
    const rowsInTable = table.querySelectorAll('tbody tr[data-issue]:not(.hidden):not(.more-hidden)').length;
    const h3 = table.previousElementSibling;
    if (h3 && h3.dataset.titlePrefix) {
      h3.innerHTML = h3.dataset.titlePrefix + ' &middot; ' + rowsInTable + ' issues';
    }
  });
  document.querySelectorAll('[data-priority-title]').forEach(h2 => {
    const priority = h2.dataset.priorityTitle;
    h2.innerHTML = h2.dataset.titlePrefix + ' &mdash; ' + (priorityCounts.get(priority) || 0) + ' issues';
  });
}

function updateBarCharts(rows) {
  document.querySelectorAll('svg.bar-chart[data-filter-dim]').forEach(svg => {
    const dim = svg.dataset.filterDim;
    if (!dim) return;
    const staleOnly = svg.dataset.staleOnly === '1';
    const counts = countVisibleRowsByDim(rows, dim, staleOnly ? tr => tr.dataset.isStale === '1' : null);
    const bars = Array.from(svg.querySelectorAll('.chart-bar'));
    const maxVal = Math.max(1, ...bars.map(g => counts.get(g.dataset.filterValue) || 0));
    const chartW = Number(svg.dataset.chartWidth || 1);
    const left = Number(svg.dataset.leftMargin || 0);
    for (const g of bars) {
      const val = counts.get(g.dataset.filterValue) || 0;
      const w = Math.round(val / maxVal * chartW);
      const rect = g.querySelector('.chart-bar-rect');
      const text = g.querySelector('.chart-bar-value');
      if (rect) rect.setAttribute('width', String(w));
      if (text) {
        text.textContent = String(val);
        text.setAttribute('x', String(left + w + 4));
      }
      g.classList.toggle('filtered-zero', val === 0);
    }
  });
}

function closedTrendMatches(obj) {
  for (const dim of FILTER_DIMS) {
    if (!rowMatchesDim(obj, dim, SELECTED[dim])) return false;
  }
  const search = (document.getElementById('filter-search').value || '').trim().toLowerCase();
  if (search && !(obj.dataset.search || '').includes(search)) return false;
  if (showDupsOnly && obj.dataset.isDup !== '1') return false;
  const dateVal = document.getElementById('filter-date').value;
  if (dateVal && obj.dataset.created && obj.dataset.created < dateVal) return false;
  return true;
}

function updateLineCharts(rows) {
  document.querySelectorAll('svg.line-chart[data-filter-dim]').forEach(svg => {
    const dim = svg.dataset.filterDim;
    if (!dim) return;
    const boundaries = JSON.parse(svg.dataset.boundaries || '[]');
    if (!boundaries.length) return;
    const marginL = Number(svg.dataset.marginL || 0);
    const marginT = Number(svg.dataset.marginT || 0);
    const plotW = Number(svg.dataset.plotW || 1);
    const plotH = Number(svg.dataset.plotH || 1);
    const seriesGroups = Array.from(svg.querySelectorAll('.chart-series'));
    const series = new Map();
    let maxVal = 1;
    for (const g of seriesGroups) {
      const value = g.dataset.filterValue;
      const vals = boundaries.map(boundary => {
        let n = rows.filter(tr => {
          const created = tr.dataset.created || '';
          return created && created <= boundary && rowValuesForDim(tr, dim).has(value);
        }).length;
        for (const obj of TREND_CLOSED) {
          const created = obj.dataset.created || '';
          const closed = obj.dataset.closed || '';
          if (created && created <= boundary && (!closed || closed > boundary)
              && closedTrendMatches(obj) && rowValuesForDim(obj, dim).has(value)) {
            n++;
          }
        }
        return n;
      });
      series.set(g, vals);
      maxVal = Math.max(maxVal, ...vals);
    }
    svg.querySelectorAll('.chart-y-label').forEach(label => {
      const idx = Number(label.dataset.gridIndex || 0);
      label.textContent = String(Math.floor(idx / 4 * maxVal));
    });
    for (const [g, vals] of series.entries()) {
      const points = vals.map((v, i) => {
        const x = marginL + Math.round(i / Math.max(boundaries.length - 1, 1) * plotW);
        const y = marginT + plotH - Math.round(v / maxVal * plotH);
        return [x, y];
      });
      const pointText = points.map(([x, y]) => x + ',' + y).join(' ');
      g.querySelectorAll('polyline').forEach(p => p.setAttribute('points', pointText));
      g.querySelectorAll('.chart-point').forEach((circle, i) => {
        const [x, y] = points[i];
        circle.setAttribute('cx', String(x));
        circle.setAttribute('cy', String(y));
      });
      const valText = g.querySelector('.chart-line-value');
      if (valText && points.length) {
        const [x, y] = points[points.length - 1];
        valText.textContent = String(vals[vals.length - 1] || 0);
        valText.setAttribute('x', String(x + 6));
        valText.setAttribute('y', String(y + 4));
      }
      g.classList.toggle('filtered-zero', (vals[vals.length - 1] || 0) === 0);
    }
  });
}

function updateDynamicReportState() {
  const rows = visibleIssueRows();
  updateVisibleCounts(rows);
  updateBarCharts(rows);
  // Trend is a global open-over-time metric: count every filter-passing row,
  // including ones paginated behind "show more" (.more-hidden), so the curve
  // reflects the full population rather than how many rows are expanded.
  const trendRows = Array.from(document.querySelectorAll('table.ar-table tbody tr[data-issue]:not(.hidden)'));
  updateLineCharts(trendRows);
}

// Tokens hidden from the dropdown for a given dim (see collectValues()).
// Rows whose ONLY tokens are hidden would otherwise be unreachable by any
// selection; treat them as always-matching for that dim so they stay
// filterable via other dimensions.
const HIDDEN_TOKEN_PREDICATES = {
  dependency: t => /^sycl kernel:/i.test(t),
};

function rowMatchesDim(tr, dim, selectedSet) {
  if (selectedSet.size === 0) return false;
  const sourceDims = (dim === 'assignee') ? ['assignee', 'owner_transferred'] : [dim];
  const rowTokens = new Set();
  for (const sd of sourceDims) {
    const key = sd.replace('_', '-');  // data attributes use dashes
    const raw = tr.dataset[sd] || tr.dataset[key] || '';
    for (const t of tokensFor(sd, raw)) rowTokens.add(t);
  }
  if (sourceDims.length > 1 && rowTokens.size > 1) rowTokens.delete(NONE_TOKEN);
  const hiddenPred = HIDDEN_TOKEN_PREDICATES[dim];
  if (hiddenPred) {
    let allHidden = rowTokens.size > 0;
    for (const t of rowTokens) { if (!hiddenPred(t)) { allHidden = false; break; } }
    if (allHidden) return true;
  }
  for (const t of rowTokens) { if (selectedSet.has(t)) return true; }
  return false;
}

// ---- Duplicate filter state ----
let showDupsOnly = false;

function buildFilterBar() {
  const bar = document.createElement('div');
  bar.className = 'filter-bar';

  // Multi-select dropdowns
  for (const dim of FILTER_DIMS) {
    const wrap = document.createElement('label');
    wrap.textContent = FILTER_LABELS[dim];
    const dd = buildMultiSelect(dim, collectValues(dim));
    dd.id = 'filter-' + dim;
    wrap.appendChild(dd);
    bar.appendChild(wrap);
  }

  // Free-text search
  const searchWrap = document.createElement('label');
  searchWrap.textContent = 'Search';
  const search = document.createElement('input');
  search.type = 'text'; search.id = 'filter-search';
  search.placeholder = 'title / action_TBD';
  search.addEventListener('input', applyFilters);
  searchWrap.appendChild(search);
  bar.appendChild(searchWrap);

  // Date filter
  const dateWrap = document.createElement('label');
  dateWrap.textContent = 'Created after';
  const dateInput = document.createElement('input');
  dateInput.type = 'date'; dateInput.id = 'filter-date';
  dateInput.addEventListener('change', applyFilters);
  dateWrap.appendChild(dateInput);
  bar.appendChild(dateWrap);

  // Duplicate filter button
  const dupBtn = document.createElement('button');
  dupBtn.id = 'filter-dup-btn';
  dupBtn.textContent = 'Duplicates only';
  dupBtn.addEventListener('click', () => {
    showDupsOnly = !showDupsOnly;
    dupBtn.classList.toggle('active', showDupsOnly);
    applyFilters();
  });
  bar.appendChild(dupBtn);

  // Hide done toggle
  const hideWrap = document.createElement('label');
  hideWrap.style.flexDirection = 'row'; hideWrap.style.alignItems = 'center'; hideWrap.style.gap = '4px';
  const hideCb = document.createElement('input');
  hideCb.type = 'checkbox'; hideCb.id = 'filter-hide-done';
  hideCb.addEventListener('change', applyFilters);
  hideWrap.appendChild(hideCb);
  hideWrap.appendChild(document.createTextNode('Hide Done'));
  bar.appendChild(hideWrap);

  // Show hidden issues (issues without open request >1 week)
  const showAllWrap = document.createElement('label');
  showAllWrap.style.flexDirection = 'row'; showAllWrap.style.alignItems = 'center'; showAllWrap.style.gap = '4px';
  const showAllCb = document.createElement('input');
  showAllCb.type = 'checkbox'; showAllCb.id = 'filter-show-hidden';
  showAllCb.addEventListener('change', applyFilters);
  showAllWrap.appendChild(showAllCb);
  showAllWrap.appendChild(document.createTextNode('Show hidden issues'));
  bar.appendChild(showAllWrap);

  // Reset
  const reset = document.createElement('button');
  reset.textContent = 'Reset filters';
  reset.addEventListener('click', () => {
    for (const dim of FILTER_DIMS) {
      const dd = document.getElementById('filter-' + dim);
      SELECTED[dim].clear();
      dd.querySelectorAll('input[type=checkbox]').forEach(cb => { cb.checked = true; SELECTED[dim].add(cb.value); });
      updateButtonLabel(dd, dim);
    }
    document.getElementById('filter-search').value = '';
    document.getElementById('filter-date').value = '';
    document.getElementById('filter-hide-done').checked = false;
    document.getElementById('filter-show-hidden').checked = false;
    showDupsOnly = false;
    document.getElementById('filter-dup-btn').classList.remove('active');
    applyFilters();
  });
  bar.appendChild(reset);

  // Export done
  const exp = document.createElement('button');
  exp.textContent = 'Export Done IDs';
  exp.addEventListener('click', () => {
    const ids = [];
    for (let i = 0; i < localStorage.length; i++) {
      const k = localStorage.key(i);
      if (k.startsWith(STORAGE_KEY_PREFIX) && localStorage.getItem(k) === '1')
        ids.push(k.slice(STORAGE_KEY_PREFIX.length));
    }
    ids.sort((a, b) => Number(a) - Number(b));
    navigator.clipboard.writeText(ids.join(',')).then(
      () => alert('Copied ' + ids.length + ' done issue IDs'),
      () => prompt('Done IDs:', ids.join(','))
    );
  });
  bar.appendChild(exp);

  // Stats
  const stats = document.createElement('span');
  stats.className = 'stats'; stats.id = 'filter-stats';
  bar.appendChild(stats);

  document.body.insertBefore(bar, document.body.firstChild);
  document.addEventListener('click', () => {
    document.querySelectorAll('.ms-dd.open').forEach(d => d.classList.remove('open'));
  });
}

function applyFilters() {
  const search = (document.getElementById('filter-search').value || '').trim().toLowerCase();
  const hideDone = document.getElementById('filter-hide-done').checked;
  const showHidden = document.getElementById('filter-show-hidden').checked;
  const dateVal = document.getElementById('filter-date').value;
  let total = 0, visible = 0;

  document.querySelectorAll('table.ar-table tbody tr[data-issue]').forEach(tr => {
    const isMoreRow = tr.classList.contains('more-hidden') || tr.classList.contains('more-shown');
    if (isMoreRow && !showHidden) {
      tr.classList.remove('more-shown');
      tr.classList.add('more-hidden');
      tr.classList.remove('hidden');
      return;
    }
    if (isMoreRow && showHidden) {
      tr.classList.remove('more-hidden');
      tr.classList.add('more-shown');
    }
    total++;
    let show = true;
    for (const dim of FILTER_DIMS) {
      if (!rowMatchesDim(tr, dim, SELECTED[dim])) { show = false; break; }
    }
    if (show && search) {
      show = (tr.dataset.search || '').includes(search);
    }
    if (show && hideDone && tr.classList.contains('done')) show = false;
    if (show && showDupsOnly && tr.dataset.isDup !== '1') show = false;
    if (show && dateVal) {
      const created = tr.dataset.created || '';
      if (created && created < dateVal) show = false;
    }
    tr.classList.toggle('hidden', !show);
    if (show) visible++;
  });

  document.querySelectorAll('.more-link').forEach(link => {
    link.style.display = showHidden ? 'none' : '';
  });

  document.getElementById('filter-stats').textContent = visible + ' / ' + total + ' rows';
  hideEmptySections();
  updateFacetCounts();
  updateDynamicReportState();
}

function hideEmptySections() {
  document.querySelectorAll('table.ar-table').forEach(table => {
    const visibleRows = table.querySelectorAll('tbody tr[data-issue]:not(.hidden):not(.more-hidden)').length;

    const h3 = table.previousElementSibling;
    const moreLink = table.nextElementSibling;
    if (visibleRows === 0) {
      table.classList.add('section-hidden');
      if (h3 && h3.tagName === 'H3') h3.classList.add('section-hidden');
      if (moreLink && moreLink.classList.contains('more-link')) moreLink.classList.add('section-hidden');
    } else {
      table.classList.remove('section-hidden');
      if (h3 && h3.tagName === 'H3') h3.classList.remove('section-hidden');
      if (moreLink && moreLink.classList.contains('more-link')) moreLink.classList.remove('section-hidden');
    }
  });

  // Hide h2 sections if all their h3 subsections are hidden
  document.querySelectorAll('h2[id^="sec-p"], h2[id="sec-other"]').forEach(h2 => {
    let sib = h2.nextElementSibling;
    let anyVisible = false;
    while (sib && sib.tagName !== 'H2') {
      if (sib.tagName === 'TABLE' && sib.classList.contains('ar-table') && !sib.classList.contains('section-hidden')) {
        anyVisible = true; break;
      }
      sib = sib.nextElementSibling;
    }
    h2.classList.toggle('section-hidden', !anyVisible);
  });
}

// ---- Init ----
document.addEventListener('DOMContentLoaded', () => {
  loadDoneState();
  document.querySelectorAll('input.ar-done').forEach(cb => cb.addEventListener('change', onDoneToggle));
  buildFilterBar();
  initChartFilters();
  // Default filter: pre-select only "Need Response" in the AR multi-select
  // so the first paint reflects the bug-scrub focus list.
  const arDd = document.getElementById('filter-ar');
  if (arDd) {
    SELECTED['ar'] = new Set(['Need Response']);
    arDd.querySelectorAll('input[type=checkbox]').forEach(cb => {
      cb.checked = (cb.value === 'Need Response');
    });
    updateButtonLabel(arDd, 'ar');
  }
  applyFilters();
  initTooltips();
  initModal();
});

// ---- Custom tooltip for full reason text ----
function escapeHtml(s) {
  return String(s).replace(/[&<>"']/g, ch => ({
    '&': '&amp;', '<': '&lt;', '>': '&gt;', '"': '&quot;', "'": '&#39;'
  }[ch]));
}

function renderTooltipText(s) {
  let out = escapeHtml(s || '');
  out = out.replace(/```([^]*?)```/g,
    '<pre><code>$1</code></pre>');
  out = out.replace(/`([^`\\n]+)`/g,
    '<code style="background:#f1f3f5;color:#333;padding:1px 4px;border-radius:3px;font-family:monospace;font-size:11.5px">$1</code>');
  return out;
}

function initTooltips() {
  const popup = document.createElement('div');
  popup.className = 'tip-popup';
  document.body.appendChild(popup);

  // Move every .tip-cell `title` into data-tip so the browser's native
  // tooltip never appears alongside our custom popup.
  document.querySelectorAll('.tip-cell[title]').forEach(cell => {
    const t = cell.getAttribute('title');
    if (t) cell.setAttribute('data-tip', t);
    cell.removeAttribute('title');
  });

  document.addEventListener('mouseover', e => {
    const cell = e.target.closest('.tip-cell[data-tip]');
    if (!cell) { popup.classList.remove('visible'); return; }
    const tip = cell.getAttribute('data-tip');
    if (!tip) { popup.classList.remove('visible'); return; }
    popup.innerHTML = renderTooltipText(tip);
    popup.classList.add('visible');
    const rect = cell.getBoundingClientRect();
    let left = rect.left;
    let top = rect.bottom + 4;
    if (left + 700 > window.innerWidth) left = window.innerWidth - 710;
    if (top + 200 > window.innerHeight) top = rect.top - popup.offsetHeight - 4;
    popup.style.left = Math.max(0, left) + 'px';
    popup.style.top = top + 'px';
  });
  document.addEventListener('mouseout', e => {
    if (e.target.closest('.tip-cell')) popup.classList.remove('visible');
  });
}

// ---- In-page modal for Issue iframe + Fix Approach markdown ----
function openModal(title, contentNode, isMarkdown) {
  const overlay = document.getElementById('modal-overlay');
  const body = document.getElementById('modal-body');
  const ttl = document.getElementById('modal-title');
  if (!overlay || !body || !ttl) return;
  ttl.textContent = title || '';
  body.innerHTML = '';
  if (isMarkdown) body.classList.add('markdown'); else body.classList.remove('markdown');
  if (typeof contentNode === 'string') {
    body.innerHTML = contentNode;
  } else if (contentNode) {
    body.appendChild(contentNode);
  }
  overlay.classList.add('visible');
}

function closeModal() {
  const overlay = document.getElementById('modal-overlay');
  const body = document.getElementById('modal-body');
  if (!overlay || !body) return;
  overlay.classList.remove('visible');
  body.innerHTML = '';
  body.classList.remove('markdown');
}

function initModal() {
  const overlay = document.getElementById('modal-overlay');
  const closeBtn = document.getElementById('modal-close');
  if (overlay) {
    overlay.addEventListener('click', e => {
      if (e.target === overlay) closeModal();
    });
  }
  if (closeBtn) closeBtn.addEventListener('click', closeModal);
  document.addEventListener('keydown', e => {
    if (e.key === 'Escape') closeModal();
  });

  document.addEventListener('click', e => {
    const fixLink = e.target.closest('a.fix-modal-link');
    if (!fixLink) return;
    e.preventDefault();
    const iid = fixLink.getAttribute('data-iid') || '';
    const title = fixLink.getAttribute('data-title') || ('Issue #' + iid);
    const src = document.getElementById('md-' + iid);
    if (!src) {
      openModal(title, '<p>Details not available.</p>', true);
    } else {
      const md = src.textContent || '';
      if (typeof marked !== 'undefined' && marked.parse) {
        openModal(title, marked.parse(md), true);
      } else {
        const tmp = document.createElement('pre');
        tmp.textContent = md;
        openModal(title, tmp, true);
      }
    }
  });
}
"""


def render_markdown_blocks(issues: list[dict]) -> str:
    parts: list[str] = []
    details_dir = RESULT_DIR / "details"
    for i in issues:
        iid = i["id"]
        path = details_dir / f"{iid}.md"
        if not path.exists():
            continue
        try:
            md = path.read_text(encoding="utf-8", errors="replace")
        except Exception:
            continue
        md = md.replace("</script>", "<\\/script>")
        parts.append(
            f'<script type="text/markdown" id="md-{iid}">{md}</script>'
        )
    return "\n".join(parts)


def render_modal_overlay() -> str:
    return (
        '<div id="modal-overlay" role="dialog" aria-modal="true">'
        '<div id="modal-frame">'
        '<div id="modal-header">'
        '<span id="modal-title"></span>'
        '<button id="modal-close" type="button" aria-label="Close">&times;</button>'
        '</div>'
        '<div id="modal-body"></div>'
        '</div>'
        '</div>'
    )


def _closed_trend_payload(closed_trend: list[dict]) -> str:
    records = []
    for i in closed_trend:
        search = (str(i["id"]) + " " + i["title"] + " " + (i.get("action_tbd") or "")).lower()
        records.append({
            "dataset": {
                "ar": i["ar"],
                "assignee": i["assignee"],
                "owner_transferred": i["owner_transferred"],
                "priority": i["priority"],
                "category": i["category"],
                "milestone": i["milestone"],
                "dependency": i["dependency"],
                "test_module": i["test_module"],
                "search": search,
                "isDup": "1" if i["is_dup"] else "0",
                "created": i["created"][:10] if i["created"] else "",
                "closed": i["closed_dt"].strftime("%Y-%m-%d") if i["closed_dt"] else "",
            }
        })
    return json.dumps(records)


def render_page(body_html: str, issues: list[dict],
                closed_trend: list[dict] | None = None,
                title_suffix: str = "") -> str:
    title = "torch-xpu-ops Bug Scrub \u2014 Highlight" + title_suffix
    md_blocks = render_markdown_blocks(issues)
    modal = render_modal_overlay()
    closed_data = _closed_trend_payload(closed_trend or [])
    return (
        "<!doctype html>\n"
        '<html lang="en">\n<head>\n'
        '<meta charset="utf-8">\n'
        f"<title>{html.escape(title)}</title>\n"
        '<meta name="viewport" content="width=device-width, initial-scale=1">\n'
        f"<style>{CSS}</style>\n"
        '<script>/* marked.min.js - inlined */'
        f"{_MARKED_LIB}</script>\n"
        "</head>\n<body>\n"
        f'<div class="content">\n{body_html}\n</div>\n'
        f"{modal}\n"
        f"{md_blocks}\n"
        f"<script>window.TREND_CLOSED_DATA = {closed_data};</script>\n"
        f"<script>{JS}</script>\n"
        "</body>\n</html>\n"
    )


def _load_subset_ids(path: Path) -> set[int]:
    return {int(s) for line in path.read_text().splitlines()
            if (s := line.strip())}


def main(argv: list[str] | None = None) -> int:
    import argparse
    parser = argparse.ArgumentParser(
        description="Generate the torch-xpu-ops bug-scrub highlight report.")
    parser.add_argument(
        "--issues-file", type=Path, default=None,
        help="File with one issue id per line; restricts the report to that "
             "subset while keeping the full highlight format.")
    parser.add_argument(
        "--out", type=Path, default=None,
        help="Output HTML path (defaults to the full-report path, or a custom "
             "path under result/custom/ when --issues-file is given).")
    parser.add_argument(
        "--title-suffix", default=None,
        help="Appended to the report title (defaults to a subset note when "
             "--issues-file is given).")
    args = parser.parse_args(argv)

    if not XLSX_PATH.exists():
        print(f"ERROR: missing {XLSX_PATH}", file=sys.stderr)
        return 2

    issues = load_issues(XLSX_PATH)
    closed_trend = load_closed_trend_issues(XLSX_PATH)

    out_path = args.out
    title_suffix = args.title_suffix or ""
    if args.issues_file is not None:
        subset = _load_subset_ids(args.issues_file)
        issues = [i for i in issues if i["id"] in subset]
        closed_trend = [i for i in closed_trend if i["id"] in subset]
        if out_path is None:
            out_path = RESULT_DIR / "custom" / "bug_scrub_custom_highlight.html"
        if not title_suffix:
            title_suffix = f" \u2014 custom selection ({len(issues)} issues)"
    if out_path is None:
        out_path = HTML_PATH

    print(f"loaded {len(issues)} issues from {XLSX_PATH.name} "
          f"(+{len(closed_trend)} closed for trend)")

    body = build_html(issues, closed_trend)
    page = render_page(body, issues, closed_trend, title_suffix=title_suffix)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(page, encoding="utf-8")
    print(f"wrote {out_path} ({out_path.stat().st_size:,} bytes)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
