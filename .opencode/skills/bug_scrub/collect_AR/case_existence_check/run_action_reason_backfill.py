"""Phase 4c follow-up: populate action_reason for check_case_avaliablity issues.

For every row in the Issues sheet whose action_TBD contains
`check_case_avaliablity` and whose action_reason is blank, aggregate the
distinct non-empty `case_existence_comments` from the Test Cases sheet
and write them back into action_reason. A single distinct comment is
written as a plain string; multiple distinct comments are written as a
JSON array (matching the existing action_reason conventions).

Usage:
    python3 run_action_reason_backfill.py

Anchors the Excel path via __file__ so it runs from any CWD.
Backs up the workbook to ..._bk_before_action_reason_backfill.xlsx before writing.
"""
from __future__ import annotations

import json
import shutil
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

COMMON_DIR = Path(__file__).resolve().parents[2] / "_common"
if str(COMMON_DIR) not in sys.path:
    sys.path.insert(0, str(COMMON_DIR))
from header_utils import row_dict, write_by_name  # type: ignore[reportMissingImports] # noqa: E402
from paths import RESULT_DIR  # type: ignore[reportMissingImports] # noqa: E402

EXCEL = RESULT_DIR / "torch_xpu_ops_issues.xlsx"
BACKUP = EXCEL.with_name(
    EXCEL.stem + "_bk_before_action_reason_backfill.xlsx"
)
TOKEN = "check_case_avaliablity"


def main() -> None:
    wb = openpyxl.load_workbook(EXCEL)
    issues = wb["Issues"]
    cases = wb["Test Cases"]

    # issue_id -> list of distinct non-empty comments (preserve insertion order)
    comments: dict[int, list[str]] = defaultdict(list)
    for row_idx in range(2, cases.max_row + 1):
        try:
            data = row_dict(cases, row_idx)
            iid = data.get("Issue ID")
            c = data["case_existence_comments"]
        except KeyError as e:
            print(f"Warning: {e}; skipping action_reason backfill")
            return
        if iid is None or c is None:
            continue
        s = str(c).strip()
        if not s or s.lower() == "none":
            continue
        if s not in comments[int(iid)]:
            comments[int(iid)].append(s)

    updated = 0
    skipped_no_blank = 0
    skipped_no_comments = 0
    for row in issues.iter_rows(min_row=2):
        row_idx = row[0].row
        data = row_dict(issues, row_idx)
        iid = data.get("Issue ID")
        if iid is None:
            continue
        tbd = (data.get("action_TBD") or "").strip()
        if TOKEN not in tbd:
            continue
        existing = (data.get("action_reason") or "")
        if str(existing).strip():
            skipped_no_blank += 1
            continue
        lst = comments.get(int(iid), [])
        if not lst:
            skipped_no_comments += 1
            continue
        if len(lst) == 1:
            write_by_name(issues, row_idx, "action_reason", lst[0])
        else:
            write_by_name(issues, row_idx, "action_reason", json.dumps(lst, ensure_ascii=False))
        updated += 1

    print(
        f"check_case_avaliablity rows updated: {updated}\n"
        f"  skipped (action_reason already set): {skipped_no_blank}\n"
        f"  skipped (no Test Cases comments):    {skipped_no_comments}"
    )

    if updated == 0:
        print("no changes; skipping save")
        return

    shutil.copy(EXCEL, BACKUP)
    print(f"backed up to {BACKUP}")
    wb.save(EXCEL)
    print(f"wrote {EXCEL}")


if __name__ == "__main__":
    main()
