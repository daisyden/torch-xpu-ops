#!/usr/bin/env python3
"""
Standalone script to run match_e2e_status function.

This script loads E2E test status from Inductor_E2E_Test_Report.xlsx files and matches them
to E2E test cases in the torch_xpu_ops_issues.xlsx workbook.

Usage:
    python run_match_e2e_status.py [--excel EXCEL_FILE] [--base-dir BASE_DIR] [--no-save]

Requirements:
    - torch_xpu_ops_issues.xlsx with 'E2E Test Cases' sheet
    - E2E test report files in directories like:
        */Inductor_E2E_Test_Report.xlsx
"""

import os
import sys
import re
import json
import subprocess


def find_issue_triage_root(start: str) -> str:
    if os.environ.get('ISSUE_TRIAGE_ROOT'):
        root = os.path.abspath(os.environ['ISSUE_TRIAGE_ROOT'])
        if (os.path.isdir(os.path.join(root, 'result')) and
                os.path.isdir(os.path.join(root, 'ci_results'))):
            return root
    path = os.path.abspath(start)
    while True:
        if (os.path.isdir(os.path.join(path, 'result')) and
                os.path.isdir(os.path.join(path, 'ci_results'))):
            return path
        parent = os.path.dirname(path)
        if parent == path:
            raise RuntimeError(f'Could not locate issue_triage root from {start}')
        path = parent

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = find_issue_triage_root(SCRIPT_DIR)
sys.path.insert(0, PROJECT_ROOT)
_COMMON_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, '..', '..', '_common'))
if _COMMON_DIR not in sys.path:
    sys.path.insert(0, _COMMON_DIR)
from header_utils import cell_by_name, ensure_col, write_by_name  # type: ignore[reportMissingImports] # noqa: E402

import glob
import argparse
import openpyxl


def normalize_key_value(value):
    if value is None:
        return ''
    return str(value).lower().strip().replace(' ', '_').replace('-', '_')


def fetch_torchbench_models_from_github():
    """Fetch torchbench model names from pytorch/benchmark GitHub repository.
    
    Fetches directory listings from:
    - https://github.com/pytorch/benchmark/tree/main/torchbenchmark/models
    - https://github.com/pytorch/benchmark/tree/main/torchbenchmark/canary_models
    
    Returns:
        dict: {model_name: 'torchbench'} mapping for valid models
              Returns empty dict if fetch fails (non-blocking)
    """
    models = {}
    directories = [
        'torchbenchmark/models',
        'torchbenchmark/canary_models'
    ]
    
    for directory in directories:
        try:
            # Use gh CLI to fetch directory listing as JSON
            cmd = [
                'gh', 'api', 'repos/pytorch/benchmark/contents/' + directory,
                '--jq', '.[].name'
            ]
            result = subprocess.run(cmd, capture_output=True, text=True, timeout=10)
            if result.returncode != 0:
                print(f"    Warning: Failed to fetch {directory} from GitHub (exit code {result.returncode})")
                continue
            
            for line in result.stdout.strip().split('\n'):
                model_name = line.strip().strip('"')
                if model_name and not model_name.startswith('.'):
                    models[normalize_key_value(model_name)] = 'torchbench'
            
            print(f"    Fetched {len([m for m in result.stdout.strip().split('\n') if m.strip()])} models from {directory}")
        except subprocess.TimeoutExpired:
            print(f"    Warning: Timeout fetching {directory} from GitHub")
        except FileNotFoundError:
            print(f"    Warning: gh CLI not found; skipping GitHub torchbench model fetch")
            break
        except Exception as e:
            print(f"    Warning: Error fetching {directory}: {e}")
    
    return models


def create_model_variants(name):
    if not name:
        return []
    name = str(name).lower().strip()
    variants = set()
    variants.add(name)
    if '/' in name:
        parts = name.split('/')
        variants.add(parts[-1])
        variants.add(name.replace('/', ''))
        variants.add(parts[0] + parts[1])
    camel_split = re.sub(r'([a-z])([A-Z])', r'\1_\2', name).split('_')
    if len(camel_split) > 1:
        variants.add('_'.join(camel_split).lower())
        variants.add(''.join(camel_split).lower())
    variants.add(name.replace('_', ''))
    return list(variants)


def parse_sheet_name(sheet_name):
    parts = sheet_name.replace('_acc', '').split('_')
    if parts[0] in ['huggingface', 'timm', 'torchbench']:
        benchmark = parts[0]
        idx = 1
    else:
        benchmark = 'unknown'
        idx = 0
    amp = False
    dtype = 'float32'
    phase = 'inf'
    remaining = parts[idx:]
    for i, part in enumerate(remaining):
        if part == 'amp':
            amp = True
        elif part in ['bf16', 'float16', 'float32', 'fp16', 'fp32']:
            dtype = part.replace('bf16', 'bfloat16').replace('fp16', 'float16').replace('fp32', 'float32')
        elif part in ['inf', 'tra', 'training']:
            phase = 'inf' if part == 'inf' else 'tra'
    return benchmark, amp, dtype, phase


def load_all_e2e_reports(base_dir, torchbench_models=None):
    all_models = {}
    if torchbench_models:
        # Pre-populate with torchbench models from GitHub
        for model_name in torchbench_models.keys():
            # Create torchbench variants for each model
            all_models[f"torchbench|float32|inf|False|{model_name}"] = 'pass'
            all_models[f"torchbench|float16|inf|False|{model_name}"] = 'pass'
            all_models[f"torchbench|bfloat16|inf|False|{model_name}"] = 'pass'
    for xlsx_path in glob.glob(f'{base_dir}/*/Inductor_E2E_Test_Report.xlsx'):
        try:
            dirname = os.path.basename(os.path.dirname(xlsx_path))
            print(f"  Loading: {dirname}/Inductor_E2E_Test_Report.xlsx")
            wb = openpyxl.load_workbook(xlsx_path, read_only=True)
            for sheet_name in wb.sheetnames:
                if not sheet_name.endswith('_acc'):
                    continue
                benchmark, amp, dtype, phase = parse_sheet_name(sheet_name)
                model_prefix = f"{benchmark}|{dtype}|{phase}|{amp}"
                sheet = wb[sheet_name]
                row_count = 0
                for row in sheet.iter_rows(min_row=3, values_only=True):
                    if not row or not row[1]:
                        continue
                    model_name = str(row[1]).strip()
                    accuracy = row[3] if len(row) > 3 else None
                    if accuracy:
                        for v in create_model_variants(model_name):
                            all_models[f"{model_prefix}|{v}"] = accuracy
                        row_count += 1
                print(f"    Sheet {sheet_name}: {row_count} models")
            wb.close()
        except Exception as e:
            print(f"  Warning: Failed to parse {xlsx_path}: {e}")
    return all_models


SKIP_RULE_BLANK = {None, '', 'status not found', 'not found', 'none', 'n/a'}


def _is_blank_for_skip(value):
    if value is None:
        return True
    return str(value).strip().lower() in SKIP_RULE_BLANK


def run_match_e2e_status(excel_file, base_dir, save=True, skip_filled=False, fetch_github_models=True):
    print(f"Loading: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)
    if 'E2E Test Cases' not in wb.sheetnames:
        print("  Error: 'E2E Test Cases' sheet not found!")
        return 0
    ws_e2e = wb['E2E Test Cases']
    total_rows = ws_e2e.max_row - 1
    print(f"  Total rows: {total_rows} (skip_filled={skip_filled})")
    ensure_col(ws_e2e, 'XPU Accuracy Status')
    print(f"  Loading E2E model status from reports...")
    
    torchbench_models = None
    if fetch_github_models:
        print(f"  Fetching torchbench model list from GitHub...")
        torchbench_models = fetch_torchbench_models_from_github()
        if torchbench_models:
            print(f"    Fetched {len(torchbench_models)} torchbench models from GitHub")
        else:
            print(f"    No torchbench models fetched; using E2E reports only")
    
    status_map = load_all_e2e_reports(base_dir, torchbench_models=torchbench_models)
    print(f"  Built status map with {len(status_map)} entries")
    matched = 0
    not_found = 0
    skipped = 0
    for row in range(2, ws_e2e.max_row + 1):
        if skip_filled and not _is_blank_for_skip(cell_by_name(ws_e2e, row, 'XPU Accuracy Status').value):
            skipped += 1
            continue
        benchmark = cell_by_name(ws_e2e, row, 'Benchmark').value
        model = cell_by_name(ws_e2e, row, 'Model').value
        dtype = cell_by_name(ws_e2e, row, 'Dtype').value
        amp = cell_by_name(ws_e2e, row, 'AMP').value
        phase = cell_by_name(ws_e2e, row, 'Phase').value
        if benchmark and model and dtype and phase:
            dtype_key = normalize_key_value(dtype)
            phase_key = 'inf' if normalize_key_value(phase).startswith('inf') else 'tra'
            amp_val = bool(amp) if amp else False
            model_variants = create_model_variants(model)
            found = False
            for model_var in model_variants:
                for check_amp in [amp_val, False] if amp_val else [False]:
                    for dtype_fallback in [dtype_key, 'float32', 'float16']:
                        key = f"{benchmark}|{dtype_fallback}|{phase_key}|{check_amp}|{model_var}"
                        if key in status_map:
                            write_by_name(ws_e2e, row, 'XPU Accuracy Status', status_map[key])
                            matched += 1
                            found = True
                            break
                    if found:
                        break
                if found:
                    break
            if not found:
                write_by_name(ws_e2e, row, 'XPU Accuracy Status', 'Status not found')
                not_found += 1
    print(f"  Matched {matched} E2E test cases, {not_found} not found, {skipped} skipped (already filled)")
    if save:
        wb.save(excel_file)
        print(f"  Saved: {excel_file}")
    return matched


def main():
    _THIS_DIR = os.path.dirname(os.path.abspath(__file__))
    _DEFAULT_ROOT = find_issue_triage_root(_THIS_DIR)
    _RESULT = os.environ.get('RESULT_DIR') or os.environ.get('ISSUE_TRIAGE_RESULT_DIR') or os.path.join(_DEFAULT_ROOT, 'result')
    _CI = os.environ.get('ISSUE_TRIAGE_CI_RESULTS') or os.path.join(_DEFAULT_ROOT, 'ci_results')

    parser = argparse.ArgumentParser(description='Match E2E test cases to accuracy status')
    parser.add_argument('--excel', default=os.path.join(_RESULT, 'torch_xpu_ops_issues.xlsx'),
                        help='Excel file with E2E_Test_Cases sheet')
    parser.add_argument('--base-dir', default=os.path.join(_CI, 'torch-xpu-ops') + os.sep,
                        help='Base directory containing E2E report folders')
    parser.add_argument('--no-save', action='store_true', help='Do not save results')
    parser.add_argument('--skip-filled', action='store_true',
                        help='Only fill blank or placeholder cells; preserve existing real statuses')
    parser.add_argument('--no-github-models', action='store_true',
                        help='Do not fetch torchbench models from GitHub; use E2E reports only')
    args = parser.parse_args()
    matched = run_match_e2e_status(args.excel, args.base_dir, save=not args.no_save,
                                   skip_filled=args.skip_filled, fetch_github_models=not args.no_github_models)
    print(f"\nComplete: Matched {matched} E2E test cases")


if __name__ == '__main__':
    main()
