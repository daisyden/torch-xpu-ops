#!/usr/bin/env python3
"""
Run Test Cases Processor - Step by Step

Usage:
    python3 run_processor_steps.py              # Run all automated Test Cases steps
    python3 run_processor_steps.py --steps 1      # Run Phase 2.1 UT CI matching
    python3 run_processor_steps.py --steps 3      # Print Phase 2.4 worklist
    python3 run_processor_steps.py --list       # List available steps
    python3 run_processor_steps.py --help       # Show help

Steps:
    1. PASS 1: Create test_cases_all.xlsx with stock and torch-xpu-ops sheets, match CI results
    3. PASS 3: Prepare check_xpu_case_existence worklist (no automated LLM endpoint)
    5. PASS 5: Duplicate detection (cross-issue)

Example:
    # Run full processor
    python3 run_processor_steps.py

    # Run specific steps
    python3 run_processor_steps.py --steps 1 3 5
"""

import os
import sys
import time
import argparse
import importlib


def find_issue_triage_root(start: str) -> str:
    if os.environ.get('ISSUE_TRIAGE_ROOT'):
        root = os.path.abspath(os.environ['ISSUE_TRIAGE_ROOT'])
        if (os.path.isdir(os.path.join(root, 'result')) and
                os.path.isdir(os.path.join(root, 'ci_results'))):
            return root
    path = os.path.abspath(start)
    while True:
        if (os.path.isdir(os.path.join(path, 'result')) and
                os.path.isdir(os.path.join(path, 'ci_results'))):
            return path
        parent = os.path.dirname(path)
        if parent == path:
            raise RuntimeError(f'Could not locate issue_triage root from {start}')
        path = parent

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_DEFAULT_ROOT = find_issue_triage_root(SCRIPT_DIR)
ISSUE_TRIAGE_ROOT = os.environ.get('ISSUE_TRIAGE_ROOT') or _DEFAULT_ROOT
ROOT_DIR = os.environ.get('ROOT_DIR', os.path.dirname(ISSUE_TRIAGE_ROOT))
RESULT_DIR = os.environ.get('RESULT_DIR') or os.environ.get('ISSUE_TRIAGE_RESULT_DIR') or os.path.join(ISSUE_TRIAGE_ROOT, 'result')
LOG_FILE = os.path.join(RESULT_DIR, "pipeline.log")

sys.path.insert(0, SCRIPT_DIR)
sys.path.insert(0, os.path.join(SCRIPT_DIR, '..'))
_COMMON_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, '..', '..', '_common'))
if _COMMON_DIR not in sys.path:
    sys.path.insert(0, _COMMON_DIR)
from header_utils import cell_by_name  # type: ignore[reportMissingImports] # noqa: E402

_PASS1_MODULE = importlib.import_module(
    f"{__package__}.pass1_ci_matcher" if __package__ else "pass1_ci_matcher"
)
_DUP_MODULE = importlib.import_module(
    f"{__package__}.pass5_duplicate_detection" if __package__ else "pass5_duplicate_detection"
)
pass1_match_ci_results = _PASS1_MODULE.pass1_match_ci_results
pass5_duplicate_detection = _DUP_MODULE.pass5_duplicate_detection

STEPS_DESC = {
    1: "PASS 1: Create test_cases_all.xlsx, collect stock & xpu CI results, match CI",
    3: "PASS 3: Prepare check_xpu_case_existence worklist (no automated LLM endpoint)",
    5: "PASS 5: Duplicate detection (cross-issue)",
}

STEP_FUNCS = {
    5: pass5_duplicate_detection,
}

llm_steps = set()


def _is_missing_ci_status(value):
    if value is None:
        return True
    return str(value).strip().lower() in {'', 'not found', 'not in stock ci', 'not_run'}


def pass3_check_xpu_case_existence_worklist(ws):
    """Print the Phase 2.4 worklist for check_xpu_case_existence/SKILL.md.

    Phase 2.4 is intentionally not automated here: the canonical implementation is
    the deep-analysis skill at analyze_ci_result/check_xpu_case_existence/SKILL.md.
    This runner only identifies the first unresolved Test Cases row per issue and
    tells the agent exactly which skill to use.
    """
    print_step_info([3])
    print("  [PASS 3/5] Preparing check_xpu_case_existence worklist...")
    print("  Skill: ${BUG_SCRUB_SKILL_ROOT}/analyze_ci_result/check_xpu_case_existence/SKILL.md")
    print("  This step only prints a manual worklist; no automated LLM endpoint is called.")

    worklist = []
    seen_issues = set()
    for row in range(2, ws.max_row + 1):
        issue_id = cell_by_name(ws, row, 'Issue ID').value
        if issue_id in seen_issues:
            continue
        try:
            xpu_status = cell_by_name(ws, row, 'XPU Status').value
            stock_status = cell_by_name(ws, row, 'Stock Status').value
        except KeyError as e:
            print(f"  Warning: {e}; skipping Phase 2.4 worklist")
            return worklist
        if _is_missing_ci_status(xpu_status) and _is_missing_ci_status(stock_status):
            seen_issues.add(issue_id)
            worklist.append((
                row,
                issue_id,
                cell_by_name(ws, row, 'Test File').value,
                cell_by_name(ws, row, 'Origin Test File').value,
                cell_by_name(ws, row, 'Test Class').value,
                cell_by_name(ws, row, 'Test Case').value,
            ))

    print(f"  Eligible issues: {len(worklist)}")
    for row, issue_id, test_file, origin_test_file, test_class, test_case in worklist[:50]:
        print(
            "  "
            f"row={row} issue={issue_id} file={test_file!r} origin={origin_test_file!r} "
            f"class={test_class!r} case={test_case!r}"
        )
    if len(worklist) > 50:
        print(f"  ... {len(worklist) - 50} more eligible issues omitted from preview")

    print("\n  To complete Phase 2.4, run check_xpu_case_existence for each listed issue")
    print("  and write results to Test Cases columns 17 (XPU Case Exist) and 18")
    print("  (case_existence_comments), as specified by that skill.")
    return worklist


def log(msg, print_also=True):
    """Log message to file and optionally print to console."""
    timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
    log_msg = f"[{timestamp}] {msg}"
    try:
        with open(LOG_FILE, "a") as f:
            f.write(log_msg + "\n")
    except Exception:
        pass
    if print_also:
        print(log_msg)


def parse_steps(steps_arg):
    """Parse step arguments like '1', '1-3', '1 2 3' into sorted list."""
    steps = set()

    for arg in steps_arg:
        arg = arg.strip()
        if '-' in arg:
            try:
                start, end = map(int, arg.split('-'))
                steps.update(range(start, end + 1))
            except ValueError:
                pass
        elif arg.isdigit():
            steps.add(int(arg))

    return sorted(steps)


def print_step_info(steps):
    """Print information about steps being.run"""
    print("\n" + "=" * 60)
    print("Test Cases Processor - Step by Step Runner")
    print("=" * 60)

    print("\nSteps to run:")
    all_llm = all(s in llm_steps for s in steps)
    has_llm = any(s in llm_steps for s in steps)
    fast_mode = not has_llm

    for step in sorted(steps):
        desc = STEPS_DESC.get(step, f"Unknown step {step}")
        print(f"  {step}. {desc}")

    if fast_mode:
        print("\nMode: FAST (pattern-based, no LLM calls)")
    elif all_llm:
        print("\nMode: LLM ONLY (all steps require LLM)")
    else:
        print("\nMode: MIXED (some LLM calls required)")

    print(f"Total steps: {len(steps)}")
    print("=" * 60 + "\n")


def run_steps(steps_to_run, input_file=None, save=True, incremental=False):
    """
    Run specified steps of the test cases processor.

    Args:
        steps_to_run: list of step numbers to execute
        input_file: optional path to input Excel file
        save: whether to save results after each step
        incremental: skip rows with already-filled result columns

    Returns:
        tuple: (workbook, issues_needing_llm, issue_duplicated_map)
    """
    import openpyxl

    start_total = time.time()

    excel_file = input_file or os.path.join(RESULT_DIR, 'torch_xpu_ops_issues.xlsx')

    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        return None, None, None

    print(f"Loading: {excel_file}")
    wb = openpyxl.load_workbook(excel_file)

    if 'Test Cases' not in wb.sheetnames:
        print("ERROR: 'Test Cases' sheet not found in workbook")
        return None, None, None

    ws = wb['Test Cases']
    total_rows = ws.max_row - 1
    print(f"Total test cases: {total_rows}\n")

    issues_needing_llm = None
    issue_duplicated_map = None

    if 1 in steps_to_run:
        print_step_info([1])
        issues_needing_llm = pass1_match_ci_results(ws, os.path.join(RESULT_DIR, 'test_cases_all.xlsx'))
        if save:
            wb.save(excel_file)
            print(f"Saved to: {excel_file}")

    if 3 in steps_to_run:
        pass3_check_xpu_case_existence_worklist(ws)

    if 5 in steps_to_run:
        print_step_info([5])
        issue_duplicated_map = pass5_duplicate_detection(ws, skip_filled=incremental)
        if save:
            wb.save(excel_file)
            print(f"Saved to: {excel_file}")

    elapsed_total = time.time() - start_total
    print("\n" + "=" * 60)
    print(f"Processing complete in {elapsed_total:.1f}s ({elapsed_total/60:.1f} min)")
    print("=" * 60)

    return wb, issues_needing_llm, issue_duplicated_map


def main():
    parser = argparse.ArgumentParser(
        description='Run Test Cases Processor - Step by Step',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__
    )

    parser.add_argument(
        '--steps', nargs='+', default=None,
        help='Steps to run (supported: 1, 3, 5; e.g., "1", "1 3 5", or "all")'
    )
    parser.add_argument(
        '--list', action='store_true',
        help='List all available steps with descriptions'
    )
    parser.add_argument(
        '--input', '-i', type=str, default=None,
        help='Input Excel file path (default: auto-detect from RESULT_DIR)'
    )
    parser.add_argument(
        '--no-save', action='store_true',
        help='Do not save results after each step (for debugging)'
    )
    parser.add_argument(
        '--incremental', action='store_true',
        help='Incremental mode: skip rows with already-filled result columns (Phase 2.3+ only)'
    )
    parser.add_argument(
        '--fast', action='store_true',
        help='Run automated Test Cases steps (1 and 5) - skip manual Phase 2.4 worklist'
    )

    args = parser.parse_args()

    if args.list:
        print("\nAvailable Steps:")
        print("-" * 60)
        for step, desc in STEPS_DESC.items():
            print(f"  {step}. {desc}")
        print("-" * 60)
        return 0

    if args.fast:
        steps_to_run = [1, 5]
        print("Fast mode: Skipping step 3 (manual check_xpu_case_existence worklist)")
    elif args.steps is None or (len(args.steps) == 1 and args.steps[0].lower() in ['all', 'a']):
        steps_to_run = [1, 3, 5]
        print("Running all steps (full processor)")
    else:
        steps_to_run = parse_steps(args.steps)
        if not steps_to_run:
            print("ERROR: No valid steps specified")
            print("Use --steps 1, --steps 3, --steps 5, or --steps 1 3 5")
            return 1


    unsupported = [s for s in steps_to_run if s not in STEPS_DESC]
    if unsupported:
        print(f"ERROR: Unsupported removed steps requested: {unsupported}")
        print("Phase 2 torch-ops extraction/dependency RAG was removed; use --steps 1 3 5.")
        return 1

    os.makedirs(RESULT_DIR, exist_ok=True)

    save = not args.no_save
    run_steps(steps_to_run, input_file=args.input, save=save, incremental=args.incremental)

    return 0


if __name__ == '__main__':
    sys.exit(main())
