#!/usr/bin/env python3
"""
PASS 1: CI Test Case Matcher

This module handles collecting CI test results from stock and XPU sources,
loading them from test_cases_all.xlsx, and matching test cases to issues.

Functions:
- collect_stock_test_cases(): Collect stock CI test cases from XML files
- collect_torch_xpu_ops_test_cases(): Collect XPU CI test cases from XML files
- create_test_cases_all_excel(): Create/merge test_cases_all.xlsx
- load_test_cases_all(): Load Excel and build lookup maps
- normalize_class_name(): Extract short class name from various formats
- pass1_match_ci_results(): Match CI results to test cases in workbook
"""

import os
import re
import glob
import sys
import zipfile
import time
import xml.etree.ElementTree as ET
import openpyxl


def find_issue_triage_root(start: str) -> str:
    if os.environ.get('ISSUE_TRIAGE_ROOT'):
        root = os.path.abspath(os.environ['ISSUE_TRIAGE_ROOT'])
        if (os.path.isdir(os.path.join(root, 'result')) and
                os.path.isdir(os.path.join(root, 'ci_results'))):
            return root
    path = os.path.abspath(start)
    while True:
        if (os.path.isdir(os.path.join(path, 'result')) and
                os.path.isdir(os.path.join(path, 'ci_results'))):
            return path
        parent = os.path.dirname(path)
        if parent == path:
            raise RuntimeError(f'Could not locate issue_triage root from {start}')
        path = parent

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_COMMON_DIR = os.path.abspath(os.path.join(_THIS_DIR, '..', '..', '_common'))
if _COMMON_DIR not in sys.path:
    sys.path.insert(0, _COMMON_DIR)
from header_utils import cell_by_name, ensure_col, row_dict, write_by_name  # type: ignore[reportMissingImports] # noqa: E402

_DEFAULT_ROOT = find_issue_triage_root(_THIS_DIR)
ISSUE_TRIAGE_ROOT = os.environ.get('ISSUE_TRIAGE_ROOT') or _DEFAULT_ROOT
RESULT_DIR = os.environ.get('ISSUE_TRIAGE_RESULT_DIR', os.path.join(ISSUE_TRIAGE_ROOT, 'result'))
CI_RESULTS_DIR = os.environ.get('ISSUE_TRIAGE_CI_RESULTS', os.path.join(ISSUE_TRIAGE_ROOT, 'ci_results'))
LOG_FILE = os.path.join(RESULT_DIR, 'pipeline.log')


def log(msg, print_also=True):
    timestamp = time.strftime('%Y-%m-%d %H:%M:%S')
    log_msg = f'[{timestamp}] {msg}'
    try:
        os.makedirs(RESULT_DIR, exist_ok=True)
        with open(LOG_FILE, 'a') as f:
            f.write(log_msg + '\n')
    except Exception:
        pass
    if print_also:
        print(log_msg)


def parse_ci_xml_content(content):
    """Parse pytest XML content and extract test cases with results."""
    try:
        root = ET.fromstring(content)
    except Exception:
        return []

    test_cases = []
    for testcase in root.findall('.//testcase'):
        name = testcase.get('name', '')
        classname = testcase.get('classname', '')
        file_path = testcase.get('file', '')

        failure = testcase.find('failure')
        skipped = testcase.find('skipped')

        status = 'passed'
        error_msg = ''
        traceback = ''

        if failure is not None:
            status = 'failed'
            msg = failure.text or failure.get('message', '') or ''
            error_msg, traceback = parse_failure_message(msg)
        elif skipped is not None:
            status = 'skipped'
            msg = skipped.text or skipped.get('message', '') or ''
            error_msg = msg[:500] if msg else 'skipped'

        test_cases.append({
            'test_file': file_path,
            'test_class': classname,
            'test_case': name,
            'status': status,
            'error_msg': error_msg,
            'traceback': traceback
        })

    return test_cases


def parse_failure_message(content):
    """Parse failure message to extract error_msg and traceback."""
    error_msg = ""
    traceback = ""

    if not content:
        return error_msg, traceback

    lines = content.split('\n')

    error_patterns = [
        (r'^RuntimeError', 'RuntimeError'),
        (r'^AssertionError', 'AssertionError'),
        (r'^ValueError', 'ValueError'),
        (r'^TypeError', 'TypeError'),
        (r'^IndexError', 'IndexError'),
        (r'^KeyError', 'KeyError'),
        (r'^ImportError', 'ImportError'),
        (r'^NotImplementedError', 'NotImplementedError'),
        (r'^AttributeError', 'AttributeError'),
        (r'^InductorError', 'InductorError'),
    ]

    error_line_idx = -1
    error_type = None
    last_error_msg = ""

    for idx, line in enumerate(lines):
        stripped = line.strip()
        for pattern, etype in error_patterns:
            if re.match(pattern, stripped):
                error_line_idx = idx
                error_type = etype
                clean_line = re.sub(r'^' + etype + r'[:\s]*', '', stripped)
                error_msg = clean_line[:200]
                break
        if error_line_idx >= 0:
            break
        for ep in [r'\braise\s+(RuntimeError|AssertionError|ValueError|TypeError|IndexError|KeyError|ImportError|NotImplementedError|AttributeError|InductorError)\s*[\(\'"]']:
            if re.search(ep, stripped):
                error_line_idx = idx
                match = re.search(r'raise\s+\w+\s*[\(\'"](.+?)[\'\"]?', stripped)
                if match:
                    last_error_msg = match.group(1).strip()[:200]

    traceback = ""
    if 'Traceback (most recent call last):' in content:
        tb_lines = []
        end_idx = error_line_idx if error_line_idx >= 0 else len(lines)
        for idx, line in enumerate(lines):
            if 'Traceback (most recent call last):' in line:
                for j in range(idx, end_idx + 1):
                    tb_lines.append(lines[j])
                break

        if tb_lines:
            traceback = '\n'.join(tb_lines)
        elif last_error_msg:
            for idx, line in enumerate(lines):
                stripped = line.strip()
                for ep in [r'\braise\s+(RuntimeError|AssertionError|ValueError|TypeError|IndexError|KeyError|ImportError|NotImplementedError|AttributeError|InductorError)\s*[\(\'"]']:
                    if re.search(ep, stripped):
                        traceback = '\n'.join(lines[idx:])
                        break
                if traceback:
                    break
    else:
        traceback = ""
        error_msg = content[:200]

    if last_error_msg and not error_msg:
        error_msg = last_error_msg

    return error_msg[:300] if error_msg else error_msg, traceback[:3000] if traceback else traceback


def collect_stock_test_cases():
    """Collect all test cases from stock PyTorch CI pytest XML files."""
    log("  Collecting stock CI test cases...")
    stock_base = os.path.join(CI_RESULTS_DIR, 'stock')

    stock_test_cases = []
    total_files = 0
    total_cases = 0

    for mount_point in glob.glob(f'{stock_base}/test-reports-runattempt1*.zip'):
        try:
            pytest_dir = os.path.join(mount_point, 'test-reports', 'python-pytest')
            if os.path.isdir(pytest_dir):
                for root, dirs, files in os.walk(pytest_dir):
                    for f in files:
                        if f.endswith('.xml'):
                            xml_path = os.path.join(root, f)
                            try:
                                with open(xml_path, 'r') as file:
                                    content = file.read()
                                test_cases = parse_ci_xml_content(content)
                                if test_cases:
                                    test_module = os.path.basename(root)
                                    for tc in test_cases:
                                        tc['source'] = f"stock:{test_module}"
                                    stock_test_cases.extend(test_cases)
                                    total_files += 1
                                    total_cases += len(test_cases)
                            except Exception:
                                pass
            else:
                with zipfile.ZipFile(mount_point, 'r') as zf:
                    for name in zf.namelist():
                        if name.endswith('.xml') and '/python-pytest/' in name:
                            try:
                                with zf.open(name) as f:
                                    content = f.read().decode('utf-8', errors='ignore')
                                test_cases = parse_ci_xml_content(content)
                                if test_cases:
                                    parts = name.split('/')
                                    test_module = parts[-2] if len(parts) >= 2 else name
                                    for tc in test_cases:
                                        tc['source'] = f"stock:{test_module}"
                                    stock_test_cases.extend(test_cases)
                                    total_files += 1
                                    total_cases += len(test_cases)
                            except Exception:
                                pass
        except Exception:
            continue

    log(f"  Stock CI: {total_cases} test cases from {total_files} XML files")
    return stock_test_cases


def collect_torch_xpu_ops_test_cases():
    """Collect all test cases from torch-xpu-ops CI pytest XML files."""
    log("  Collecting torch-xpu-ops CI test cases...")
    base_dir = os.path.join(CI_RESULTS_DIR, 'torch-xpu-ops')

    xpu_test_cases = []
    total_files = 0
    total_cases = 0

    for d in os.listdir(base_dir):
        if d.startswith('Inductor-XPU-UT-Data-'):
            match = re.match(r'Inductor-XPU-UT-Data-([a-f0-9]+)-.*-(\d+)-1$', d)
            if match:
                folder_path = os.path.join(base_dir, d, d)
                if not os.path.exists(folder_path):
                    continue
                for f in os.listdir(folder_path):
                    if f.endswith('.xml') and (f.startswith('op_ut_with_all') or f.startswith('op_ut_with_skip') or f.startswith('op_ut_with_exe') or f == 'op_extended.xml'):
                        xml_path = os.path.join(folder_path, f)
                        try:
                            tree = ET.parse(xml_path)
                            root = tree.getroot()
                            count = len(root.findall('.//testcase'))
                            if count > 0:
                                prefix = f.replace('.xml', '')
                                for testcase in root.findall('.//testcase'):
                                    name = testcase.get('name', '')
                                    classname = testcase.get('classname', '')
                                    file_path = testcase.get('file', '')

                                    failure = testcase.find('failure')
                                    skipped = testcase.find('skipped')

                                    status = 'passed'
                                    error_msg = ''
                                    traceback = ''

                                    if failure is not None:
                                        status = 'failed'
                                        msg = failure.text or failure.get('message', '') or ''
                                        error_msg, traceback = parse_failure_message(msg)
                                    elif skipped is not None:
                                        status = 'skipped'
                                        msg = skipped.text or skipped.get('message', '') or ''
                                        error_msg = msg[:500] if msg else 'skipped'

                                    xpu_test_cases.append({
                                        'prefix': prefix,
                                        'test_file': file_path,
                                        'test_class': classname,
                                        'test_case': name,
                                        'status': status,
                                        'error_msg': error_msg,
                                        'traceback': traceback
                                    })
                                    total_cases += 1
                                total_files += 1
                        except Exception:
                            pass

    log(f"  XPU CI: {total_cases} test cases from {total_files} XML files")
    return xpu_test_cases


def create_test_cases_all_excel(stock_cases, xpu_cases):
    """Create test_cases_all.xlsx with stock and torch-xpu-ops sheets."""
    output_path = os.path.join(RESULT_DIR, 'test_cases_all.xlsx')
    wb = openpyxl.Workbook()

    ws_stock = wb.active
    ws_stock.title = 'stock'
    stock_headers = ['Test File', 'Test Class', 'Test Case', 'Status', 'Error Message', 'Traceback']
    for col, header in enumerate(stock_headers, 1):
        write_by_name(ws_stock, 1, header, header)

    log(f"  Writing {len(stock_cases)} stock test cases...")
    for row_idx, tc in enumerate(stock_cases, 2):
        write_by_name(ws_stock, row_idx, 'Test File', tc.get('test_file', ''))
        write_by_name(ws_stock, row_idx, 'Test Class', tc.get('test_class', ''))
        write_by_name(ws_stock, row_idx, 'Test Case', tc.get('test_case', ''))
        write_by_name(ws_stock, row_idx, 'Status', tc.get('status', ''))
        write_by_name(ws_stock, row_idx, 'Error Message', tc.get('error_msg', ''))
        write_by_name(ws_stock, row_idx, 'Traceback', tc.get('traceback', '')[:3000] if tc.get('traceback') else '')

    ws_xpu = wb.create_sheet('torch-xpu-ops')
    xpu_headers = ['Test File', 'Test Class', 'Test Case', 'XML Prefix', 'Status', 'Error Message', 'Traceback']
    for col, header in enumerate(xpu_headers, 1):
        write_by_name(ws_xpu, 1, header, header)

    log(f"  Writing {len(xpu_cases)} XPU test cases...")
    for row_idx, tc in enumerate(xpu_cases, 2):
        write_by_name(ws_xpu, row_idx, 'Test File', tc.get('test_file', ''))
        write_by_name(ws_xpu, row_idx, 'Test Class', tc.get('test_class', ''))
        write_by_name(ws_xpu, row_idx, 'Test Case', tc.get('test_case', ''))
        write_by_name(ws_xpu, row_idx, 'XML Prefix', tc.get('prefix', ''))
        write_by_name(ws_xpu, row_idx, 'Status', tc.get('status', ''))
        write_by_name(ws_xpu, row_idx, 'Error Message', tc.get('error_msg', ''))
        write_by_name(ws_xpu, row_idx, 'Traceback', tc.get('traceback', '')[:3000] if tc.get('traceback') else '')

    log(f"  Saving Excel file (skipping column width formatting for performance)...")
    wb.save(output_path)
    log(f"  Created: {output_path}")
    return output_path


def build_stock_status_map(stock_cases):
    """Build a map for fast lookup: (test_class, test_case) -> status, error_msg, traceback."""
    status_map = {}
    for tc in stock_cases:
        key = (tc.get('test_class', ''), tc.get('test_case', ''))
        status_map[key] = {
            'status': tc.get('status', ''),
            'error_msg': tc.get('error_msg', ''),
            'traceback': tc.get('traceback', '')
        }
    return status_map


def build_xpu_status_map(xpu_cases):
    """Build a map for fast lookup: (test_class, test_case, prefix) -> status, error_msg, traceback."""
    status_map = {}
    for tc in xpu_cases:
        key = (tc.get('test_class', ''), tc.get('test_case', ''), tc.get('prefix', ''))
        status_map[key] = {
            'status': tc.get('status', ''),
            'error_msg': tc.get('error_msg', ''),
            'traceback': tc.get('traceback', '')
        }
    return status_map


def normalize_class_name(class_name):
    """Extract short class name from various formats."""
    if not class_name:
        return None
    basename = class_name.split('.')[-1] if '.' in class_name else class_name
    short = basename.replace('XPU', '').replace('Tests', '').replace('Test', '')
    if short and short != basename and short.strip():
        return short.strip()
    return basename


def extract_test_case_from_path(test_file, test_class=None, test_case=None):
    """
    Extract test_class and test_case from test_file if they are missing.
    Handles pytest format: path/to/file.py::ClassName::method_name
    Returns: (test_class, test_case)
    """
    if not test_file:
        return test_class, test_case
    
    test_file = str(test_file).strip()
    
    if '::' not in test_file:
        return test_class, test_case
    
    parts = test_file.split('::')
    if len(parts) >= 3:
        extracted_class = parts[1].strip()
        extracted_method = parts[2].strip()
        if extracted_method and not test_case:
            test_case = extracted_method
        if extracted_class and not test_class:
            test_class = extracted_class
    elif len(parts) == 2:
        extracted_class = parts[1].strip()
        if extracted_class and not test_class:
            test_class = extracted_class
    
    return test_class, test_case


def normalize_file_path(file_path):
    """Get basename of file path."""
    if not file_path:
        return None
    return file_path.split('/')[-1].replace('.py', '').replace('_xpu', '').replace('_cuda', '').strip()


def load_test_cases_all():
    """Load test_cases_all.xlsx and build lookup maps for stock and xpu results."""
    test_cases_all_path = os.path.join(RESULT_DIR, 'test_cases_all.xlsx')

    if not os.path.exists(test_cases_all_path):
        log(f"  Warning: {test_cases_all_path} not found, will collect from XML")
        return None, None, None, None

    wb_all = openpyxl.load_workbook(test_cases_all_path)

    stock_case_map = {}
    stock_short_class_map = {}
    stock_classes = set()

    if 'stock' in wb_all.sheetnames:
        ws_stock = wb_all['stock']
        for row in range(2, ws_stock.max_row + 1):
            data = row_dict(ws_stock, row)
            test_file = data.get('Test File')
            test_class = data.get('Test Class')
            test_case = data.get('Test Case')
            if test_class and test_case:
                short_class = normalize_class_name(test_class)
                stock_case_map[(test_class, test_case)] = {
                    'status': data.get('Status') or '',
                    'error_msg': data.get('Error Message') or '',
                    'traceback': data.get('Traceback') or ''
                }
                stock_classes.add(test_class)
                stock_classes.add(short_class)
                if short_class != test_class:
                    stock_short_class_map[(short_class, test_case)] = stock_case_map[(test_class, test_case)]
                else:
                    stock_short_class_map[(test_class, test_case)] = stock_case_map[(test_class, test_case)]

    xpu_case_map = {}
    xpu_short_class_map = {}
    xpu_classes = set()

    if 'torch-xpu-ops' in wb_all.sheetnames:
        ws_xpu = wb_all['torch-xpu-ops']
        for row in range(2, ws_xpu.max_row + 1):
            data = row_dict(ws_xpu, row)
            test_file = data.get('Test File')
            test_class = data.get('Test Class')
            test_case = data.get('Test Case')
            if test_class and test_case:
                short_class = normalize_class_name(test_class)
                xpu_case_map[(test_class, test_case)] = {
                    'status': data.get('Status') or '',
                    'error_msg': data.get('Error Message') or '',
                    'traceback': data.get('Traceback') or ''
                }
                xpu_classes.add(test_class)
                xpu_classes.add(short_class)
                if short_class != test_class:
                    xpu_short_class_map[(short_class, test_case)] = xpu_case_map[(test_class, test_case)]
                else:
                    xpu_short_class_map[(short_class, test_case)] = xpu_case_map[(test_class, test_case)]

    return (stock_case_map, stock_short_class_map, stock_classes), (xpu_case_map, xpu_short_class_map, xpu_classes)


def pass1_match_ci_results(ws, output_path):
    """
    PASS 1: Create test_cases_all.xlsx, collect stock & xpu CI results, match CI.

    Reads test results from test_cases_all.xlsx 'stock' and 'torch-xpu-ops' sheets.

    Updates:
        Col 8: Error Message
        Col 9: Traceback
        Col 12: XPU Status
        Col 13: Stock Status

    Column mapping for Test Cases sheet:
        Col 1: Issue ID
        Col 2: Test Reproducer
        Col 3: Test Type
        Col 4: Test File
        Col 5: Origin Test File
        Col 6: Test Class
        Col 7: Test Case
        Col 8: Error Message
        Col 9: Traceback
        Col 12: XPU Status
        Col 13: Stock Status
        Col 16: CUDA Case Exist
        Col 17: XPU Case Exist
        Col 18: case_existence_comments
        Col 19: can_enable_on_xpu
        Col 20: duplicated_issue
    """

    for name in ["Error Message", "Traceback", "XPU Status", "Stock Status", "No Match Reason"]:
        ensure_col(ws, name)

    # Collect CI test cases directly without creating slow Excel intermediate
    log(f"  Collecting CI test cases from XML files...")
    stock_cases = collect_stock_test_cases()
    xpu_cases = collect_torch_xpu_ops_test_cases()

    # Build lookup maps directly from collected cases (skip Excel creation for performance)
    log(f"  Building lookup maps...")
    stock_case_map = build_stock_status_map(stock_cases)
    
    # For XPU, build map WITHOUT prefix since matching just needs test class/case
    xpu_case_map = {}
    for tc in xpu_cases:
        key = (tc.get('test_class', ''), tc.get('test_case', ''))
        xpu_case_map[key] = {
            'status': tc.get('status', ''),
            'error_msg': tc.get('error_msg', ''),
            'traceback': tc.get('traceback', '')
        }
    
    # Also build short class name maps for fuzzy matching
    stock_short_class_map = {}
    stock_classes = set()
    for tc in stock_cases:
        test_class = tc.get('test_class', '')
        test_case = tc.get('test_case', '')
        if test_class and test_case:
            short_class = normalize_class_name(test_class)
            stock_classes.add(test_class)
            stock_classes.add(short_class)
            if short_class != test_class:
                stock_short_class_map[(short_class, test_case)] = stock_case_map[(test_class, test_case)]
            else:
                stock_short_class_map[(test_class, test_case)] = stock_case_map[(test_class, test_case)]
    
    xpu_short_class_map = {}
    xpu_classes = set()
    for tc in xpu_cases:
        test_class = tc.get('test_class', '')
        test_case = tc.get('test_case', '')
        if test_class and test_case:
            short_class = normalize_class_name(test_class)
            xpu_classes.add(test_class)
            xpu_classes.add(short_class)
            key = (test_class, test_case)
            if short_class != test_class:
                xpu_short_class_map[(short_class, test_case)] = xpu_case_map[key]
            else:
                xpu_short_class_map[(test_class, test_case)] = xpu_case_map[key]
    
    stock_lookup = (stock_case_map, stock_short_class_map, stock_classes)
    xpu_lookup = (xpu_case_map, xpu_short_class_map, xpu_classes)
    stock_case_map, stock_short_class_map, stock_classes = stock_lookup
    xpu_case_map, xpu_short_class_map, xpu_classes = xpu_lookup

    issues_needing_llm = {}
    total = ws.max_row - 1
    found_count = 0
    not_found_count = 0

    log("  Matching CI results from test_cases_all.xlsx...")
    rows_to_move_to_others = []
    
    for i, row_idx in enumerate(range(2, ws.max_row + 1), 1):
        test_file = cell_by_name(ws, row_idx, 'Test File').value
        test_class = cell_by_name(ws, row_idx, 'Test Class').value
        test_case = cell_by_name(ws, row_idx, 'Test Case').value
        issue_id = cell_by_name(ws, row_idx, 'Issue ID').value
        test_type = cell_by_name(ws, row_idx, 'Test Type').value

        if not test_case:
            test_class, test_case = extract_test_case_from_path(test_file, test_class, test_case)
            if test_case:
                write_by_name(ws, row_idx, 'Test Class', test_class if test_class else '')
                write_by_name(ws, row_idx, 'Test Case', test_case)
        
        if not test_case:
            test_type_str = test_type if test_type else 'unknown'
            if test_type_str.lower() != 'e2e':
                rows_to_move_to_others.append((row_idx, issue_id, 'No unittest test case found'))
            continue

        xpu_status = None
        xpu_error_msg = None
        xpu_traceback = None
        stock_status = None
        stock_error_msg = None
        stock_traceback = None

        short_class = normalize_class_name(test_class)

        if test_class and test_case:
            xpu_key = (test_class, test_case)
            if xpu_key in xpu_case_map:
                xpu_status = xpu_case_map[xpu_key]['status']
                xpu_error_msg = xpu_case_map[xpu_key]['error_msg']
                xpu_traceback = xpu_case_map[xpu_key]['traceback']
            elif xpu_short_class_map.get((short_class, test_case)):
                xpu_status = xpu_short_class_map[(short_class, test_case)]['status']
                xpu_error_msg = xpu_short_class_map[(short_class, test_case)]['error_msg']
                xpu_traceback = xpu_short_class_map[(short_class, test_case)]['traceback']

            stock_key = (test_class, test_case)
            if stock_key in stock_case_map:
                stock_status = stock_case_map[stock_key]['status']
                stock_error_msg = stock_case_map[stock_key]['error_msg']
                stock_traceback = stock_case_map[stock_key]['traceback']
            elif stock_short_class_map.get((short_class, test_case)):
                stock_status = stock_short_class_map[(short_class, test_case)]['status']
                stock_error_msg = stock_short_class_map[(short_class, test_case)]['error_msg']
                stock_traceback = stock_short_class_map[(short_class, test_case)]['traceback']

        write_by_name(ws, row_idx, 'XPU Status', xpu_status if xpu_status else '')
        write_by_name(ws, row_idx, 'Stock Status', stock_status if stock_status else '')

        if xpu_error_msg:
            write_by_name(ws, row_idx, 'Error Message', xpu_error_msg[:3000] if xpu_error_msg else '')
        elif stock_error_msg:
            write_by_name(ws, row_idx, 'Error Message', stock_error_msg[:3000] if stock_error_msg else '')

        if xpu_traceback:
            write_by_name(ws, row_idx, 'Traceback', xpu_traceback[:3000] if xpu_traceback else '')
        elif stock_traceback and not cell_by_name(ws, row_idx, 'Traceback').value:
            write_by_name(ws, row_idx, 'Traceback', stock_traceback[:3000] if stock_traceback else '')

        if xpu_status or stock_status:
            found_count += 1
            write_by_name(ws, row_idx, 'No Match Reason', '')
        else:
            not_found_count += 1
            issues_needing_llm[issue_id] = {
                'test_file': test_file,
                'test_class': test_class,
                'test_case': test_case,
                'origin_test_file': cell_by_name(ws, row_idx, 'Origin Test File').value
            }
            if not test_case:
                write_by_name(ws, row_idx, 'No Match Reason', 'No test case')
            elif not test_class:
                write_by_name(ws, row_idx, 'No Match Reason', 'No test class')
            elif not xpu_status and not stock_status:
                write_by_name(ws, row_idx, 'No Match Reason', 'No test case')

        if i % 500 == 0:
            log(f"    Progress: {i}/{total}")

    log(f"  PASS 1 complete: {found_count} matched, {not_found_count} not found")
    
    if rows_to_move_to_others:
        log(f"  Moving {len(rows_to_move_to_others)} rows with no test case to Others sheet...")
        workbook = ws.parent
        
        if 'Others' not in workbook.sheetnames:
            log("    Warning: Others sheet not found, skipping move")
        else:
            ws_others = workbook['Others']
            
            for row_idx, issue_id, reason in rows_to_move_to_others:
                try:
                    issue_id_cell = cell_by_name(ws, row_idx, 'Issue ID').value
                    title_cell = ws.cell(row_idx, 2).value if row_idx <= ws.max_row else None
                    test_repro_cell = cell_by_name(ws, row_idx, 'Test Reproducer').value
                    
                    next_row = ws_others.max_row + 1
                    write_by_name(ws_others, next_row, 'ID', issue_id_cell if issue_id_cell else '')
                    write_by_name(ws_others, next_row, 'Title', title_cell if title_cell else test_repro_cell if test_repro_cell else '')
                    write_by_name(ws_others, next_row, 'reproduce step', reason)
                except Exception as e:
                    log(f"    Warning: Failed to move row {row_idx} to Others: {e}")
            
            log(f"  Deleting {len(rows_to_move_to_others)} rows from Test Cases sheet...")
            for row_idx in sorted(rows_to_move_to_others, key=lambda x: x[0], reverse=True):
                ws.delete_rows(row_idx[0], 1)
    
    return issues_needing_llm
