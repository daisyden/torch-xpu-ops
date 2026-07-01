"""Merge Phase 4b wave results into Issues sheet.

Loads all 305 result_<id>.json under agent_space/phase4b/wave{1..7}/
and writes their action_TBD / action_reason / owner_transferred arrays
into the Issues sheet.

Separator for Phase 4b within-result joins is " | " (pipe) because
action_reason sentences contain commas. Each agent result is the
authoritative value for its row — re-running the agent for an issue
REPLACES the prior row contents (overwrite, not append). This is what
the agent's idempotent re-discovery implies: the latest run wins.

INVARIANT (post-merge): For any row containing "No action — investigate
further" in action_TBD, owner_transferred MUST NOT equal Reporter when
Assignee is empty. The merge enforces this by clearing such values
after writing agent emissions, guarding against stale data and agent
regressions.
"""
import json
import os
import sys
from pathlib import Path
from collections import defaultdict

import openpyxl

COMMON_DIR = Path(__file__).resolve().parents[2] / "_common"
if str(COMMON_DIR) not in sys.path:
    sys.path.insert(0, str(COMMON_DIR))
from header_utils import cell_by_name, row_dict, write_by_name  # type: ignore[reportMissingImports] # noqa: E402
from paths import RESULT_DIR, AGENT_SPACE  # type: ignore[reportMissingImports] # noqa: E402

EXCEL  = RESULT_DIR / "torch_xpu_ops_issues.xlsx"
WAVES  = Path(os.environ.get("PHASE4B_WAVES", AGENT_SPACE / "phase4b"))

SEP = " | "

# ---- load all 305 results --------------------------------------------------
results = {}
wave_dirs = sorted(WAVES.glob("wave*"), key=lambda p: int(p.name.replace("wave", "")))
for wave_dir in wave_dirs:
    for f in sorted(wave_dir.glob("result_*.json")):
        with open(f) as fh:
            r = json.load(fh)
        iid = r["issue_number"]
        if iid in results:
            raise RuntimeError(f"duplicate issue {iid}")
        results[iid] = r

print(f"loaded {len(results)} Phase 4b results")

# ---- per-status counters ---------------------------------------------------
by_status = defaultdict(int)
for r in results.values():
    by_status[r.get("validation_status", "?")] += 1
print(f"status breakdown: {dict(by_status)}")

# ---- open Excel ------------------------------------------------------------
wb  = openpyxl.load_workbook(EXCEL)
ws  = wb["Issues"]

INVESTIGATE_VERB = "No action — investigate further"

def join_dedup(items):
    seen, out = set(), []
    for x in items:
        if not x:
            continue
        s = x.strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return SEP.join(out) if out else None

updated = 0
missing = []
for row in ws.iter_rows(min_row=2):
    row_idx = row[0].row
    iid = cell_by_name(ws, row_idx, "Issue ID").value
    if iid is None:
        continue
    r = results.get(iid)
    if r is None:
        continue
    acts = r.get("action_TBD") or []
    rsns = r.get("action_reason") or []
    owns = r.get("owner_transferred") or []
    if isinstance(acts, str): acts = [acts]
    if isinstance(rsns, str): rsns = [rsns]
    if isinstance(owns, str): owns = [owns]
    if not (acts or rsns or owns):
        continue
    write_by_name(ws, row_idx, "action_TBD", join_dedup(acts))
    write_by_name(ws, row_idx, "action_reason", join_dedup(rsns))
    cell_by_name(ws, row_idx, "owner_transferred").value = join_dedup(owns)
    updated += 1

# Any result not matched to an Excel row?
excel_ids = {row_dict(ws, row_idx).get("Issue ID") for row_idx in range(2, ws.max_row + 1)}
missing   = sorted(set(results) - excel_ids)

# ---- INVARIANT: Reporter is NEVER a valid owner_transferred ----------------
# Rule (per AGENT_INSTRUCTIONS.md DERIVATION RULE — owner_transferred):
#   owner_transferred = Reporter is ONLY valid when action_TBD is purely a
#   combination of the v4.17 carve-out verbs (close/verify family) OR the
#   Phase 4e D3 Reporter-verify/re-investigate verbs. For ANY other verb
#   (investigate further, Land PR, Address CI, Wait for fix PR, Resolve
#   review, @<user> response, Submit issue, reassess fix path, etc.),
#   owner_transferred MUST be sourced from Assignee | comment-AR-owner |
#   blank — never from Reporter.
# Action: when owner == Reporter and verb is NOT a pure carve-out, overwrite
# with Assignee if non-empty, else clear.
CARVE_OUT_VERB_FRAGMENTS = (
    "Verify fix from merged PR",
    "Close the fixed issue",
    "label_not_target_and_close",
    "label not_target and close",
    "close_as_not_planned",
    "Confirm fix and close",
    "Reporter to verify the fix",
    "Reporter to re-investigate",
)

def _is_pure_carveout(act_value):
    tokens = [t.strip() for t in (act_value or "").split("|") if t.strip()]
    return bool(tokens) and all(
        any(frag in tok for frag in CARVE_OUT_VERB_FRAGMENTS) for tok in tokens
    )

# INVARIANT (AGENT_INSTRUCTIONS.md v4.17): pure close/verify carve-out rows are
# Reporter-owned even when an Assignee exists. Merge is authoritative here.
carveout_reporter = []
for row in ws.iter_rows(min_row=2):
    row_idx = row[0].row
    act = (cell_by_name(ws, row_idx, "action_TBD").value or "")
    if not _is_pure_carveout(act):
        continue
    rep = (cell_by_name(ws, row_idx, "Reporter").value or "").strip()
    if not rep:
        continue
    own = (cell_by_name(ws, row_idx, "owner_transferred").value or "").strip()
    if own != rep:
        write_by_name(ws, row_idx, "owner_transferred", rep)
        carveout_reporter.append(cell_by_name(ws, row_idx, "Issue ID").value)
if carveout_reporter:
    print(f"invariant: set owner_transferred=Reporter on {len(carveout_reporter)} pure carve-out rows")
    print(f"  IDs: {carveout_reporter[:20]}{'...' if len(carveout_reporter) > 20 else ''}")
cleared = []
reassigned = []
for row in ws.iter_rows(min_row=2):
    row_idx = row[0].row
    own = (cell_by_name(ws, row_idx, "owner_transferred").value or "").strip()
    rep = (cell_by_name(ws, row_idx, "Reporter").value or "").strip()
    if not own or own != rep:
        continue
    asg = (cell_by_name(ws, row_idx, "Assignee").value or "").strip()
    if asg and asg == rep:
        continue
    act = (cell_by_name(ws, row_idx, "action_TBD").value or "")
    tokens = [t.strip() for t in act.split("|") if t.strip()]
    if tokens and all(any(frag in tok for frag in CARVE_OUT_VERB_FRAGMENTS) for tok in tokens):
        continue
    issue_id = cell_by_name(ws, row_idx, "Issue ID").value
    if asg and asg.lower() != "none":
        write_by_name(ws, row_idx, "owner_transferred", asg)
        reassigned.append(issue_id)
    else:
        cell_by_name(ws, row_idx, "owner_transferred").value = None
        cleared.append(issue_id)
if reassigned:
    print(f"invariant: reassigned owner_transferred=Reporter -> Assignee on {len(reassigned)} non-carve-out rows")
    print(f"  IDs: {reassigned[:20]}{'...' if len(reassigned) > 20 else ''}")
if cleared:
    print(f"invariant: cleared owner_transferred on {len(cleared)} non-carve-out rows where owner==Reporter and Assignee was empty")
    print(f"  IDs: {cleared[:20]}{'...' if len(cleared) > 20 else ''}")

wb.save(EXCEL)
print(f"merged {updated} issues into Issues sheet")
if missing:
    print(f"WARN: {len(missing)} result IDs not in Excel: {missing[:10]}...")
else:
    print("all 305 result IDs mapped to Excel rows")
