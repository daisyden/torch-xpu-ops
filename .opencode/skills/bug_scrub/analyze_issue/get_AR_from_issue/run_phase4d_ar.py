"""Phase 4d: Derive AR bucket(s) from Phase 4a-4c outputs.

Deterministic post-processing. Reads:
  action_TBD, action_reason, Root Cause, Fix Approach, Assignee, Reporter,
  and (v4.16) Test Cases sheet `XPU Status` / `Stock Status` columns.
Writes `AR` column (creates if missing). Multi-value `; `-delimited.

v4.17 changes (May 24, 2026):
- Fixed 4 systemic bugs surfaced by #3433 review:
  (1) `Verify fix from merged PR <ref> and close` removed from
      LAND_PR_VERBS; now exclusively routes to `Verify` bucket via
      fires_verify() Path A (explicit verb + MERGED PR via pr_cache or
      live gh fallback; Reporter==owner_transferred constraint removed
      for Path A). Path B retains legacy silent inference.
  (2) load_pr_analysis_cache() now reads the correct JSON key
      `pr_candidates` (was reading nonexistent `pr_analysis` for 11
      versions, silently returning empty cache; v4.15 Verify=6 came
      entirely from the live gh fallback). Entries normalized to
      {state, url, repo, pr_number, relationship, verdict}.
  (3) owner_transferred carve-out: for rows whose action_TBD is PURELY
      a combination of close/verify verbs (Verify fix from merged PR,
      Close the fixed issue, label_not_target_and_close,
      close_as_not_planned, Confirm fix and close), owner_transferred
      MUST be the Reporter (next-actor for close/verify is the Reporter
      signing off, not the maintainer/tracker). 25 retroactive rewrites
      applied (21 Verify-verb + 4 Close/Skip). For all other verbs the
      legacy rule still applies (Assignee | comment-AR-owner | blank;
      never Reporter).
  (4) Need Owner suppressed when buckets <= {Close/Skip, Verify}
      (reporter-owned scope). Does NOT suppress when paired with
      Land PR / Need Response / Wait for PR (legacy Assignee gating).
- Final v4.17 AR counts: Close/Skip=10, Need Owner=26 (-2),
  Land PR=101 (-23), Wait for PR=30, Need Response=140 (+3),
  Verify=34 (+28), UNCLASSIFIED=0.

v4.16 changes (May 24, 2026):
- Added RULE 1 audit guard. When `Close the fixed issue` is present in
  `action_TBD`, Phase 4d cross-checks every Test Cases row for that issue
  against strict RULE 1 (`XPU Status in {passed, fixed}` AND
  `Stock Status not in {fail, error, timeout}`). On violation, Close/Skip
  is suppressed and the row is routed to Need Response instead. Issues
  with zero Test Cases sheet rows are treated as `out_of_scope`
  (alt-path close: manual-verification / performance-investigation /
  won't-fix) and Close/Skip remains valid.
- Added pending-ack guard. Parallel to RULE 1, Phase 4d scans
  `action_reason` for ack-pending admissions (`pending @X's ack`,
  `awaiting @X's confirmation`, `pending a final verification`,
  `awaiting confirmation from @X`, and the approval/sign-off/response
  variants). On match, Close/Skip is suppressed even when RULE 1 passes,
  because Phase 4b's own reason field admits the close is gated on a
  still-open maintainer ack. Spot-fixed #2766 (BBBela's 2026-05-07 close
  request awaiting @EikanWang for 17 days).

v4.15 changes (May 24, 2026):
- D: Need Response now matches any `@<user>: please <action>` template.
- C: `RETRIAGE_PRS` trigger dropped (3 prior rows rewritten as real verbs).
- F: `Verify` bucket has a live `gh pr view` fallback (24h-TTL cache at
  agent_space/phase4d_verify_pr_state_cache.json) used when an issue has
  no pr_analysis entry in the wave*/result_*.json cache.
- H: canonical Wait-for-PR verb is `Wait for fix PR`; `Wait for PR` and
  `Monitor ...` are retained as defensive aliases.

Spec deviation still active (planned follow-up):
- The live `gh issue view` no_response leg is OMITTED (300 issues = 300
  gh calls). Other Need-Response triggers (Request info, (>1 week)
  staleness, maintainer-question rewrites) still fire correctly.
"""
from pathlib import Path
import re
import json
import shutil
import subprocess
import sys
import time
from openpyxl import load_workbook

COMMON_DIR = Path(__file__).resolve().parents[2] / "_common"
if str(COMMON_DIR) not in sys.path:
    sys.path.insert(0, str(COMMON_DIR))
from header_utils import ensure_col, get_col, header_index, row_dict, write_by_name  # type: ignore[reportMissingImports] # noqa: E402
from paths import RESULT_DIR, AGENT_SPACE  # type: ignore[reportMissingImports] # noqa: E402

XLSX_PATH = RESULT_DIR / "torch_xpu_ops_issues.xlsx"
BACKUP = RESULT_DIR / "torch_xpu_ops_issues_bk_before_phase4d.xlsx"
PHASE4B_ROOT = AGENT_SPACE / "phase4b"
VERIFY_CACHE_PATH = AGENT_SPACE / "phase4d_verify_pr_state_cache.json"
VERIFY_CACHE_TTL_SEC = 24 * 3600

CANONICAL_ORDER = [
    "Close/Skip",
    "Need Owner",
    "Land PR",
    "Wait for PR",
    "Wait for dependency fix",
    "Need Response",
    "Need check case existence",
    "Add label",
    "Submit issue",
    "Verify",
]

PR_REF_RE = re.compile(
    r"(?:https://github\.com/[\w.-]+/[\w.-]+/pull/\d+|[\w.-]+/[\w.-]+#\d+|\bPR\s*#?\d+)",
    re.IGNORECASE,
)

LAND_PR_VERBS = (
    "Land PR ",
    "Resolve unresolved review comments on PR ",
    "Address CI failures on PR ",
    "Wait for review on PR ",
)

CASE_EXISTENCE_KEYWORDS = (
    "case removed", "test renamed", "missing case", "case missing", "test removed",
)


def is_blank_assignee(s):
    if s is None:
        return True
    s = str(s).strip()
    if not s:
        return True
    return s.lower() in {"none", "unassigned"}


def load_pr_analysis_cache():
    # v4.17: phase4b wave files use key "pr_candidates" (not "pr_analysis" -
    # the v4.6 code used the wrong key, so the cache was silently empty for
    # 11 versions, and Verify=6 in v4.15/v4.16 came entirely from the live
    # gh fallback). Normalize entries to {state, url, repo, pr_number,
    # relationship, verdict} for downstream consumers.
    cache = {}
    for p in sorted(PHASE4B_ROOT.glob("wave*/result_*.json")):
        try:
            d = json.loads(p.read_text())
        except Exception:
            continue
        n = d.get("issue_number")
        if not n:
            m = re.search(r"result_(\d+)", p.name)
            if m:
                n = int(m.group(1))
        if not n:
            continue
        entries = []
        for pc in d.get("pr_candidates", []) or []:
            repo = pc.get("repo") or ""
            num = pc.get("pr_number")
            entries.append({
                "state": (pc.get("live_state") or "").upper(),
                "url": f"https://github.com/{repo}/pull/{num}" if repo and num else "",
                "repo": repo,
                "pr_number": num,
                "relationship": pc.get("relationship") or "",
                "verdict": pc.get("verdict") or "",
            })
        cache[int(n)] = entries
    return cache


def split_actions(action_TBD):
    primary = [a.strip() for a in action_TBD.split(" | ") if a.strip()]
    fallback = [a.strip() for a in action_TBD.split("; ") if a.strip()]
    return fallback if len(fallback) > len(primary) else primary


RULE1_PASS_OK = {"passed", "fixed"}
RULE1_STOCK_BAD = {"fail", "error", "timeout"}


def load_test_cases_status(wb):
    status = {}
    for sname in ("Test Cases", "E2E Test Cases"):
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        H = {name: col - 1 for name, col in header_index(ws).items()}
        iid_col = H.get("Issue ID")
        xpu_col = H.get("XPU Status")
        stock_col = H.get("Stock Status")
        case_col = H.get("Test Case") or H.get("Model")
        if iid_col is None or xpu_col is None:
            continue
        for r in ws.iter_rows(min_row=2):
            v = r[iid_col].value
            if v is None:
                continue
            ids = []
            for tok in str(v).replace(";", ",").split(","):
                tok = tok.strip()
                if tok.isdigit():
                    ids.append(int(tok))
            if not ids:
                continue
            xpu = (str(r[xpu_col].value).strip().lower() if r[xpu_col].value is not None else "")
            stock = (str(r[stock_col].value).strip().lower() if stock_col is not None and r[stock_col].value is not None else "")
            case = r[case_col].value if case_col is not None else None
            for iid in ids:
                status.setdefault(iid, []).append({"xpu": xpu, "stock": stock, "case": case, "sheet": sname})
    return status


def rule1_audit(issue_id, tc_status):
    try:
        n = int(issue_id)
    except (TypeError, ValueError):
        return "out_of_scope", "non-integer issue id"
    cases = tc_status.get(n, [])
    if not cases:
        return "out_of_scope", "no Test Cases sheet rows (alt-path close eligible)"
    violations = []
    for c in cases:
        xpu_ok = c["xpu"] in RULE1_PASS_OK
        stock_bad = c["stock"] in RULE1_STOCK_BAD
        if (not xpu_ok) or stock_bad:
            violations.append(c)
    if not violations:
        return "ok", f"all {len(cases)} cases satisfy RULE 1"
    parts = []
    for v in violations[:3]:
        parts.append(f"[{v['sheet']}] {str(v['case'])[:40]} xpu={v['xpu']!r} stock={v['stock']!r}")
    if len(violations) > 3:
        parts.append(f"+{len(violations)-3} more")
    return "violation", "; ".join(parts)


def fires_close_skip(actions):
    for a in actions:
        if a in {"Close the fixed issue", "label_not_target_and_close", "Skip issue"}:
            return True
        if a.startswith("Skip issue"):
            return True
    return False


def has_close_fixed_verb(actions):
    return any(a == "Close the fixed issue" for a in actions)


_PENDING_ACK_RE = re.compile(
    r"(pending\s+@\w[\w-]*'?s?\s+ack"
    r"|pending\s+@\w[\w-]*'?s?\s+(?:approval|sign-?off|confirmation|response|reply|review)"
    r"|awaiting\s+@\w[\w-]*'?s?\s+(?:ack|approval|sign-?off|confirmation|response|reply|review)"
    r"|pending\s+a\s+final\s+verification"
    r"|awaiting\s+confirmation\s+from\s+@\w[\w-]*)",
    re.IGNORECASE,
)


def pending_ack_audit(action_reason):
    m = _PENDING_ACK_RE.search(action_reason or "")
    if not m:
        return "ok", ""
    return "pending_ack", m.group(1)


def fires_land_pr(actions):
    for a in actions:
        for verb in LAND_PR_VERBS:
            if a.startswith(verb) and PR_REF_RE.search(a):
                return True
    return False


def fires_wait_for_pr(actions):
    for a in actions:
        # v4.15: canonical verb is `Wait for fix PR`. `Wait for PR` (v4.14)
        # is retained as a defensive alias; `Monitor ...` (v4.6 prefix) is
        # also retained for backwards compatibility with old data.
        if a.startswith("Monitor"):
            return True
        if a == "Wait for fix PR":
            return True
        if a == "Wait for PR":
            return True
    return False


def fires_need_response(actions):
    for a in actions:
        if a.startswith("Request info") or a.startswith("Need Response"):
            return True
        if "(>1 week)" in a:
            return True
        # v4.15 D: generic @<user>: please <action> template captures all
        # reporter/assignee-directed asks (reply/verify/re-run/confirm/test/...).
        if a.startswith("@") and "please " in a.lower():
            return True
        if "closed unmerged" in a:
            return True
        # v4.15 B1: bare `No action - investigate further` is intentionally
        # routed here so triage leads see these rows in the weekly stale-
        # request review. v4.15 C: `RETRIAGE_PRS` token is no longer emitted
        # by Phase 4b (3 prior rows rewritten); the rule below is dropped.
        if a.startswith("No action"):
            return True
    return False


def fires_case_existence(action_TBD, root_cause, fix_approach):
    if "check_case_avaliablity" in action_TBD or "check_case_availability" in action_TBD:
        return True
    pool = " ".join([action_TBD, root_cause, fix_approach]).lower()
    return any(kw in pool for kw in CASE_EXISTENCE_KEYWORDS)


def fires_wait_for_dependency_fix(actions):
    # v4.25 Phase 4e: AR bucket fires when action_TBD contains a verb of the
    # form `Wait for dependency fix <ref>` (with or without trailing repo+#N).
    for a in actions:
        if a.startswith("Wait for dependency fix"):
            return True
    return False


def fires_add_label(actions):
    # v4.26 Phase 4e D2: `Add label '<canonical>'` -> bucket "Add label".
    for a in actions:
        if a.startswith("Add label "):
            return True
    return False


def fires_submit_issue(actions):
    # v4.26 Phase 4e D3 no-ref: `Assignee to submit issue to <component>
    # upstream` -> bucket "Submit issue".
    for a in actions:
        if a.startswith("Assignee to submit issue to ") and " upstream" in a:
            return True
    return False


def fires_dep_verify(actions):
    # v4.26 Phase 4e D3 merged: `Reporter to verify the fix from <ref>...` ->
    # bucket "Verify" (in addition to existing pr_cache-based Verify path).
    for a in actions:
        if a.startswith("Reporter to verify the fix from "):
            return True
    return False


def fires_dep_need_response(actions):
    # v4.26 Phase 4e D3 closed-unmerged: `Reporter to re-investigate: upstream
    # ref <ref> was closed without resolving...` -> bucket "Need Response".
    for a in actions:
        if a.startswith("Reporter to re-investigate: upstream ref "):
            return True
    return False


def fires_verify(issue_id, reporter, owner_t, pr_cache, actions, live_cache):
    try:
        n = int(issue_id)
    except (TypeError, ValueError):
        return False

    # Path A (v4.17): explicit "Verify fix from merged PR <ref>" verb fires
    # Verify whenever the referenced PR is MERGED. Reporter/owner alignment
    # is irrelevant - the verb itself was emitted by phase4b as the signal.
    pr_analyses = pr_cache.get(n, [])
    for a in actions:
        if not a.startswith("Verify fix from merged PR "):
            continue
        verb_urls = set(_extract_pr_urls(a))
        for pa in pr_analyses:
            if pa.get("state") == "MERGED" and pa.get("url") in verb_urls:
                return True
        for url in verb_urls:
            if _live_pr_state(url, live_cache) == "MERGED":
                return True

    # Path B (v4.6 legacy): silent inference when Reporter is in the
    # owner_transferred set AND a merged fixing PR exists in cache, even
    # without an explicit Verify verb in action_TBD.
    if not reporter:
        return False
    owners = {o.strip().lower() for o in re.split(r"[,;|]", owner_t) if o.strip()}
    if reporter.strip().lower() not in owners:
        return False
    if any(pa.get("state") == "MERGED" for pa in pr_analyses):
        return True
    return False


_PR_URL_RE = re.compile(r"https://github\.com/[\w.-]+/[\w.-]+/pull/\d+")
_PR_SHORT_RE = re.compile(r"\b([\w.-]+/[\w.-]+)#(\d+)\b")


def _extract_pr_urls(text):
    urls = list(_PR_URL_RE.findall(text))
    for repo, num in _PR_SHORT_RE.findall(text):
        urls.append(f"https://github.com/{repo}/pull/{num}")
    return urls


def _live_pr_state(url, live_cache):
    now = time.time()
    entry = live_cache.get(url)
    if entry and (now - entry.get("fetched_at", 0)) < VERIFY_CACHE_TTL_SEC:
        return entry.get("state")
    try:
        out = subprocess.run(
            ["gh", "pr", "view", url, "--json", "state,mergedAt"],
            capture_output=True, text=True, timeout=20, check=False,
        )
        if out.returncode != 0:
            live_cache[url] = {"state": None, "fetched_at": now, "error": out.stderr.strip()[:200]}
            return None
        data = json.loads(out.stdout or "{}")
        state = (data.get("state") or "").upper() or None
        live_cache[url] = {"state": state, "fetched_at": now, "mergedAt": data.get("mergedAt")}
        return state
    except Exception as e:
        live_cache[url] = {"state": None, "fetched_at": now, "error": str(e)[:200]}
        return None


def _load_live_cache():
    if VERIFY_CACHE_PATH.exists():
        try:
            return json.loads(VERIFY_CACHE_PATH.read_text())
        except Exception:
            return {}
    return {}


def _save_live_cache(cache):
    VERIFY_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    VERIFY_CACHE_PATH.write_text(json.dumps(cache, indent=2, sort_keys=True))


def derive_ar(row_vals, headers, pr_cache, live_cache, tc_status, rule1_log, pending_ack_log):
    def col(name):
        if name not in headers:
            return ""
        v = row_vals[headers[name]]
        return "" if v is None else str(v)

    issue_id = col("Issue ID")
    action_TBD = col("action_TBD")
    action_reason = col("action_reason")
    root_cause = col("Root Cause")
    fix_approach = col("Fix Approach")
    assignee = col("Assignee")
    reporter = col("Reporter")
    owner_t = col("owner_transferred")
    actions = split_actions(action_TBD)

    # v4.16: RULE 1 + pending-ack audits. When `Close the fixed issue` is
    # present, two parallel checks can block Close/Skip and force Need Response:
    #   (a) RULE 1 violation: one or more Test Cases rows fail strict RULE 1
    #       (Phase 4a should have caught this; Phase 4d is the safety net).
    #   (b) pending-ack admission: action_reason contains "pending @X's ack",
    #       "awaiting @X's confirmation", "pending a final verification", etc.
    #       Phase 4b emitted Close but admitted in its own reason field that
    #       the close is gated on a still-open maintainer ack. We refuse to
    #       Close/Skip until that ack lands.
    # `out_of_scope` (no Test Cases rows) plus no pending-ack admission ->
    # legitimate alt-path close (manual-verification / perf-investigation /
    # won't-fix); Close/Skip remains valid.
    rule1_blocks_close = False
    pending_ack_blocks_close = False
    if has_close_fixed_verb(actions):
        verdict, detail = rule1_audit(issue_id, tc_status)
        if verdict == "violation":
            rule1_blocks_close = True
        rule1_log.append((issue_id, verdict, detail))
        ack_verdict, ack_detail = pending_ack_audit(action_reason)
        if ack_verdict == "pending_ack":
            pending_ack_blocks_close = True
            pending_ack_log.append((issue_id, ack_detail))

    block_close = rule1_blocks_close or pending_ack_blocks_close

    buckets = set()
    if fires_close_skip(actions) and not block_close:
        buckets.add("Close/Skip")
    land = fires_land_pr(actions)
    if land:
        buckets.add("Land PR")
    if not land and fires_wait_for_pr(actions):
        buckets.add("Wait for PR")
    if fires_wait_for_dependency_fix(actions):
        buckets.add("Wait for dependency fix")
    if fires_add_label(actions):
        buckets.add("Add label")
    if fires_submit_issue(actions):
        buckets.add("Submit issue")
    if not land and (fires_need_response(actions) or block_close or fires_dep_need_response(actions)):
        buckets.add("Need Response")
    if fires_case_existence(action_TBD, root_cause, fix_approach):
        buckets.add("Need check case existence")
    if fires_verify(issue_id, reporter, owner_t, pr_cache, actions, live_cache) or fires_dep_verify(actions):
        buckets.add("Verify")

    # v4.17 + v4.26: Need Owner fires when Assignee is blank, EXCEPT when
    # the only buckets are Reporter-owned (Close/Skip, Verify) or dep-audit
    # buckets that already routed the next action to a specific role
    # (Submit issue -> Assignee, Add label -> reporter/maintainer pair).
    # For Wait for dependency fix and the other open-PR buckets the legacy
    # rule still applies: blank Assignee means a maintainer must be assigned.
    if is_blank_assignee(assignee):
        reporter_owned_only = buckets and buckets <= {"Close/Skip", "Verify"}
        if not reporter_owned_only:
            buckets.add("Need Owner")

    ordered = [b for b in CANONICAL_ORDER if b in buckets]
    return "; ".join(ordered)


def main():
    if not XLSX_PATH.exists():
        print(f"ERROR: xlsx not found at {XLSX_PATH}", file=sys.stderr)
        sys.exit(1)
    shutil.copy2(XLSX_PATH, BACKUP)
    print(f"Backed up workbook -> {BACKUP.name}")

    wb = load_workbook(XLSX_PATH)
    ws = wb["Issues"]
    headers = {name: col - 1 for name, col in header_index(ws).items()}

    existing_ar_col = get_col(ws, "AR")
    ar_col_idx = ensure_col(ws, "AR")
    headers = {name: col - 1 for name, col in header_index(ws).items()}
    if existing_ar_col is not None:
        print(f"AR column already exists at col {ar_col_idx}")
    else:
        print(f"Created AR column at col {ar_col_idx}")

    pr_cache = load_pr_analysis_cache()
    print(f"Loaded pr_analysis cache for {len(pr_cache)} issues")
    live_cache = _load_live_cache()
    live_cache_initial = len(live_cache)
    print(f"Loaded live PR-state cache: {live_cache_initial} entries (TTL {VERIFY_CACHE_TTL_SEC}s)")
    tc_status = load_test_cases_status(wb)
    print(f"Loaded Test Cases status for {len(tc_status)} issues")

    counts = {b: 0 for b in CANONICAL_ORDER}
    counts["UNCLASSIFIED"] = 0
    written = 0
    rule1_log = []
    pending_ack_log = []
    for r in range(2, ws.max_row + 1):
        data = row_dict(ws, r)
        row_vals = [None] * len(headers)
        for name, idx in headers.items():
            row_vals[idx] = data.get(name)
        ar = derive_ar(row_vals, headers, pr_cache, live_cache, tc_status, rule1_log, pending_ack_log)
        write_by_name(ws, r, "AR", ar if ar else "")
        if not ar:
            counts["UNCLASSIFIED"] += 1
        else:
            for b in ar.split("; "):
                counts[b] = counts.get(b, 0) + 1
        written += 1

    wb.save(XLSX_PATH)
    _save_live_cache(live_cache)
    print(f"Wrote AR for {written} rows to {XLSX_PATH.name}")
    print(f"Live PR-state cache: {len(live_cache)} entries ({len(live_cache) - live_cache_initial} new)")
    violations = [x for x in rule1_log if x[1] == "violation"]
    oos = [x for x in rule1_log if x[1] == "out_of_scope"]
    ok = [x for x in rule1_log if x[1] == "ok"]
    print(f"=== RULE 1 audit: {len(ok)} ok, {len(oos)} out_of_scope (alt-path close), {len(violations)} VIOLATIONS ===")
    for iid, _, detail in violations:
        print(f"  VIOLATION #{iid}: {detail}")
    if oos:
        print(f"  out_of_scope issues: {[x[0] for x in oos]}")
    print(f"=== Pending-ack audit: {len(pending_ack_log)} Close-fixed rows admitted a pending ack (force-routed to Need Response) ===")
    for iid, detail in pending_ack_log:
        print(f"  PENDING-ACK #{iid}: matched {detail!r}")
    print("=== AR bucket counts (open + closed; multi-value expanded) ===")
    for b in CANONICAL_ORDER + ["UNCLASSIFIED"]:
        print(f"  {counts.get(b, 0):4d}  {b}")


if __name__ == "__main__":
    main()
