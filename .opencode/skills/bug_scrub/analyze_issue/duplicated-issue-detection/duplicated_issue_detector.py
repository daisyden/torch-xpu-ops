#!/usr/bin/env python3
"""Duplicated Issue Detection.

Detects duplicate issues from the Test Cases sheet and writes a
``duplicated_issue`` column to the Issues sheet, populated with a
comma-separated list of the other issue IDs each issue duplicates.

Duplicate relationships are derived from two sources and unioned:
  1. Exact (Test Class, Test Case) matches across different issues
     (the documented detection criteria).
  2. Any pre-existing per-test-case ``duplicated_issue`` values already
     present on the Test Cases sheet (preserves prior, possibly
     semantic, matches so re-running is non-destructive).

Columns are resolved by header name, never by hardcoded index.
"""

import argparse
import os
from collections import defaultdict

import openpyxl

ISSUES_SHEET = "Issues"
TEST_CASES_SHEET = "Test Cases"
DUP_COL = "duplicated_issue"
_BLANK = (None, "", "none", "None")


def _norm(value):
    return str(value).strip() if value is not None else ""


def _sort_key(issue_id):
    try:
        return (0, int(issue_id))
    except (TypeError, ValueError):
        return (1, str(issue_id))


def _header_index(sheet):
    return {c.value: i + 1 for i, c in enumerate(sheet[1]) if c.value is not None}


def _first_blank_column(sheet):
    col = 1
    while sheet.cell(1, col).value not in (None, ""):
        col += 1
    return col


def _compute_issue_duplicates(tc_sheet):
    hdr = _header_index(tc_sheet)
    iid_c = hdr["Issue ID"]
    class_c = hdr["Test Class"]
    case_c = hdr["Test Case"]
    dup_c = hdr.get(DUP_COL)

    # Source 1: exact (Test Class, Test Case) match across issues.
    key_to_issues = defaultdict(set)
    for r in range(2, tc_sheet.max_row + 1):
        case = _norm(tc_sheet.cell(r, case_c).value)
        if not case:
            continue
        key = (_norm(tc_sheet.cell(r, class_c).value), case)
        key_to_issues[key].add(_norm(tc_sheet.cell(r, iid_c).value))

    pairs = defaultdict(set)
    for issues in key_to_issues.values():
        if len(issues) > 1:
            for a in issues:
                pairs[a] |= issues - {a}

    # Source 2: pre-existing per-test-case duplicated_issue values.
    if dup_c is not None:
        for r in range(2, tc_sheet.max_row + 1):
            raw = tc_sheet.cell(r, dup_c).value
            if raw in _BLANK:
                continue
            owner = _norm(tc_sheet.cell(r, iid_c).value)
            for part in str(raw).replace(";", ",").split(","):
                part = part.strip()
                if part and part != owner:
                    pairs[owner].add(part)

    return {k: v for k, v in pairs.items() if v}


def add_duplicated_column(excel_file):
    wb = openpyxl.load_workbook(excel_file)
    tc_sheet = wb[TEST_CASES_SHEET]
    iss_sheet = wb[ISSUES_SHEET]

    issue_dups = _compute_issue_duplicates(tc_sheet)

    iss_hdr = _header_index(iss_sheet)
    iss_id_c = iss_hdr["Issue ID"]
    dup_c = iss_hdr.get(DUP_COL)
    if dup_c is None:
        dup_c = _first_blank_column(iss_sheet)
        iss_sheet.cell(1, dup_c).value = DUP_COL

    written = 0
    for r in range(2, iss_sheet.max_row + 1):
        issue_id = _norm(iss_sheet.cell(r, iss_id_c).value)
        if not issue_id:
            continue
        dups = issue_dups.get(issue_id)
        if not dups:
            continue
        # Incremental: only fill blanks, never overwrite existing analysis.
        if iss_sheet.cell(r, dup_c).value not in _BLANK:
            continue
        iss_sheet.cell(r, dup_c).value = ",".join(sorted(dups, key=_sort_key))
        written += 1

    wb.save(excel_file)
    print(f"Issues with duplicates detected: {len(issue_dups)}")
    print(f"duplicated_issue column: {iss_sheet.cell(1, dup_c).coordinate} ('{DUP_COL}')")
    print(f"Rows written (blank cells filled): {written}")
    return issue_dups


def main():
    parser = argparse.ArgumentParser(description="Duplicated Issue Detection")
    parser.add_argument("--excel", "-e", required=True, help="Path to torch_xpu_ops_issues.xlsx")
    args = parser.parse_args()
    if not os.path.exists(args.excel):
        print(f"ERROR: Excel file not found: {args.excel}")
        return 1
    add_duplicated_column(args.excel)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
