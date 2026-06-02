#!/usr/bin/env python3
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
import re
import os
import requests
import subprocess
import argparse
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
COMMON_DIR = os.path.abspath(os.path.join(SCRIPT_DIR, "..", "..", "_common"))
if COMMON_DIR not in sys.path:
    sys.path.insert(0, COMMON_DIR)
from header_utils import ensure_col, header_index, row_dict, write_by_name  # type: ignore[reportMissingImports] # noqa: E402
from paths import DATA_DIR as _DATA_DIR_PATH, RESULT_DIR as _RESULT_DIR_PATH  # type: ignore[reportMissingImports] # noqa: E402

RESULT_DIR = str(_RESULT_DIR_PATH)
DATA_DIR = str(_DATA_DIR_PATH)

GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")
GITHUB_HEADERS = {
    "Accept": "application/vnd.github.v3+json",
    "Authorization": f"token {GITHUB_TOKEN}"
} if GITHUB_TOKEN else {}

# PyTorchXPU ProjectV2 (Intel org, project number 61). Fetched in one batched
# GraphQL call via `gh api graphql` to bypass per-token scope filtering.
PYTORCHXPU_PROJECT_OWNER = "intel"
PYTORCHXPU_PROJECT_NUMBER = 61
VALID_PRIORITIES = {"P0", "P1", "P2", "P3"}

# Bare field names in PyTorchXPU project -> issue dict key.
PYTORCHXPU_FIELD_MAP = {
    "Priority": "project_priority",
    "Status": "project_status",
    "Estimate": "project_estimate",
    "Depending": "project_depending",
    "Short Comment": "project_short_comments",  # singular per project schema
    "Short Comments": "project_short_comments",  # tolerate plural alias
}
ALL_PROJECT_KEYS = {
    "project_priority",
    "project_status",
    "project_estimate",
    "project_depending",
    "project_short_comments",
}

# Excel cell limits
EXCEL_MAX_CELL_LEN = 32767
try:
    from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
except ImportError:
    ILLEGAL_CHARACTERS_RE = re.compile(r"[\x00-\x08\x0b-\x0c\x0e-\x1f]")


def sanitize_cell(value):
    """Strip Excel-illegal control chars and truncate to Excel cell limit."""
    if value is None:
        return ""
    text = str(value)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    if len(text) > EXCEL_MAX_CELL_LEN:
        text = text[:EXCEL_MAX_CELL_LEN]
    return text


def fetch_all_project_fields():
    """Fetch field values for every PyTorchXPU project item in one batched call.

    Uses `gh api graphql` (not requests) because gh CLI's auth handles the
    read:project scope correctly, while $GITHUB_TOKEN-based OAuth tokens get
    that scope stripped by GitHub's API gateway.

    Returns: dict[issue_number:int -> dict[cache_key:str -> value:str]]
    """
    query = '''
    query($owner: String!, $number: Int!, $cursor: String) {
      organization(login: $owner) {
        projectV2(number: $number) {
          items(first: 100, after: $cursor) {
            pageInfo { hasNextPage endCursor }
            nodes {
              content { ... on Issue { number repository { nameWithOwner } } }
              fieldValues(first: 30) {
                nodes {
                  ... on ProjectV2ItemFieldSingleSelectValue {
                    name field { ... on ProjectV2FieldCommon { name } }
                  }
                  ... on ProjectV2ItemFieldTextValue {
                    text field { ... on ProjectV2FieldCommon { name } }
                  }
                  ... on ProjectV2ItemFieldNumberValue {
                    number field { ... on ProjectV2FieldCommon { name } }
                  }
                }
              }
            }
          }
        }
      }
    }
    '''
    by_issue = {}
    cursor = None
    page = 0
    while True:
        page += 1
        args = [
            "gh", "api", "graphql",
            "-f", f"query={query}",
            "-F", f"owner={PYTORCHXPU_PROJECT_OWNER}",
            "-F", f"number={PYTORCHXPU_PROJECT_NUMBER}",
        ]
        if cursor:
            args.extend(["-F", f"cursor={cursor}"])
        try:
            proc = subprocess.run(args, capture_output=True, text=True, timeout=120)
        except (subprocess.TimeoutExpired, FileNotFoundError) as exc:
            print(f"PyTorchXPU project fetch failed (page {page}): {exc}")
            return by_issue
        if proc.returncode != 0:
            print(f"PyTorchXPU project fetch failed (page {page}): {proc.stderr.strip()}")
            return by_issue
        try:
            data = json.loads(proc.stdout)
        except json.JSONDecodeError as exc:
            print(f"PyTorchXPU project fetch returned non-JSON (page {page}): {exc}")
            return by_issue
        if data.get("errors"):
            print(f"PyTorchXPU project GraphQL errors: {data['errors']}")
            return by_issue
        proj = (((data.get("data") or {}).get("organization") or {}).get("projectV2") or {})
        items = (proj.get("items") or {}).get("nodes") or []
        for item in items:
            content = item.get("content") or {}
            issue_number = content.get("number")
            if issue_number is None:
                continue
            repo = ((content.get("repository") or {}).get("nameWithOwner") or "")
            if repo and repo != "intel/torch-xpu-ops":
                continue
            fields = {key: "" for key in ALL_PROJECT_KEYS}
            for fv in (item.get("fieldValues") or {}).get("nodes") or []:
                field = fv.get("field") or {}
                field_name = str(field.get("name") or "").strip()
                cache_key = PYTORCHXPU_FIELD_MAP.get(field_name)
                if cache_key is None:
                    continue
                raw = ""
                for key in ("name", "text", "number"):
                    val = fv.get(key)
                    if val is not None and str(val).strip():
                        raw = str(val).strip()
                        break
                if cache_key == "project_priority":
                    match = re.search(r"\bP[0-3]\b", raw.upper())
                    if match:
                        fields[cache_key] = match.group(0)
                else:
                    fields[cache_key] = raw
            by_issue[int(issue_number)] = fields
        page_info = (proj.get("items") or {}).get("pageInfo") or {}
        if not page_info.get("hasNextPage"):
            break
        cursor = page_info.get("endCursor")
        if not cursor:
            break
    print(f"PyTorchXPU project: loaded fields for {len(by_issue)} issues across {page} page(s)")
    return by_issue


def populate_project_fields(issues):
    missing = [issue for issue in issues if not ALL_PROJECT_KEYS.issubset(issue.keys())]
    if not missing:
        return
    fields_by_number = fetch_all_project_fields()
    if not fields_by_number:
        for issue in missing:
            for key in ALL_PROJECT_KEYS:
                issue.setdefault(key, "")
        return
    for issue in missing:
        number = issue.get("number")
        fields = fields_by_number.get(number) if number is not None else None
        if fields:
            for key, value in fields.items():
                issue[key] = value
        else:
            for key in ALL_PROJECT_KEYS:
                issue.setdefault(key, "")


def fetch_all_issue_types():
    query = '''
    query($endCursor: String) {
      search(query: "repo:intel/torch-xpu-ops is:issue", type: ISSUE, first: 100, after: $endCursor) {
        pageInfo { hasNextPage endCursor }
        nodes { ... on Issue { number issueType { name } } }
      }
    }
    '''
    by_number = {}
    cursor = None
    page = 0
    while True:
        page += 1
        args = ["gh", "api", "graphql", "-f", f"query={query}"]
        if cursor:
            args.extend(["-F", f"endCursor={cursor}"])
        try:
            proc = subprocess.run(args, capture_output=True, text=True, timeout=120)
        except (subprocess.TimeoutExpired, FileNotFoundError) as exc:
            print(f"issueType fetch failed (page {page}): {exc}")
            return by_number
        if proc.returncode != 0:
            print(f"issueType fetch failed (page {page}): {proc.stderr.strip()}")
            return by_number
        try:
            data = json.loads(proc.stdout)
        except json.JSONDecodeError as exc:
            print(f"issueType fetch non-JSON (page {page}): {exc}")
            return by_number
        if data.get("errors"):
            print(f"issueType GraphQL errors: {data['errors']}")
            return by_number
        search = (data.get("data") or {}).get("search") or {}
        for node in search.get("nodes") or []:
            num = node.get("number")
            if num is None:
                continue
            t = (node.get("issueType") or {}).get("name") or ""
            by_number[int(num)] = t
        page_info = search.get("pageInfo") or {}
        if not page_info.get("hasNextPage"):
            break
        cursor = page_info.get("endCursor")
        if not cursor:
            break
    print(f"issueType: loaded for {len(by_number)} issues across {page} page(s)")
    return by_number


def populate_issue_types(issues):
    type_by_num = fetch_all_issue_types()
    for issue in issues:
        num = issue.get("number")
        issue["github_type"] = type_by_num.get(int(num), "") if num is not None else ""

# Parse command line arguments
parser = argparse.ArgumentParser(description="Generate Excel report for torch-xpu-ops issues")
parser.add_argument("--issues", type=str, default="", help="Comma-separated list of issue IDs to process (default: all)")
args = parser.parse_args()

# Parse target issue IDs
TARGET_ISSUE_IDS = None
if args.issues:
    TARGET_ISSUE_IDS = set()
    for part in args.issues.split(','):
        part = part.strip()
        if part:
            try:
                TARGET_ISSUE_IDS.add(int(part))
            except ValueError:
                pass

# Load data - try to load from JSON, or fetch from GitHub if not exists
issues_json_path = os.path.join(DATA_DIR, "torch_xpu_ops_issues.json")
comments_json_path = os.path.join(DATA_DIR, "torch_xpu_ops_comments.json")

if os.path.exists(issues_json_path):
    with open(issues_json_path) as f:
        issues = json.load(f)
else:
    print("Fetching issues from GitHub...")
    issues = []
    
    # Fetch open issues
    page = 1
    print("Fetching OPEN issues...")
    while len(issues) < 500:
        url = f"https://api.github.com/repos/intel/torch-xpu-ops/issues?state=open&per_page=100&page={page}"
        response = requests.get(url, headers=GITHUB_HEADERS, timeout=30)
        if response.status_code != 200:
            break
        batch = response.json()
        if not batch:
            break
        issues.extend([i for i in batch if 'pull_request' not in i])
        print(f"Fetched {len(issues)} open issues...")
        page += 1
    
    # Fetch closed issues with wontfix/not_target labels for Not Applicable sheet population
    print("Fetching CLOSED issues with wontfix/not_target labels...")
    closed_issues = []
    # GitHub REST API /issues endpoint uses AND logic for multiple labels.
    # To get issues with EITHER wontfix OR not_target, fetch each separately.
    for label in ['wontfix', 'not_target']:
        print(f"  Fetching closed issues labeled '{label}'...")
        page = 1
        while True:
            url = f"https://api.github.com/repos/intel/torch-xpu-ops/issues?state=closed&labels={label}&per_page=100&page={page}"
            response = requests.get(url, headers=GITHUB_HEADERS, timeout=30)
            if response.status_code != 200:
                break
            batch = response.json()
            if not batch:
                break
            closed_issues.extend([i for i in batch if 'pull_request' not in i])
            print(f"    Fetched {len(closed_issues)} closed issues so far...")
            page += 1
    
    # Merge closed issues (dedup by issue number)
    existing_numbers = {i['number'] for i in issues}
    for issue in closed_issues:
        if issue['number'] not in existing_numbers:
            issues.append(issue)
            existing_numbers.add(issue['number'])
    
    print(f"Total issues (open + closed with wontfix/not_target): {len(issues)}")
    
    with open(issues_json_path, 'w') as f:
        json.dump(issues, f)

# Filter to target issues if specified
if TARGET_ISSUE_IDS:
    issues = [i for i in issues if i['number'] in TARGET_ISSUE_IDS]
    print(f"Filtered to {len(issues)} target issues: {sorted(TARGET_ISSUE_IDS)}")

populate_project_fields(issues)
populate_issue_types(issues)
before = len(issues)
issues = [i for i in issues if (i.get("github_type") or "").strip().lower() != "task"]
dropped = before - len(issues)
if dropped:
    print(f"Dropped {dropped} Task issue(s) per v4.30 policy")
if GITHUB_TOKEN and not TARGET_ISSUE_IDS:
    with open(issues_json_path, 'w') as f:
        json.dump(issues, f)

# Known test types
KNOWN_TEST_TYPES = ['op_ut', 'op_extend', 'op_extended', 'e2e', 'benchmark', 'ut', 'test_xpu']

# Model lists from benchmarks
HUGGINGFACE_MODELS = [
    'AlbertForMaskedLM', 'AlbertForQuestionAnswering', 'AllenaiLongformerBase',
    'BartForCausalLM', 'BartForConditionalGeneration', 'BertForMaskedLM',
    'BertForQuestionAnswering', 'BlenderbotForCausalLM', 'BlenderbotForConditionalGeneration',
    'BlenderbotSmallForCausalLM', 'BlenderbotSmallForConditionalGeneration', 'CamemBert',
    'DebertaV2ForMaskedLM', 'DebertaV2ForQuestionAnswering', 'DistilBertForMaskedLM',
    'DistilBertForQuestionAnswering', 'DistillGPT2', 'ElectraForCausalLM',
    'ElectraForQuestionAnswering', 'GoogleFnet', 'google/gemma-2-2b', 'google/gemma-3-4b-it',
    'GPT2ForSequenceClassification', 'GPTJForCausalLM', 'GPTJForQuestionAnswering', 'GPTNeoForCausalLM',
    'GPTNeoForSequenceClassification', 'LayoutLMForMaskedLM', 'LayoutLMForSequenceClassification',
    'M2M100ForConditionalGeneration', 'MBartForCausalLM', 'MBartForConditionalGeneration',
    'MegatronBertForCausalLM', 'MegatronBertForQuestionAnswering', 'meta-llama/Llama-3.2-1B',
    'mistralai/Mistral-7B-Instruct-v0.3', 'MobileBertForMaskedLM', 'MobileBertForQuestionAnswering',
    'MT5ForConditionalGeneration', 'openai/gpt-oss-20b', 'openai/whisper-tiny', 'OPTForCausalLM',
    'PegasusForCausalLM', 'PegasusForConditionalGeneration', 'PLBartForCausalLM',
    'PLBartForConditionalGeneration', 'Qwen/Qwen3-0.6B', 'RobertaForCausalLM', 'RobertaForQuestionAnswering',
    'T5ForConditionalGeneration', 'T5Small', 'TrOCRForCausalLM', 'XGLMForCausalLM',
    'XLNetLMHeadModel', 'YituTechConvBert'
]

TIMM_MODELS = [
    'adv_inception_v3', 'beit_base_patch16_224', 'botnet26t_256', 'cait_m36_384',
    'coat_lite_mini', 'convit_base', 'convmixer_768_32', 'convnext_base',
    'convnextv2_nano.fcmae_ft_in22k_in1k', 'crossvit_9_240', 'cspdarknet53', 'deit_base_distilled_patch16_224',
    'deit_tiny_patch16_224.fb_in1k', 'dla102', 'dm_nfnet_f0', 'dpn107', 'eca_botnext26ts_256',
    'eca_halonext26ts', 'ese_vovnet19b_dw', 'fbnetc_100', 'fbnetv3_b', 'gernet_l',
    'ghostnet_100', 'gluon_inception_v3', 'gmixer_24_224', 'gmlp_s16_224', 'hrnet_w18',
    'inception_v3', 'jx_nest_base', 'lcnet_050', 'levit_128', 'mixer_b16_224',
    'mixnet_l', 'mnasnet_100', 'mobilenetv2_100', 'mobilenetv3_large_100', 'mobilevit_s',
    'nfnet_l0', 'pit_b_224', 'pnasnet5large', 'poolformer_m36', 'regnety_002',
    'repvgg_a2', 'res2net101_26w_4s', 'res2net50_14w_8s', 'res2next50', 'resmlp_12_224',
    'resnest101e', 'rexnet_100', 'sebotnet33ts_256', 'selecsls42b', 'spnasnet_100',
    'swin_base_patch4_window7_224', 'swsl_resnext101_32x16d', 'tf_efficientnet_b0',
    'tf_mixnet_l', 'tinynet_a', 'tnt_s_patch16_224', 'twins_pcpvt_base', 'visformer_small',
    'vit_base_patch14_dinov2.lvd142m', 'vit_base_patch16_224', 'vit_base_patch16_siglip_256',
    'volo_d1_224', 'xcit_large_24_p8_224'
]

TORCHBENCH_MODELS = [
    'alexnet', 'Background_Matting', 'basic_gnn_edgecnn', 'basic_gnn_gcn', 'basic_gnn_gin',
    'basic_gnn_sage', 'BERT_pytorch', 'cm3leon_generate', 'dcgan', 'demucs', 'densenet121',
    'detectron2_fasterrcnn_r_101_c4', 'detectron2_fasterrcnn_r_101_dc5', 'detectron2_fasterrcnn_r_101_fpn',
    'detectron2_fasterrcnn_r_50_c4', 'detectron2_fasterrcnn_r_50_dc5', 'detectron2_fasterrcnn_r_50_fpn',
    'detectron2_fcos_r_50_fpn', 'detectron2_maskrcnn', 'detectron2_maskrcnn_r_101_c4', 'detectron2_maskrcnn_r_101_fpn',
    'detectron2_maskrcnn_r_50_c4', 'detectron2_maskrcnn_r_50_fpn', 'dlrm', 'doctr_det_predictor',
    'doctr_reco_predictor', 'drq', 'fastNLP_Bert', 'functorch_dp_cifar10', 'functorch_maml_omniglot',
    'hf_Albert', 'hf_Bart', 'hf_Bert', 'hf_Bert_large', 'hf_BigBird', 'hf_clip', 'hf_DistilBert',
    'hf_distil_whisper', 'hf_GPT2', 'hf_GPT2_large', 'hf_Longformer', 'hf_Reformer', 'hf_Roberta_base',
    'hf_T5', 'hf_T5_base', 'hf_T5_generate', 'hf_T5_large', 'hf_Whisper',
    'LearningToPaint', 'lennard_jones', 'llama', 'llama_v2_7b_16h', 'llava', 'maml', 'maml_omniglot',
    'microbench_unbacked_tolist_sum', 'mnasnet1_0', 'mobilenet_v2', 'mobilenet_v2_quantized_qat',
    'mobilenet_v3_large', 'moco', 'modded_nanogpt', 'moondream', 'nanogpt', 'nvidia_deeprecommender',
    'opacus_cifar10', 'phlippe_densenet', 'phlippe_resnet', 'pyhpc_equation_of_state',
    'pyhpc_isoneutral_mixing', 'pyhpc_turbulent_kinetic_energy', 'pytorch_CycleGAN_and_pix2pix',
    'pytorch_stargan', 'pytorch_unet', 'resnet152', 'resnet18', 'resnet50', 'resnet50_quantized_qat',
    'resnext50_32x4d', 'sam', 'sam_fast', 'shufflenet_v2_x1_0', 'simple_gpt', 'simple_gpt_tp_manual',
    'soft_actor_critic', 'speech_transformer', 'squeezenet1_1', 'stable_diffusion_text_encoder',
    'stable_diffusion_unet', 'Super_SloMo', 'tacotron2', 'timm_efficientdet', 'timm_efficientnet',
    'timm_nfnet', 'timm_regnet', 'timm_resnest', 'timm_vision_transformer', 'timm_vision_transformer_large',
    'timm_vovnet', 'torch_multimodal_clip', 'tts_angular', 'vgg16', 'vision_maskrcnn', 'yolov3',
    'codellama', 'DALLE2_pytorch', 'diffuser_instruct_pix2pix', 'fambench_dlrm', 'fambench_xlmr',
    'gat', 'gcn', 'hf_GPT2_generate', 'hf_mixtral', 'hf_MPT_7b_instruct', 'hf_Yi', 'lit_llama',
    'lit_llama_generate', 'lit_llama_lora', 'llama_v2_13b', 'llama_v2_70b', 'llama_v31_8b',
    'mistral_7b_instruct', 'orca_2', 'phi_1_5', 'phi_2', 'sage', 'stable_diffusion_xl', 'torchrec_dlrm'
]

def identify_benchmark(model_name):
    """Identify benchmark from model name using exact matching"""
    model_lower = model_name.lower()
    
    # Check for prefixed models (hf_*, timm_*) first - these are exact
    if model_lower.startswith('hf_'):
        return 'huggingface'
    elif model_lower.startswith('timm_'):
        return 'timm'
    
    # Check huggingface models (unprefixed variants)
    for m in HUGGINGFACE_MODELS:
        m_lower = m.lower()
        # Only match if it's a prefixed model OR exact match, not substring
        if m_lower.startswith('hf_'):
            # Skip prefixed variants here - handled above
            continue
        if m_lower == model_lower or m_lower.replace('_', '') == model_lower.replace('_', ''):
            return 'huggingface'
    
    # Check timm models
    for m in TIMM_MODELS:
        m_lower = m.lower()
        if m_lower.startswith('timm_'):
            continue
        if m_lower == model_lower or m_lower.replace('_', '') == model_lower.replace('_', ''):
            return 'timm'
    
    # Check torchbench models
    for m in TORCHBENCH_MODELS:
        m_lower = m.lower()
        if m_lower == model_lower or m_lower.replace('_', '') == model_lower.replace('_', ''):
            return 'torchbench'
    
    return 'unknown'

def extract_e2e_reproducer(body, title):
    """Extract reproducer command from issue body"""
    text = f"{title} {body}"
    
    reproducer_lines = []
    
    # Look for code blocks with commands (between ``` and ```)
    if '```' in text:
        parts = text.split('```')
        for i, part in enumerate(parts):
            # Code blocks are odd-indexed (1, 3, 5, ...)
            if i % 2 == 1:  # This is a code block content
                part_stripped = part.strip()
                if part_stripped:
                    lines = part_stripped.split('\n')
                    for line in lines:
                        line_stripped = line.strip()
                        # Look for actual commands (python, pytest, etc.)
                        if line_stripped and (line_stripped.startswith(('python', 'pytest', 'XPU_', './')) or 'python' in line_stripped.lower()):
                            if not line_stripped.startswith('#'):
                                reproducer_lines.append(line_stripped)
                    # If we found a command, use it
                    if reproducer_lines:
                        break
    
    # Also look for command patterns without code blocks
    if not reproducer_lines:
        # Look for python or pytest command patterns
        cmd_patterns = [
            r'(pytest\s+[^\n]+)',
            r'(python\s+test/[^\n]+)',
            r'(python\s+-m\s+pytest[^\n]+)',
            r'(XPU_QUANT_CONFIG=[^\n]+python[^\n]+)',
            r'(python\s+benchmarks/dynamo/[^\n]+)',
            r'(python\s+[^\n]+run_benchmark[^\n]+)',
        ]
        
        for pattern in cmd_patterns:
            matches = re.findall(pattern, text)
            for match in matches:
                reproducer_lines.append(match.strip())
    
    if not reproducer_lines:
        # Generic reproducer from title
        return title[:200]
    
    # Join and limit to 3 lines
    return '\n'.join(reproducer_lines[:3])


def parse_e2e_info(body, title):
    """Parse e2e benchmark information from issue body"""
    e2e_info = []
    
    text = f"{title} {body}"
    
    # Get reproducer
    reproducer = extract_e2e_reproducer(body, title)
    
    # Check for model names in title or body
    all_model_names = HUGGINGFACE_MODELS + TIMM_MODELS + TORCHBENCH_MODELS
    
    # Extract phase (training/inference)
    phase = 'inference'
    if 'training' in text.lower():
        phase = 'training'
    elif 'train' in text.lower():
        phase = 'training'
    
    # Extract dtype
    dtype = 'float32'
    dtype_patterns = [
        (r'bfloat16|bf16', 'bfloat16'),
        (r'float16|fp16', 'float16'),
        (r'float32|fp32', 'float32'),
        (r'int8|int\s*8', 'int8'),
    ]
    for pattern, dt in dtype_patterns:
        if re.search(pattern, text, re.IGNORECASE):
            dtype = dt
            break
    
    # Extract AMP (automatic mixed precision)
    amp = False
    if '--amp' in text.lower() or 'amp' in text.lower():
        amp = True
    
    # Extract test type
    test_type = 'accuracy'
    if 'throughputs' in text.lower() or 'performance' in text.lower() or 'latency' in text.lower():
        test_type = 'performance'
    
    # Extract backend
    backend = 'inductor'
    if '--backend=' in text:
        match = re.search(r'--backend=(\w+)', text)
        if match:
            backend = match.group(1)
    elif 'eager' in text.lower():
        backend = 'eager'
    elif 'inductor' in text.lower():
        backend = 'inductor'
    
    # Extract disable-cudagraphs
    disable_cudagraphs = 'no'
    if 'disable-cudagraphs' in text.lower() or 'disable_cudagraphs' in text.lower():
        disable_cudagraphs = 'yes'
    
    # Find model in body - need exact model name, not partial match
    found_models = set()
    for model in all_model_names:
        # Use word boundary to avoid partial matches
        if re.search(r'\b' + re.escape(model.lower()) + r'\b', text.lower()):
            benchmark = identify_benchmark(model)
            if benchmark != 'unknown' and model not in found_models:
                found_models.add(model)
                e2e_info.append({
                    'reproducer': reproducer,
                    'benchmark': benchmark,
                    'model': model,
                    'phase': phase,
                    'dtype': dtype,
                    'amp': amp,
                    'test_type': test_type,
                    'backend': backend,
                    'disable_cudagraphs': disable_cudagraphs,
                })
    
    # If no specific model found but looks like e2e issue
    if not e2e_info:
        if 'benchmark' in text.lower() or 'huggingface' in text.lower() or 'timm' in text.lower() or 'torchbench' in text.lower():
            # Try to identify benchmark from context
            if 'hf_' in text.lower() or 'huggingface' in text.lower():
                benchmark = 'huggingface'
            elif 'timm_' in text.lower() or 'timm.' in text.lower():
                benchmark = 'timm'
            elif 'torchbench' in text.lower():
                benchmark = 'torchbench'
            else:
                benchmark = 'unknown'
            
            e2e_info.append({
                'reproducer': reproducer,
                'benchmark': benchmark,
                'model': 'unknown',
                'phase': phase,
                'dtype': dtype,
                'test_type': test_type,
                'backend': backend,
                'disable_cudagraphs': disable_cudagraphs,
            })
    
    return e2e_info



def map_origin_test_file(test_file):
    if not test_file:
        return ""
    match = re.search(r'test/xpu/(.+?)(?:_xpu)?\.py$', test_file)
    if match:
        return f"test/{match.group(1)}.py"
    if 'benchmarks/' in test_file:
        return test_file
    return test_file

# Local checkouts used to verify whether a parsed test path corresponds to a
# real test file. A case is only routed to the Test Cases sheet when its file
# resolves to one of these two roots (issues whose path does not resolve fall
# through to the Others sheet).
_PYTORCH_REPO_ROOT = os.environ.get("PYTORCH_REPO_ROOT", "").strip()
if not _PYTORCH_REPO_ROOT or not os.path.isdir(_PYTORCH_REPO_ROOT):
    _PYTORCH_REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
    for _ in range(10):
        if (os.path.isdir(os.path.join(_PYTORCH_REPO_ROOT, 'test')) and
                os.path.isdir(os.path.join(_PYTORCH_REPO_ROOT, 'third_party', 'torch-xpu-ops', 'test', 'xpu'))):
            break
        _PYTORCH_REPO_ROOT = os.path.dirname(_PYTORCH_REPO_ROOT)
    else:
        _PYTORCH_REPO_ROOT = os.path.expanduser("~/upstream/pytorch")

_PYTORCH_TEST_DIR = os.path.join(_PYTORCH_REPO_ROOT, 'test')
_XPU_TEST_DIR = os.path.join(_PYTORCH_REPO_ROOT, 'third_party', 'torch-xpu-ops', 'test')


def resolve_test_file(test_path):
    """Map a dotted test path to (test_file_rel, class_suffix, origin_file_rel).

    Tries on-disk resolution first under <repo>/test/ and
    <repo>/third_party/torch-xpu-ops/test/. When no candidate file exists
    locally (e.g. shallow checkout), falls back to a best-effort string
    mapping so downstream phases still see a usable test_file. The class
    suffix is heuristically split on the longest trailing run of
    PascalCase segments (treated as a dotted class chain).

    Returns ("", "", "") only when test_path is empty.
    """
    if not test_path:
        return "", "", ""
    parts = test_path.split('.')

    candidates = []
    if 'torch-xpu-ops' in parts:
        try:
            i = parts.index('torch-xpu-ops')
            sub = parts[i + 1:]
            if sub and sub[0] == 'test':
                rel = sub[1:]
                for k in range(len(rel), 0, -1):
                    fp_abs = os.path.join(_XPU_TEST_DIR, *rel[:k]) + '.py'
                    fp_rel = 'torch-xpu-ops/test/' + '/'.join(rel[:k]) + '.py'
                    suffix = '.'.join(rel[k:])
                    candidates.append((fp_abs, fp_rel, suffix, 'xpu'))
        except ValueError:
            pass

    if parts and parts[0] == 'test':
        rel = parts[1:]
    else:
        rel = parts
    for k in range(len(rel), 0, -1):
        fp_abs = os.path.join(_PYTORCH_TEST_DIR, *rel[:k]) + '.py'
        fp_rel = 'test/' + '/'.join(rel[:k]) + '.py'
        suffix = '.'.join(rel[k:])
        candidates.append((fp_abs, fp_rel, suffix, 'pytorch'))

    for fp_abs, fp_rel, suffix, kind in candidates:
        if os.path.isfile(fp_abs):
            origin = map_origin_test_file(fp_rel) if kind == 'xpu' else fp_rel
            return fp_rel, suffix, origin

    # Best-effort fallback: file not present in this checkout. Reconstruct
    # the most plausible relative path + class suffix so Phase 2.4 (case
    # existence) and Phase 2.5 (local verification) have something to
    # investigate rather than an empty cell.
    def _split_class_suffix(rel_parts):
        # Pop trailing PascalCase tokens as class chain (e.g. ['ReproTests'])
        cls = []
        while rel_parts and rel_parts[-1] and rel_parts[-1][:1].isupper():
            cls.insert(0, rel_parts.pop())
        return rel_parts, '.'.join(cls)

    if 'torch-xpu-ops' in parts:
        try:
            i = parts.index('torch-xpu-ops')
            sub = parts[i + 1:]
            if sub and sub[0] == 'test':
                rel = list(sub[1:])
                rel, class_suffix = _split_class_suffix(rel)
                if rel:
                    fp_rel = 'torch-xpu-ops/test/' + '/'.join(rel) + '.py'
                    return fp_rel, class_suffix, map_origin_test_file(fp_rel)
        except ValueError:
            pass

    rel = list(parts[1:] if parts and parts[0] == 'test' else parts)
    rel, class_suffix = _split_class_suffix(rel)
    if rel:
        fp_rel = 'test/' + '/'.join(rel) + '.py'
        return fp_rel, class_suffix, fp_rel
    return "", "", ""


_PYTEST_FILE_RE = re.compile(r'pytest[^\n`]*?\b(test[\w/]*\.py)((?:::[\w.]+)*)')
_STACK_FILE_RE = re.compile(r'File\s+"[^"]*?(test[/\\][\w/\\.]+\.py)"')
_CLASS_METHOD_RE = re.compile(r'\b([A-Z][A-Za-z0-9_]*)\.(test_\w+)\b')
_TEST_METHOD_RE = re.compile(r'\b(test_\w+)\b')


def _disk_match_rel_file(rel):
    """Check if a relative test file path (e.g. 'test/foo.py' or 'test_vmap_xpu.py')
    exists under either pytorch test/ or xpu-ops test/.

    Returns canonical relative path or "".
    """
    if not rel:
        return ""
    rel = rel.replace('\\', '/').lstrip('./')
    if rel.startswith('test/'):
        if os.path.isfile(os.path.join(_PYTORCH_REPO_ROOT, rel)):
            return rel
        sub = rel[len('test/'):]
        for root_rel, root_abs in (('torch-xpu-ops/test/', _XPU_TEST_DIR),):
            if os.path.isfile(os.path.join(root_abs, sub)):
                return root_rel + sub
    basename = os.path.basename(rel)
    for root_abs, root_rel in ((_PYTORCH_TEST_DIR, 'test/'), (_XPU_TEST_DIR, 'torch-xpu-ops/test/')):
        for cur, _dirs, files in os.walk(root_abs):
            if basename in files:
                full = os.path.join(cur, basename)
                return root_rel + os.path.relpath(full, root_abs).replace('\\', '/')
    return ""


def best_effort_test_info(body, title):
    """Best-effort mining of test_file/class/case from issue body+title for
    issues without a parseable 'Cases:' block (used as label-fallback).

    Returns (test_file_rel, test_class, test_case) where any field may be "".
    """
    body = body or ''
    title = title or ''
    test_file = ''
    test_class = ''
    test_case = ''

    for m in _PYTEST_FILE_RE.finditer(body):
        rel = _disk_match_rel_file(m.group(1))
        if rel:
            test_file = rel
            tail = m.group(2)
            if tail:
                segs = [s for s in tail.split('::') if s]
                if len(segs) >= 2:
                    test_class, test_case = segs[0], segs[1]
                elif len(segs) == 1:
                    test_case = segs[0]
            break

    if not test_file:
        for m in _STACK_FILE_RE.finditer(body):
            rel = _disk_match_rel_file(m.group(1))
            if rel:
                test_file = rel
                break

    if not test_case:
        m = _CLASS_METHOD_RE.search(body)
        if m:
            test_class = test_class or m.group(1)
            test_case = m.group(2)

    if not test_case:
        m = _TEST_METHOD_RE.search(title)
        if m:
            test_case = m.group(1)

    return test_file, test_class, test_case


import hashlib as _hashlib

_LLM_CACHE_PATH = os.path.join(DATA_DIR, 'llm_extracted.json')
_LLM_CACHE = None


def _load_llm_cache():
    global _LLM_CACHE
    if _LLM_CACHE is not None:
        return _LLM_CACHE
    try:
        with open(_LLM_CACHE_PATH) as f:
            _LLM_CACHE = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        _LLM_CACHE = {}
    return _LLM_CACHE


def get_llm_extraction(issue_id, body):
    cache = _load_llm_cache()
    entry = cache.get(str(issue_id))
    if not entry:
        return None
    body_hash = _hashlib.sha256((body or '').encode('utf-8')).hexdigest()[:16]
    if entry.get('body_hash') and entry['body_hash'] != body_hash:
        return None
    return entry


def llm_test_cases_for_issue(issue_id, body):
    """Return disk-verified test_cases list for an issue from the LLM cache.

    Each item: {'test_type', 'test_file', 'origin_test_file', 'test_class', 'test_case'}.
    Paths that do not resolve under <repo>/test/ or torch-xpu-ops/test/ are dropped.
    """
    entry = get_llm_extraction(issue_id, body)
    if not entry:
        return []
    out = []
    for tc in entry.get('test_cases', []) or []:
        raw = (tc.get('test_file') or '').strip().replace('\\', '/')
        cls = (tc.get('test_class') or '').strip()
        case = (tc.get('test_method') or '').strip()
        if not raw or not case:
            continue
        rel = _disk_match_rel_file(raw)
        if not rel:
            continue
        out.append({
            'test_type': 'ut',
            'test_file': rel,
            'origin_test_file': map_origin_test_file(rel) if rel.startswith('torch-xpu-ops/') else rel,
            'test_class': cls,
            'test_case': case,
        })
    return out


def parse_test_cases_from_body(body):
    cases = []

    if 'Cases:' in body:
        cases_section = body.split('Cases:')[1]

        end_markers = ['\n###', '\nVersions', '\n```']
        min_end = len(cases_section)
        for marker in end_markers:
            idx = cases_section.find(marker)
            if idx > 0 and idx < min_end:
                min_end = idx
        cases_section = cases_section[:min_end]

        lines = cases_section.split('\n')

        for line in lines:
            line = line.strip()

            if not line:
                continue

            if line.startswith('###') or line.startswith('...'):
                continue

            if line.startswith('~~') and line.endswith('~~'):
                continue

            parts = line.split(',')
            if len(parts) < 3:
                continue

            test_type = parts[0].strip()
            if test_type not in KNOWN_TEST_TYPES:
                continue

            field1 = parts[1].strip()
            field2 = parts[2].strip()

            # Two formats observed in the wild:
            #   A) op_ut,<dotted.module[.Class]>,<test_case>
            #   B) op_ut,,<dotted.module>            (module-level import error)
            # In (B) field1 is empty and field2 is the module path with no case.
            if field1:
                test_path = field1
                test_case = field2
                module_level = False
            else:
                test_path = field2
                test_case = ''
                module_level = True

            if not test_path:
                continue
            if not module_level:
                if not test_case or len(test_case) < 3:
                    continue
                if ' ' in test_case:
                    continue

            test_file, class_suffix, origin_file = resolve_test_file(test_path)
            test_class = class_suffix

            if not module_level and not test_class and '.' in test_case:
                head, _, tail = test_case.rpartition('.')
                if head and tail:
                    test_class = head
                    test_case = tail

            cases.append({
                'test_type': test_type,
                'test_file': test_file,
                'origin_test_file': origin_file,
                'test_class': test_class,
                'test_case': test_case,
                'module_level': module_level,
            })

    # Extract from pytest code blocks (format: pytest -v test/test_ops.py -k test_name)
    if '```' in body:
        code_blocks = body.split('```')
        for block in code_blocks:
            # Look for pytest patterns with test path and test method
            # Handles formats: test/test_ops.py or test/distributed/test_c10d_xccl.py::ClassName::method
            pytest_pattern = r'pytest\s+-v\s+(test[/a-zA-Z0-9_/.]+\.py(?:::[a-zA-Z0-9_]+)*)'
            matches = re.findall(pytest_pattern, block)
            for match in matches:
                test_path = match.strip()
                if '::' in test_path:
                    parts = test_path.split('::')
                    file_path = parts[0]
                    test_class = parts[1] if len(parts) > 1 else ""
                    # Only emit test_case when an explicit ::method segment is present.
                    # With just file::Class, the -k handler below produces the real
                    # method row; emitting test_method=class here yields a degenerate
                    # row where test_class == test_case.
                    test_method = parts[2] if len(parts) > 2 else ""
                    if test_method:
                        cases.append({
                            'test_type': 'ut',
                            'test_file': file_path,
                            'origin_test_file': file_path,
                            'test_class': test_class,
                            'test_case': test_method
                        })
                else:
                    # No class/method, just file
                    cases.append({
                        'test_type': 'ut',
                        'test_file': test_path,
                        'origin_test_file': test_path,
                        'test_class': '',
                        'test_case': ''
                    })

            # Also look for test_xpu,...,... format in code blocks
            test_xpu_pattern = r'(test_xpu),([a-zA-Z0-9_\.]+),([a-zA-Z0-9_]+)'
            matches = re.findall(test_xpu_pattern, block)
            for match in matches:
                test_type, test_path, test_method = match[0], match[1], match[2]
                test_class = ""
                if '.test_' in test_path:
                    # e.g., test.test_xpu.TestXpuAutocast -> TestXpuAutocast
                    class_parts = test_path.split('.test_')
                    if len(class_parts) > 1:
                        class_name = class_parts[1]
                        if '.' in class_name:
                            test_class = class_name.rsplit('.', 1)[1] if '.' in class_name else class_name
                        else:
                            test_class = class_name
                cases.append({
                    'test_type': test_type,
                    'test_file': test_path.replace('.', '/') + '.py',
                    'origin_test_file': test_path.replace('.', '/') + '.py',
                    'test_class': test_class,
                    'test_case': test_method
                })

            # Also handle pytest commands with -k pattern (extract test method from -k value)
            # Look for: pytest ... -k test_python_ref__refs_logspace_tensor_overload_xpu_float64
            k_pattern_matches = re.findall(r'-k\s+([a-zA-Z0-9_]+)', block)
            for test_name in k_pattern_matches:
                # Try to find associated test file in the same block
                pytest_v_match = re.search(r'pytest\s+-v\s+(test[/a-zA-Z0-9_]+\.py)', block)
                if pytest_v_match:
                    file_path = pytest_v_match.group(1)
                    cases.append({
                        'test_type': 'ut',
                        'test_file': file_path,
                        'origin_test_file': file_path,
                        'test_class': '',
                        'test_case': test_name
                    })

    # Extract from pytest commands outside code blocks
    # Look for patterns like: pytest -v test/test_ops.py -k test_name
    re_pattern = r'pytest\s+-v\s+(test[/a-zA-Z0-9_]+\.py)\s*-k\s+([a-zA-Z0-9_]+)'
    matches = re.findall(re_pattern, body)
    for file_path, test_name in matches:
        cases.append({
            'test_type': 'ut',
            'test_file': file_path,
            'origin_test_file': file_path,
            'test_class': '',
            'test_case': test_name
        })

    if 'benchmarks/dynamo/' in body:
        matches = re.findall(r'(python\s+benchmarks/dynamo/[^\s]+)', body)
        for match in matches:
            test_file = match.replace('python ', '').strip()
            cases.append({
                'test_type': 'e2e',
                'test_file': test_file,
                'origin_test_file': test_file,
                'test_class': '',
                'test_case': match.strip()
            })

    if 'pytest' in body:
        k_match = re.search(r'pytest[^-]*(-k\s+[^\s]+)?', body)
        if k_match and k_match.group(1):
            cases.append({
                'test_type': 'ut',
                'test_file': '',
                'origin_test_file': '',
                'test_class': '',
                'test_case': k_match.group(1).strip()
            })

    return cases

def generate_summary(body, title):
    # Summary based on issue title
    return title[:150]

def classify_issue_type(body, title, labels):
    text = f"{title} {body}".lower()
    
    for label in labels:
        ln = label.get('name', '').lower()
        if 'task' == ln or 'internal task' in ln:
            return 'internal task'
    
    performance_keywords = [
        'performance regression', 'performance dropped', 'performance issue',
        'latency', 'throughput', 'slow performance', 'performance slow',
        'execution time', 'runtime performance', 'performance fail'
    ]
    
    has_performance_keyword = any(k in text for k in performance_keywords)
    
    bug_keywords = [
        'assertionerror', 'runtimeerror', 'valueerror', 'typeerror', 'indexerror',
        'keyerror', 'importerror', 'notimplementederror', 'attributeerror',
        'inductorerror', 'crash', 'fail', 'bug', 'error', 'not implemented',
        'not supported', 'missing', 'incorrect', 'wrong', 'unexpected'
    ]
    
    has_bug_keyword = any(k in text for k in bug_keywords)
    
    feature_keywords = ['feature request', 'support for', 'implement', 'add support', 'need feature']
    has_feature_keyword = any(k in text for k in feature_keywords)
    
    if has_feature_keyword:
        return 'feature request'
    
    if has_performance_keyword:
        return 'performance issue'
    
    if has_bug_keyword:
        return 'functionality bug'
    
    return 'unknown'

def is_e2e_issue(body, title, labels):
    """Check if issue is related to E2E benchmark"""
    text = f"{title} {body}".lower()
    
    # Check labels first - only exact 'e2e' label
    for label in labels:
        ln = label.get('name', '').lower()
        if ln == 'e2e':
            return True
    
    # Check for specific E2E benchmark paths (not just the word 'benchmark')
    e2e_patterns = [
        r'benchmarks/dynamo/',           # torch-xpu-ops benchmark scripts
        r'benchmarks/timm/',             # timm benchmark
        r'benchmarks/huggingface/',     # huggingface benchmark
        r'benchmarks/torchbench/',      # torchbench benchmark
        r'run_benchmark\.py',            # torchbenchmark runner
    ]
    
    for pattern in e2e_patterns:
        if re.search(pattern, text):
            return True
    
    # Check for model names from benchmark model lists with explicit benchmark framework mention
    # Only for specific benchmark prefixes
    benchmark_model_prefixes = ['hf_', 'timm_']  # e.g., hf_Albert, timm_resnet50
    
    has_model = False
    has_benchmark_context = False
    
    for prefix in benchmark_model_prefixes:
        if prefix in text:
            has_model = True
            break
    
    # Must have explicit benchmark framework mention (as test framework)
    if has_model:
        benchmark_paths = ['benchmarks/dynamo', 'run_benchmark', 'torchbenchmark', 'benchmark.py']
        for kw in benchmark_paths:
            if kw in text:
                has_benchmark_context = True
                break
    
    if has_model and has_benchmark_context:
        return True
    
    return False


def classify_test_module(body, title, labels):
    text = f"{title} {body}".lower()
    
    # Check if it's an E2E issue first
    if is_e2e_issue(body, title, labels):
        return 'e2e'
    
    pytest_patterns = [
        r'pytest\s+.*test[/._]',
        r'python\s+.*test[/._]',
        r'test/test_',
        r'test/xpu/test_',
    ]
    
    has_test_pattern = False
    for pattern in pytest_patterns:
        if re.search(pattern, text):
            has_test_pattern = True
            break
    
    build_patterns = [
        r'\[win\]\[build\]',
        r'build from source',
        r'compile from source', 
        r'source build',
        r'build script',
        r'BUILD_SEPARATE',
        r'BUILD_SHARED',
        r'cmake build',
        r'cmake error',
        r'cmake fail',
        r'setup\.py install',
        r'pip install -e \.',
        r'python setup\.py develop',
    ]
    
    has_build = any(re.search(p, text, re.IGNORECASE) for p in build_patterns)
    
    infra_patterns = [
        r'workflow\s+(error|fail|issue|problem)',
        r'github\s+action\s+(error|fail|issue)',
        r'azure\s+pipeline\s+(error|fail)',
        r'ci\s+(runner|config|setup)\s+(error|fail)',
        r'runner\s+(error|fail|timeout)\s+in\s+ci',
        r'checkout\s+(error|fail)\s+in\s+(workflow|ci)',
        r'githubaction',
    ]
    
    has_infra = any(re.search(p, text) for p in infra_patterns)
    
    for label in labels:
        ln = label.get('name', '').lower()
        if 'infrastructure' in ln and ('ci' in ln or 'workflow' in ln or 'action' in ln):
            has_infra = True
            break
    
    if has_build:
        return 'build'
    
    if has_infra:
        return 'infrastructure'
    
    if has_test_pattern:
        if 'benchmarks/dynamo/' in text or 'benchmark' in text:
            return 'e2e'
        return 'ut'
    
    return 'ut'

def classify_module(body, title, labels):
    text = f"{title} {body}".lower()
    labels_str = ', '.join([l.get('name', '') for l in labels]).lower()
    
    # Check labels first
    for label in labels:
        ln = label.get('name', '').lower()
        if 'module: distributed' in ln:
            return 'distributed'
        if 'module: inductor' in ln:
            return 'inductor'
        if 'module: ao' in ln:
            return 'AO'
        if 'module: ut' in ln:
            return 'aten_ops'
        if 'module: quant' in ln:
            return 'low_precision'
        if 'module: profiler' in ln:
            return 'profiling'
        if 'module: dynamo' in ln:
            return 'dynamo'
        if 'module: op impl' in ln:
            return 'aten_ops'
    
    # Special case - "Torch not compiled with CUDA enabled" means test configuration issue, not inductor
    if 'torch not compiled with cuda enabled' in text:
        return 'unknown'
    
    # Random failures are not module-specific
    if 'random failure' in text or 'random failures' in text:
        return 'unknown'
    
    # Torch operations (from PyTorch docs)
    torch_ops = [
        'add', 'sub', 'mul', 'div', 'matmul', 'mm', 'dot', 'vdot', 'bmm',
        'addmm', 'addmv', 'addbmm', 'smm', 'spmm', 'mm', 'mv', 'vecdot',
        'conv', 'conv1d', 'conv2d', 'conv3d', 'conv_transpose',
        'batch_norm', 'layer_norm', 'group_norm', 'instance_norm',
        'dropout', 'embedding', 'linear', 'lstm', 'gru', 'rnn',
        'softmax', 'log_softmax', 'sigmoid', 'tanh', 'relu', 'leaky_relu',
        'pool', 'avg_pool', 'max_pool', 'adaptive_pool',
        'fft', 'ifft', 'fft2', 'ifft2',
        'chunk', 'split', 'view', 'reshape', 'transpose', 'permute',
        'cat', 'stack', 'gather', 'scatter', 'index', 'where',
        'sum', 'mean', 'std', 'var', 'min', 'max', 'argmin', 'argmax',
        'norm', 'linalg.norm', 'linalg.matrix_norm', 'linalg.vector_norm',
        'eig', 'svd', 'qr', 'cholesky', 'solve', 'inverse',
        'det', 'logdet', 'slogdet', 'trace',
        'clone', 'copy_', 'to', 'cuda', 'cpu', 'xpu', 'device',
        'zeros', 'ones', 'empty', 'full', 'arange', 'linspace', 'logspace',
        'tensor', 'scalar_tensor', 'tensor.tensor',
        'getitem', 'setitem', 'call', 'forward', 'backward',
        'relu', 'gelu', 'silu', 'mish', 'softplus', 'elu', 'selu', 'celu',
        'flash_attention', 'scaled_dot_product_attention', 'sdpa',
        'interpolate', 'grid_sample', 'affine_grid',
        'grid_sampler', 'grid_sampler_2d',
        'bernoulli', 'normal', 'uniform', 'randn', 'rand', 'randint',
        'multinomial', ' poisson', 'exponential', 'geometric',
        'lerp', 'lerp_', 'fmod', 'remainder', 'nextafter',
        'linspace', 'logspace', 'geomspace',
        'complex', 'real', 'imag', 'angle',
        'conj', 'view_as_real', 'view_as_complex',
    ]
    
    module_keywords = [
        ('distributed', ['distributed', 'device_mesh', 'ProcessGroup', 'FSDP', 'DDP', 'c10d', 'tensor parallel']),
        ('inductor', ['inductor', 'inductor error', 'compile error', 'lower', 'kernel code']),
        ('dynamo', ['dynamo', 'torch.compile', '_dynamo', 'dynamo']),
        ('autograd', ['autograd', 'backward', 'grad', 'gradient']),
        ('aten_ops', ['aten::', 'torch.ops.aten', 'test_ops']),
        ('low_precision', ['quantization', 'int8', 'fp8', 'int4', 'amp', 'bf16', 'fp16', 'float8']),
        ('optimizer', ['optimizer', 'lr_scheduler', 'adam', 'sgd']),
        ('profiling', ['profiling', 'profile', 'benchmark']),
        ('fx', ['torch.fx', 'fx.', 'symbolic']),
        ('export', ['torch.export', 'exported']),
    ]
    
    # Check torch ops first
    for op in torch_ops:
        if re.search(rf'\b{re.escape(op)}\b', text):
            return 'aten_ops'
    
    for m, kw in module_keywords:
        if any(k in text for k in kw):
            return m
    
    return 'unknown'

def get_dependency_from_body(body, labels=None):
    if labels is None:
        labels = []
    
    labels_str = ', '.join([l.get('name', '') for l in labels]).lower()
    
    # Check labels first for 'dependency component:'
    if 'dependency component: onednn' in labels_str or 'dependency component: mkl-dnn' in labels_str or 'dependency component: dnnl' in labels_str:
        return 'oneDNN'
    if 'dependency component: onemkl' in labels_str or 'dependency component: mkl' in labels_str:
        return 'oneMKL'
    if 'dependency component: triton' in labels_str:
        return 'Triton'
    if 'dependency component: torchao' in labels_str:
        return 'AO'
    if 'dependency component: transformers' in labels_str or 'dependency component: huggingface' in labels_str:
        return 'transformers'
    if 'dependency component: oneapi' in labels_str or 'dependency component: sycl' in labels_str:
        return 'oneAPI'
    if 'dependency component: driver' in labels_str:
        return 'driver'
    if 'dependency component: oneccl' in labels_str or 'dependency component: ccl' in labels_str or 'dependency component: xccl' in labels_str:
        return 'oneCCL'
    
    # Filter out version/environment sections
    if not body:
        return ''
    
    text = body.lower()
    
    # Remove version/environment sections
    version_headers = [
        r'###\s*version',
        r'###\s*versions',
        r'###\s*environment',
        r'###\s*reproduction',
        r'###\s*steps?\s+to\s+reproduce',
        r'###\s*additional\s*context',
    ]
    
    for header in version_headers:
        match = re.search(header, text, re.IGNORECASE)
        if match:
            text = text[:match.start()]
            break
    
    # Check for actual dependency in body (require context like "caused by", "issue", "depend on")
    dep_keywords = [
        ('transformers', [
            'caused by transformers', 'transformers issue', 'transformers bug',
            'depends on transformers', 'need transformers fix', 'waiting for transformers',
            'huggingface issue', 'huggingface bug', 'depends on huggingface'
        ]),
        ('AO', [
            'caused by torchao', 'torchao issue', 'torchao bug',
            'depends on torchao', 'need torchao fix', 'waiting for torchao'
        ]),
        ('oneDNN', [
            'caused by onednn', 'onednn issue', 'onednn bug', 'oneDNN issue',
            'depends on onednn', 'need onednn fix', 'waiting for onednn',
            'mkl-dnn issue', 'dnnl issue'
        ]),
        ('oneCCL', [
            'caused by oneccl', 'oneccl issue', 'oneccl bug',
            'depends on oneccl', 'need oneccl fix', 'waiting for oneccl',
            'xccl issue', 'ccl issue', 'depends on ccl'
        ]),
        ('oneMKL', [
            'caused by onemkl', 'onemkl issue', 'onemkl bug',
            'depends on onemkl', 'need onemkl fix', 'waiting for onemkl',
            'caused by mkl', 'mkl issue'
        ]),
        ('driver', [
            'caused by driver', 'driver issue', 'driver bug',
            'depends on driver', 'need driver fix', 'waiting for driver'
        ]),
        ('Triton', [
            'caused by triton', 'triton issue', 'triton bug',
            'depends on triton', 'need triton fix', 'waiting for triton',
            'triton-xpu issue', 'tl\\. issue'
        ]),
        ('oneAPI', [
            'caused by oneapi', 'oneapi issue', 'oneapi bug', 'sycl issue',
            'depends on oneapi', 'need oneapi fix', 'waiting for oneapi',
            'icpx issue', 'dpcpp issue', 'sycl compiler issue'
        ]),
    ]
    
    for d, kw in dep_keywords:
        if any(k in text for k in kw):
            return d
    
    return ''


# Patterns for extracting error/traceback when an issue has no parsed test cases.
ERROR_LINE_RE = re.compile(
    r'^\s*(?:[A-Za-z_][\w\.]*(?:Error|Exception|Warning)|RuntimeError|AssertionError|'
    r'ValueError|TypeError|IndexError|KeyError|ImportError|NotImplementedError|'
    r'AttributeError|InductorError):\s*.+',
    re.MULTILINE,
)
TRACEBACK_RE = re.compile(
    r'Traceback \(most recent call last\):.*?(?=\n\s*\n|\n###|\n```|\Z)',
    re.DOTALL,
)


def extract_error_message(body):
    if not body:
        return ""
    match = ERROR_LINE_RE.search(body)
    if match:
        return match.group(0).strip()
    return ""


def extract_traceback(body):
    if not body:
        return ""
    match = TRACEBACK_RE.search(body)
    if match:
        return match.group(0).strip()
    return ""


# Create Excel
wb = openpyxl.Workbook()

# Sheet 1: Issues
ws_issues = wb.active
ws_issues.title = "Issues"

# Core columns for basic issue info
# Note: PR columns (PR, PR Owner, PR Status, PR Description) populated by ../pr-extraction/
# Note: owner_transfer, action_TBD, Category, Root Cause columns populated by update_test_results/
# Priority is initialized from GitHub Projects when PyTorchXPU Priority is set;
# otherwise Phase 3 fills it from triage analysis.
headers = ["Issue ID", "Title", "Status", "Assignee", "Reporter", "Labels",
            "Created Time", "Updated Time", "Milestone", "Summary", "Type", "GitHub Type",
           "Module", "Test Module", "Dependency", "Priority",
           "PyTorchXPU Status", "PyTorchXPU Estimate",
           "PyTorchXPU Depending", "PyTorchXPU Short Comments"]

for col, header in enumerate(headers, 1):
    cell = ws_issues.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Sheet 2: Test Cases (ut)
ws_cases = wb.create_sheet("Test Cases")
case_headers = ["Issue ID", "Test Reproducer", "Test Type", "Test File",
                "Origin Test File", "Test Class", "Test Case"]

for col, header in enumerate(case_headers, 1):
    cell = ws_cases.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Sheet 3: E2E Test Cases
ws_e2e = wb.create_sheet("E2E Test Cases")

# Core columns for E2E test case basic info
# Note: Error Message, Traceback filled by bug_scrub/analyze_ci_result/e2e_test_cases
e2e_headers = ["Issue ID", "Test Reproducer", "Benchmark", "Model", "Phase", "Dtype", "AMP",
               "Backend", "Test Type", "Cudagraph", "Error Message", "Traceback"]

for col, header in enumerate(e2e_headers, 1):
    cell = ws_e2e.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

# Sheet 4: Others (issues without UT or E2E test cases)
ws_others = wb.create_sheet("Others")
others_headers = ["ID", "Title", "Labels", "reproduce step", "Error Message", "Traceback"]

for col, header in enumerate(others_headers, 1):
    cell = ws_others.cell(row=1, column=col, value=header)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

issue_row = 2
case_row = 2
e2e_row = 2
others_row = 2

issues_with_ut = set()
issues_with_e2e = set()

# Track test case duplicates: (test_file, test_class, test_case) per issue
# Also skip cases where test_case or test_class cannot be extracted
seen_test_cases = set()
valid_test_types = set(KNOWN_TEST_TYPES + ['e2e'])

for issue in issues:
    num = issue['number']
    title = issue['title']
    body = issue.get('body', '') or ''
    
    status = issue['state']
    assignee = ", ".join([a['login'] for a in issue.get('assignees', [])])
    reporter = issue['user']['login']
    labels = issue.get('labels', [])
    label_str = ", ".join([l['name'] for l in labels])
    created = issue['created_at']
    updated = issue['updated_at']
    milestone = issue.get('milestone', {})
    milestone_name = (milestone.get('title') or '') if milestone else ''
    
    issue_type = classify_issue_type(body, title, labels)
    module = classify_module(body, title, labels)
    test_module = classify_test_module(body, title, labels)
    dependency = get_dependency_from_body(body, labels)
    project_priority = issue.get('project_priority', '') or ''
    
    # Error Message and Traceback are populated later by bug_scrub/analyze_ci_result/.
    summary = generate_summary(body, title)
    
    write_by_name(ws_issues, issue_row, "Issue ID", num)
    write_by_name(ws_issues, issue_row, "Title", title)
    write_by_name(ws_issues, issue_row, "Status", status)
    write_by_name(ws_issues, issue_row, "Assignee", assignee)
    write_by_name(ws_issues, issue_row, "Reporter", reporter)
    write_by_name(ws_issues, issue_row, "Labels", label_str)
    write_by_name(ws_issues, issue_row, "Created Time", created)
    write_by_name(ws_issues, issue_row, "Updated Time", updated)
    write_by_name(ws_issues, issue_row, "Milestone", milestone_name)
    write_by_name(ws_issues, issue_row, "Summary", summary)
    write_by_name(ws_issues, issue_row, "Type", issue_type)
    write_by_name(ws_issues, issue_row, "GitHub Type", issue.get("github_type", ""))
    write_by_name(ws_issues, issue_row, "Module", module)
    write_by_name(ws_issues, issue_row, "Test Module", test_module)
    write_by_name(ws_issues, issue_row, "Dependency", dependency)
    write_by_name(ws_issues, issue_row, "Priority", project_priority)
    write_by_name(ws_issues, issue_row, "PyTorchXPU Status", sanitize_cell(issue.get('project_status', '')))
    write_by_name(ws_issues, issue_row, "PyTorchXPU Estimate", sanitize_cell(issue.get('project_estimate', '')))
    write_by_name(ws_issues, issue_row, "PyTorchXPU Depending", sanitize_cell(issue.get('project_depending', '')))
    write_by_name(ws_issues, issue_row, "PyTorchXPU Short Comments", sanitize_cell(issue.get('project_short_comments', '')))
    
    # Parse test cases and e2e info
    test_cases = parse_test_cases_from_body(body)

    llm_entry = get_llm_extraction(num, body)
    # LLM is a FALLBACK: only consult it when the script extractor found nothing.
    if llm_entry and not test_cases:
        for llm_tc in llm_test_cases_for_issue(num, body):
            test_cases.append(llm_tc)
    
    # Only parse e2e info if it's actually an e2e issue
    e2e_info = []
    if test_module == 'e2e':
        e2e_info = parse_e2e_info(body, title)
    
    # Add to test cases sheet (non-e2e)
    if test_cases:
        # Deduplicate within this issue, filter invalid cases
        for tc in test_cases:
            # Skip e2e cases - they go to e2e sheet
            if tc.get('test_type') == 'e2e':
                continue

            test_class = tc.get('test_class', '')
            test_case = tc.get('test_case', '')
            test_file = tc.get('test_file', '')
            module_level = bool(tc.get('module_level'))

            if not test_file:
                continue
            if not module_level:
                if not test_case:
                    continue
                if len(test_case) < 3:
                    continue
                if any(c in test_case for c in ['~', '`', '@', '#', '$', '%', '^', '&', '*', '(', ')']):
                    continue
            if test_class and any(c in test_class for c in ['~', '`', '@', '#', '$', '%', '^', '&', '*', '(', ')']):
                continue

            dup_key = (test_file, test_class, test_case)
            if dup_key in seen_test_cases:
                continue
            seen_test_cases.add(dup_key)

            write_by_name(ws_cases, case_row, "Issue ID", num)
            write_by_name(ws_cases, case_row, "Test Reproducer", title[:150])
            write_by_name(ws_cases, case_row, "Test Type", tc.get('test_type', ''))
            write_by_name(ws_cases, case_row, "Test File", test_file)
            write_by_name(ws_cases, case_row, "Origin Test File", tc.get('origin_test_file', ''))
            write_by_name(ws_cases, case_row, "Test Class", test_class)
            write_by_name(ws_cases, case_row, "Test Case", test_case)
            # Columns 8-9: Error Message and Traceback are left blank for bug_scrub/analyze_ci_result/.
            case_row += 1
            issues_with_ut.add(num)

    if num not in issues_with_ut and test_module != 'e2e':
        label_lc = label_str.lower()
        llm_kind = llm_entry.get('kind') if llm_entry else None
        if 'module: ut' in label_lc or 'skipped' in label_lc or llm_kind == 'unittest':
            be_file, be_class, be_case = best_effort_test_info(body, title)
            if llm_entry and (not be_file or not be_case):
                for llm_tc in llm_test_cases_for_issue(num, body):
                    be_file = be_file or llm_tc['test_file']
                    be_class = be_class or llm_tc['test_class']
                    be_case = be_case or llm_tc['test_case']
                    if be_file and be_case:
                        break
            write_by_name(ws_cases, case_row, "Issue ID", num)
            write_by_name(ws_cases, case_row, "Test Reproducer", title[:150])
            write_by_name(ws_cases, case_row, "Test Type", 'ut')
            write_by_name(ws_cases, case_row, "Test File", be_file)
            write_by_name(ws_cases, case_row, "Origin Test File", map_origin_test_file(be_file) if be_file else '')
            write_by_name(ws_cases, case_row, "Test Class", be_class)
            write_by_name(ws_cases, case_row, "Test Case", be_case)
            case_row += 1
            issues_with_ut.add(num)

    # Add to e2e sheet
    if e2e_info and num not in issues_with_ut:
        for info in e2e_info:
            reproducer = info.get('reproducer', title[:150])
            write_by_name(ws_e2e, e2e_row, "Issue ID", num)
            write_by_name(ws_e2e, e2e_row, "Test Reproducer", reproducer[:200] if reproducer else title[:150])
            write_by_name(ws_e2e, e2e_row, "Benchmark", info.get('benchmark', ''))
            write_by_name(ws_e2e, e2e_row, "Model", info.get('model', ''))
            write_by_name(ws_e2e, e2e_row, "Phase", info.get('phase', ''))
            write_by_name(ws_e2e, e2e_row, "Dtype", info.get('dtype', ''))
            write_by_name(ws_e2e, e2e_row, "AMP", info.get('amp', False))
            write_by_name(ws_e2e, e2e_row, "Backend", info.get('backend', ''))
            write_by_name(ws_e2e, e2e_row, "Test Type", info.get('test_type', ''))
            write_by_name(ws_e2e, e2e_row, "Cudagraph", info.get('disable_cudagraphs', ''))
            # Columns 11-12: Error Message, Traceback - left blank for bug_scrub/analyze_ci_result/
            e2e_row += 1
            issues_with_e2e.add(num)
    elif test_module == 'e2e' and num not in issues_with_ut:
        # Add e2e issues without specific model info
        write_by_name(ws_e2e, e2e_row, "Issue ID", num)
        write_by_name(ws_e2e, e2e_row, "Test Reproducer", title[:150])
        write_by_name(ws_e2e, e2e_row, "Benchmark", 'unknown')
        # Columns 11-12: Error Message, Traceback - left blank for bug_scrub/analyze_ci_result/
        e2e_row += 1
        issues_with_e2e.add(num)

    if (num not in issues_with_e2e and num not in issues_with_ut
            and llm_entry and llm_entry.get('kind') == 'e2e'
            and llm_entry.get('test_cases')):
        llm_repro = llm_entry.get('reproducer') or title[:200]
        for llm_tc in llm_entry['test_cases']:
            benchmark = (llm_tc.get('test_class') or '').strip()
            model = (llm_tc.get('test_method') or '').strip()
            if not benchmark and not model:
                continue
            write_by_name(ws_e2e, e2e_row, "Issue ID", num)
            write_by_name(ws_e2e, e2e_row, "Test Reproducer", llm_repro[:200])
            write_by_name(ws_e2e, e2e_row, "Benchmark", benchmark)
            write_by_name(ws_e2e, e2e_row, "Model", model)
            e2e_row += 1
            issues_with_e2e.add(num)

    issue_row += 1

    if issue_row % 50 == 0:
        print(f"Processed {issue_row-1} issues...")

print(f"\nTotal issues: {issue_row-2}")
print(f"Total test case rows: {case_row-2}")
print(f"Total e2e case rows: {e2e_row-2}")

for issue in issues:
    num = issue['number']
    if num in issues_with_ut or num in issues_with_e2e:
        continue
    title = issue['title']
    body = issue.get('body', '') or ''
    labels = issue.get('labels', [])
    label_str = ", ".join([l['name'] for l in labels])

    write_by_name(ws_others, others_row, "ID", num)
    write_by_name(ws_others, others_row, "Title", sanitize_cell(title))
    write_by_name(ws_others, others_row, "Labels", sanitize_cell(label_str))
    # Phase 2.5 owns deep reproducer/error extraction for Others rows. Keep
    # these fields blank in Phase 1 so stale regex/LLM snippets do not drive
    # local verification.
    write_by_name(ws_others, others_row, "reproduce step", "")
    write_by_name(ws_others, others_row, "Error Message", "")
    write_by_name(ws_others, others_row, "Traceback", "")
    others_row += 1

print(f"Total others rows: {others_row-2}")

others_ids = set()
for row_idx in range(2, ws_others.max_row + 1):
    data = row_dict(ws_others, row_idx)
    if data.get("ID") is not None:
        others_ids.add(data["ID"])
for row_idx in range(2, ws_issues.max_row + 1):
    num_val = row_dict(ws_issues, row_idx).get("Issue ID")
    if num_val in others_ids:
        write_by_name(ws_issues, row_idx, "Test Module", 'others')
    elif num_val in issues_with_e2e:
        write_by_name(ws_issues, row_idx, "Test Module", 'e2e')
    elif num_val in issues_with_ut:
        write_by_name(ws_issues, row_idx, "Test Module", 'ut')

for ws in [ws_issues, ws_cases, ws_e2e, ws_others]:
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

# Ensure result directory exists
os.makedirs(RESULT_DIR, exist_ok=True)

output_path = os.path.join(RESULT_DIR, "torch_xpu_ops_issues.xlsx")


def _merge_incremental_from_prior(new_wb, prior_xlsx_path):
    """Preserve downstream analysis columns + 'Not applicable' sheet when prior xlsx exists.

    Phase 1.1 rebuilds the workbook fresh from GitHub. Without this merge, Phase 2/3/4
    output columns (Category, Root Cause, Fix Approach, action_TBD, action_reason,
    owner_transferred, AR, duplicated_issue, XPU Status, Stock Status, XPU Case Exist,
    case_existence_comments, Local Status, etc.) are wiped on every re-fetch.

    Row matching keys:
      Issues          -> Issue ID
      Test Cases      -> (Issue ID, Test File, Test Class, Test Case)
      E2E Test Cases  -> (Issue ID, Benchmark, Model, Phase, Dtype, Backend)
      Others          -> ID
      Not applicable  -> copied verbatim
    """
    if not os.path.exists(prior_xlsx_path):
        print("[incremental-merge] No prior xlsx; clean run, nothing to merge.")
        return 0

    try:
        prior_wb = openpyxl.load_workbook(prior_xlsx_path)
    except Exception as e:
        print(f"[incremental-merge] Could not load prior xlsx ({e}); skipping merge.")
        return 0

    merged_cells = 0

    canonical_headers = {
        "Issues": headers + ["Category", "Root Cause", "Fix Approach", "action_TBD", "action_reason",
                   "owner_transferred", "AR", "duplicated_issue"],
        "Test Cases": case_headers + ["Error Message", "Traceback", "XPU Status", "Stock Status",
                       "No Match Reason", "XPU Case Exist", "case_existence_comments",
                       "duplicated_issue", "Local Status"],
        "E2E Test Cases": e2e_headers + ["XPU Accuracy Status", "Local Status"],
        "Others": others_headers + ["Local Status"],
    }

    def _headers(ws):
        return list(header_index(ws))

    def _ensure_canonical(ws, sheet_name):
        for hdr in canonical_headers.get(sheet_name, []):
            ensure_col(ws, hdr)

    def _merge_sheet(sheet_name, key_fn):
        nonlocal merged_cells
        if sheet_name not in prior_wb.sheetnames or sheet_name not in new_wb.sheetnames:
            return
        prior_ws = prior_wb[sheet_name]
        new_ws = new_wb[sheet_name]
        _ensure_canonical(new_ws, sheet_name)
        prior_headers = _headers(prior_ws)
        new_headers = _headers(new_ws)
        preserved = [h for h in prior_headers if h is not None]
        for hdr in preserved:
            if hdr not in new_headers:
                ensure_col(new_ws, hdr)
        prior_index = {}
        issue_fallback = {}
        for r in range(2, prior_ws.max_row + 1):
            data = row_dict(prior_ws, r)
            try:
                k = key_fn(data)
            except Exception:
                continue
            if k is None:
                continue
            prior_index.setdefault(k, []).append(r)
            issue_key = data.get("Issue ID") if sheet_name != "Others" else data.get("ID")
            if issue_key is not None:
                issue_fallback.setdefault(issue_key, []).append(r)
        used_prior_rows = set()
        for r in range(2, new_ws.max_row + 1):
            data = row_dict(new_ws, r)
            try:
                k = key_fn(data)
            except Exception:
                continue
            if k is None:
                continue
            candidates = prior_index.get(k, [])
            chosen = next((c for c in candidates if c not in used_prior_rows), None)
            if chosen is None:
                issue_key = data.get("Issue ID") if sheet_name != "Others" else data.get("ID")
                issue_candidates = issue_fallback.get(issue_key, []) if issue_key is not None else []
                chosen = next((c for c in issue_candidates if c not in used_prior_rows), None)
            if chosen is None:
                continue
            used_prior_rows.add(chosen)
            for hdr in preserved:
                val = row_dict(prior_ws, chosen).get(hdr)
                if val not in (None, ""):
                    write_by_name(new_ws, r, hdr, val)
                    merged_cells += 1
        print(f"[incremental-merge] {sheet_name}: preserved {len(preserved)} named column(s)")

    _merge_sheet("Issues", lambda d: d.get("Issue ID"))
    _merge_sheet("Test Cases", lambda d: (d.get("Issue ID"), d.get("Test File"),
                                          d.get("Test Class"), d.get("Test Case")))
    _merge_sheet("E2E Test Cases", lambda d: (d.get("Issue ID"), d.get("Benchmark"),
                                              d.get("Model"), d.get("Phase"),
                                              d.get("Dtype"), d.get("Backend")))
    _merge_sheet("Others", lambda d: d.get("ID"))

    if "Not applicable" in prior_wb.sheetnames:
        if "Not applicable" in new_wb.sheetnames:
            del new_wb["Not applicable"]
        src_na = prior_wb["Not applicable"]
        dst_na = new_wb.create_sheet("Not applicable")
        for row in src_na.iter_rows():
            for cell in row:
                new_cell = dst_na.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy()
                    new_cell.fill = cell.fill.copy()
        print(f"[incremental-merge] Not applicable: restored {src_na.max_row - 1} row(s) verbatim")

    return merged_cells


merged = _merge_incremental_from_prior(wb, output_path)
print(f"[incremental-merge] total preserved cells: {merged}")
wb.save(output_path)
print(f"\nSaved to {output_path}")
