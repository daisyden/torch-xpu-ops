#!/usr/bin/env python3
"""
Merge one or more columns from a source Excel sheet into a target Excel sheet.

By default (--copy) values are only written where the target cell is blank.
With --update, target cells are overwritten even if they already have a value.

Adds an "Updated" column to the target sheet: "Yes" for rows that were updated.

Usage:
    python merge_reason_columns.py \
        --source <source_file.xlsx> --source-sheet <sheet_name> \
        --target <target_file.xlsx> --target-sheet <sheet_name> \
        --output <output_file.xlsx> \
        --columns V --columns W
        [--copy | --update]

Columns are specified by column letter, repeatable. To map a source column to a
differently positioned target column use "SRC:TGT", e.g.:
        --columns V:X        # source col V -> target col X
        --columns V          # source col V -> target col V

The rows are matched by the value in Column A (first column) of each sheet.
"""

import argparse
import sys

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter


def find_header_row(ws):
    """Return the 1-based row index of the header row (first non-empty row)."""
    for row in ws.iter_rows():
        if any(cell.value is not None for cell in row):
            return row[0].row
    return 1


def parse_column_specs(specs):
    """
    Parse a list of column specifications into a list of
    (source_col_index, target_col_index, source_letter, target_letter) tuples.

    Each spec is a column letter, optionally "SRC:TGT" to map a source column
    to a different target column.
    """
    parsed = []
    for spec in specs:
        spec = spec.strip()
        if not spec:
            continue
        if ":" in spec:
            src_letter, tgt_letter = (part.strip() for part in spec.split(":", 1))
        else:
            src_letter = tgt_letter = spec
        if not src_letter or not tgt_letter:
            print(f"ERROR: invalid column spec {spec!r}. Use 'V' or 'V:X'.")
            sys.exit(1)
        try:
            src_idx = column_index_from_string(src_letter.upper())
            tgt_idx = column_index_from_string(tgt_letter.upper())
        except ValueError:
            print(f"ERROR: invalid column letter in spec {spec!r}.")
            sys.exit(1)
        parsed.append((src_idx, tgt_idx, src_letter.upper(), tgt_letter.upper()))
    if not parsed:
        print("ERROR: no valid columns specified via --columns.")
        sys.exit(1)
    return parsed


def load_source_map(ws, header_row, key_col, src_cols):
    """
    Build a dict mapping key_value -> tuple of source column values
    (in the same order as src_cols), skipping the header row.
    """
    mapping = {}
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        key = row[key_col - 1] if len(row) >= key_col else None
        if key is None:
            continue
        values = tuple(
            row[c - 1] if len(row) >= c else None for c in src_cols
        )
        mapping[key] = values
    return mapping


def merge(
    source_path,
    source_sheet,
    target_path,
    target_sheet,
    output_path,
    column_specs,
    overwrite,
):
    print(f"Loading source:  {source_path!r}  sheet={source_sheet!r}")
    src_wb = openpyxl.load_workbook(source_path, data_only=True)
    if source_sheet not in src_wb.sheetnames:
        print(
            f"ERROR: sheet {source_sheet!r} not found in source. Available: {src_wb.sheetnames}"
        )
        sys.exit(1)
    src_ws = src_wb[source_sheet]

    print(f"Loading target:  {target_path!r}  sheet={target_sheet!r}")
    tgt_wb = openpyxl.load_workbook(target_path)
    if target_sheet not in tgt_wb.sheetnames:
        print(
            f"ERROR: sheet {target_sheet!r} not found in target. Available: {tgt_wb.sheetnames}"
        )
        sys.exit(1)
    tgt_ws = tgt_wb[target_sheet]

    src_cols = [s[0] for s in column_specs]
    tgt_cols = [s[1] for s in column_specs]

    mode = "update (overwrite)" if overwrite else "copy (blank only)"
    mapping_desc = ", ".join(
        f"{src_l}->{tgt_l}" for _, _, src_l, tgt_l in column_specs
    )
    print(f"Mode: {mode}")
    print(f"Columns: {mapping_desc}")

    # --- Locate header and key column in source ---
    src_header_row = find_header_row(src_ws)
    src_key_col = 1  # Column A
    print(f"Source header row={src_header_row}, key=col {src_key_col}(A)")

    # --- Build source lookup ---
    source_map = load_source_map(src_ws, src_header_row, src_key_col, src_cols)
    print(f"Source rows loaded: {len(source_map)}")

    # --- Locate header in target ---
    tgt_header_row = find_header_row(tgt_ws)
    tgt_key_col = 1
    print(f"Target header row={tgt_header_row}, key=col {tgt_key_col}(A)")

    # --- Add "Updated" column header to target (next free column) ---
    max_col = tgt_ws.max_column
    updated_col = max_col + 1
    tgt_ws.cell(row=tgt_header_row, column=updated_col, value="Updated")
    print(
        f"'Updated' column added at col {updated_col}({get_column_letter(updated_col)})"
    )

    # --- Iterate target rows and fill in / overwrite the configured columns ---
    updated_count = 0
    skipped_no_match = 0
    skipped_already_filled = 0

    for row_idx in range(tgt_header_row + 1, tgt_ws.max_row + 1):
        key_cell = tgt_ws.cell(row=row_idx, column=tgt_key_col)
        key = key_cell.value
        if key is None:
            continue

        if key not in source_map:
            skipped_no_match += 1
            continue

        src_values = source_map[key]

        # Skip rows where source has nothing to contribute at all.
        if all(v is None for v in src_values):
            skipped_no_match += 1
            continue

        row_changed = False
        any_target_blocked = False

        for (_, tgt_col, _, _), src_value in zip(column_specs, src_values):
            cell = tgt_ws.cell(row=row_idx, column=tgt_col)
            cell_blank = cell.value is None or str(cell.value).strip() == ""

            if not overwrite and not cell_blank:
                # In copy mode, never touch a non-blank target cell.
                any_target_blocked = True
                continue

            if src_value is None:
                # Nothing to write from source for this column.
                continue

            if cell.value != src_value:
                cell.value = src_value
                row_changed = True

        if row_changed:
            tgt_ws.cell(row=row_idx, column=updated_col, value="Yes")
            updated_count += 1
        elif any_target_blocked:
            skipped_already_filled += 1

    print(f"\nSummary:")
    print(f"  Rows updated:              {updated_count}")
    print(f"  Rows skipped (no match):   {skipped_no_match}")
    print(f"  Rows skipped (not blank):  {skipped_already_filled}")

    tgt_wb.save(output_path)
    print(f"\nSaved output to: {output_path!r}")


def main():
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--source", required=True, help="Path to source Excel file")
    parser.add_argument(
        "--source-sheet", required=True, help="Sheet name in source file"
    )
    parser.add_argument("--target", required=True, help="Path to target Excel file")
    parser.add_argument(
        "--target-sheet", required=True, help="Sheet name in target file"
    )
    parser.add_argument(
        "--output",
        required=True,
        help="Path to save the updated target file (can be same as --target to overwrite)",
    )
    parser.add_argument(
        "--columns",
        action="append",
        default=None,
        metavar="SPEC",
        help=(
            "Column to copy/update by letter, repeatable. "
            "Use 'V' for same source/target column, or 'V:X' to map "
            "source column V to target column X. "
            "Defaults to 'V' and 'W' (Reason / DetailReason) if omitted."
        ),
    )
    mode_group = parser.add_mutually_exclusive_group()
    mode_group.add_argument(
        "--copy",
        dest="overwrite",
        action="store_false",
        help="Only fill target cells that are blank (default).",
    )
    mode_group.add_argument(
        "--update",
        dest="overwrite",
        action="store_true",
        help="Overwrite target cells even if they already have a value.",
    )
    parser.set_defaults(overwrite=False)
    args = parser.parse_args()

    specs = args.columns if args.columns else ["V", "W"]
    column_specs = parse_column_specs(specs)

    merge(
        source_path=args.source,
        source_sheet=args.source_sheet,
        target_path=args.target,
        target_sheet=args.target_sheet,
        output_path=args.output,
        column_specs=column_specs,
        overwrite=args.overwrite,
    )


if __name__ == "__main__":
    main()
